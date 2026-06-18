import csv
import os
from io import BytesIO, StringIO

from django.conf import settings

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from rest_framework import viewsets, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination

from .models import (
    InstallationProject,
    InstallationProduct,
    InstallationTask,
    InstallationDocument,
    InstallationReport,
    InstallationNotification,
)
from .serializers import (
    InstallationProjectSerializer,
    InstallationProductSerializer,
    InstallationTaskSerializer,
    InstallationDocumentSerializer,
    InstallationReportSerializer,
    InstallationNotificationSerializer,
)


class InstallationPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 200


class BaseInstallationViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = InstallationPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering = ['-id']


def _project_filter(qs, request):
    project = request.query_params.get('project')
    if project:
        qs = qs.filter(project_id=project)
    return qs


class InstallationProjectViewSet(BaseInstallationViewSet):
    queryset = InstallationProject.objects.all().order_by('-id')
    serializer_class = InstallationProjectSerializer
    search_fields = ['name', 'client', 'status', 'supervisor']

    @action(detail=True, methods=['get'])
    def tasks(self, request, pk=None):
        qs = InstallationTask.objects.filter(project_id=pk).order_by('due_date', '-id')
        status = request.query_params.get('status')
        priority = request.query_params.get('priority')
        if status:
            qs = qs.filter(status=status)
        if priority:
            qs = qs.filter(priority=priority)
        return Response(InstallationTaskSerializer(qs, many=True, context={'request': request}).data)

    @action(detail=True, methods=['get'])
    def products(self, request, pk=None):
        qs = InstallationProduct.objects.filter(project_id=pk).order_by('-id')
        return Response(InstallationProductSerializer(qs, many=True, context={'request': request}).data)


class InstallationProductViewSet(BaseInstallationViewSet):
    queryset = InstallationProduct.objects.select_related('project').all().order_by('-id')
    serializer_class = InstallationProductSerializer
    search_fields = ['reference', 'name', 'description', 'status', 'project__name']

    def get_queryset(self):
        return _project_filter(super().get_queryset(), self.request)


class InstallationTaskViewSet(BaseInstallationViewSet):
    queryset = InstallationTask.objects.select_related('project').all().order_by('due_date', '-id')
    serializer_class = InstallationTaskSerializer
    search_fields = ['name', 'description', 'status', 'priority', 'comment', 'project__name']

    def get_queryset(self):
        qs = _project_filter(super().get_queryset(), self.request)
        status = self.request.query_params.get('status')
        priority = self.request.query_params.get('priority')
        if status:
            qs = qs.filter(status=status)
        if priority:
            qs = qs.filter(priority=priority)
        return qs


class InstallationDocumentViewSet(BaseInstallationViewSet):
    queryset = InstallationDocument.objects.select_related('project', 'uploaded_by').all().order_by('-id')
    serializer_class = InstallationDocumentSerializer
    search_fields = ['title', 'document_type', 'status', 'project__name']

    def get_queryset(self):
        return _project_filter(super().get_queryset(), self.request)

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)


class InstallationReportViewSet(BaseInstallationViewSet):
    queryset = InstallationReport.objects.select_related('project', 'generated_by').all().order_by('-id')
    serializer_class = InstallationReportSerializer
    search_fields = ['reference', 'title', 'report_type', 'status', 'project__name']

    def get_queryset(self):
        return _project_filter(super().get_queryset(), self.request)

    def perform_create(self, serializer):
        serializer.save(generated_by=self.request.user)


class InstallationNotificationViewSet(BaseInstallationViewSet):
    queryset = InstallationNotification.objects.select_related('project').all().order_by('-id')
    serializer_class = InstallationNotificationSerializer
    search_fields = ['title', 'message', 'level', 'project__name']

    def get_queryset(self):
        return _project_filter(super().get_queryset(), self.request)


class InstallationDashboardViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        today = timezone.localdate()
        open_tasks = InstallationTask.objects.exclude(status='Terminée')
        return Response({
            'projects': InstallationProject.objects.count(),
            'activeProjects': InstallationProject.objects.filter(status='En cours').count(),
            'finishedProjects': InstallationProject.objects.filter(status='Terminé').count(),
            'lateProjects': InstallationProject.objects.exclude(status__in=['Terminé', 'Annulé']).filter(planned_end_date__lt=today).count(),
            'products': InstallationProduct.objects.count(),
            'tasks': InstallationTask.objects.count(),
            'openTasks': open_tasks.count(),
            'finishedTasks': InstallationTask.objects.filter(status='Terminée').count(),
            'lateTasks': open_tasks.filter(due_date__lt=today).count(),
            'documents': InstallationDocument.objects.count(),
            'reports': InstallationReport.objects.count(),
            'notifications': InstallationNotification.objects.filter(is_read=False).count(),
        })


def _rows_for(resource, request):
    if resource == 'projects':
        qs = InstallationProject.objects.all().order_by('-id')
        headers = ['ID', 'Projet', 'Client', 'Début', 'Fin prévue', 'Statut', 'Avancement %', 'Produits', 'Tâches']
        rows = [[o.id, o.name, o.client, o.start_date or '', o.planned_end_date or '', o.status, o.progress, o.products_count, o.tasks_count] for o in qs]
    elif resource == 'products':
        qs = _project_filter(InstallationProduct.objects.select_related('project').all().order_by('-id'), request)
        headers = ['ID', 'Référence', 'Produit', 'Projet', 'Date', 'Statut']
        rows = [[o.id, o.reference, o.name, o.project.name, o.date or '', o.status] for o in qs]
    elif resource == 'tasks':
        qs = _project_filter(InstallationTask.objects.select_related('project').all().order_by('due_date', '-id'), request)
        headers = ['ID', 'Tâche', 'Projet', 'Statut', 'Priorité', 'Début', 'Échéance', 'Commentaire']
        rows = [[o.id, o.name, o.project.name, o.status, o.priority, o.start_date or '', o.due_date or '', o.comment] for o in qs]
    elif resource == 'documents':
        qs = _project_filter(InstallationDocument.objects.select_related('project').all().order_by('-id'), request)
        headers = ['ID', 'Titre', 'Type', 'Projet', 'Statut']
        rows = [[o.id, o.title, o.document_type, o.project.name if o.project else '', o.status] for o in qs]
    else:
        qs = _project_filter(InstallationReport.objects.select_related('project').all().order_by('-id'), request)
        headers = ['ID', 'Référence', 'Titre', 'Type', 'Projet', 'Statut']
        rows = [[o.id, o.reference, o.title, o.report_type, o.project.name if o.project else '', o.status] for o in qs]
    return headers, rows



def _logo_path():
    candidates = [
        os.path.join(settings.BASE_DIR, 'static', 'hacint_logo.png'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'hacint_logo.png'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _clean(value):
    return str(value if value is not None else '').replace('\n', ' ').strip()


def _report_title(resource):
    names = {
        'projects': 'RAPPORT INSTALLATION - PROJETS',
        'products': 'RAPPORT INSTALLATION - PRODUITS',
        'tasks': 'RAPPORT INSTALLATION - TÂCHES',
        'documents': 'RAPPORT INSTALLATION - DOCUMENTS',
        'reports': 'RAPPORT INSTALLATION - RAPPORTS',
    }
    return names.get(resource, 'RAPPORT INSTALLATION')


def _simple_pdf(title, headers, rows):
    # Fallback PDF minimal si ReportLab n'est pas disponible.
    lines = ['HACINT ERP', title, f'Date de génération: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}', ''] + [' | '.join(map(_clean, headers))]
    for row in rows[:200]:
        lines.append(' | '.join(_clean(x) for x in row))
    text = '\n'.join(lines).replace('(', '[').replace(')', ']')
    stream = 'BT /F1 9 Tf 35 805 Td 12 TL ' + ''.join(f'({_clean(line)[:145]}) Tj T* ' for line in text.split('\n')) + 'ET'
    objects = []
    objects.append('1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj')
    objects.append('2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj')
    objects.append('3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj')
    objects.append('4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj')
    objects.append(f'5 0 obj << /Length {len(stream.encode("latin-1", "ignore"))} >> stream\n{stream}\nendstream endobj')
    pdf = '%PDF-1.4\n'
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf.encode('latin-1')))
        pdf += obj + '\n'
    xref_pos = len(pdf.encode('latin-1'))
    pdf += f'xref\n0 {len(objects)+1}\n0000000000 65535 f \n'
    for off in offsets[1:]:
        pdf += f'{off:010d} 00000 n \n'
    pdf += f'trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF'
    return pdf.encode('latin-1', 'ignore')


def _professional_pdf(title, headers, rows):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    except Exception:
        return _simple_pdf(title, headers, rows)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    logo = _logo_path()
    header_table_data = []
    if logo:
        header_table_data.append([
            RLImage(logo, width=42 * mm, height=12 * mm),
            Paragraph('<b>HACINT ERP</b><br/><font size="9">Document généré automatiquement</font>', styles['Normal']),
            Paragraph(f'<b>{title}</b><br/><font size="9">Date de génération : {timezone.localtime().strftime("%d/%m/%Y %H:%M")}</font>', styles['Normal']),
        ])
    else:
        header_table_data.append([
            Paragraph('<b>HACINT ERP</b>', styles['Title']),
            '',
            Paragraph(f'<b>{title}</b><br/><font size="9">Date de génération : {timezone.localtime().strftime("%d/%m/%Y %H:%M")}</font>', styles['Normal']),
        ])
    header_table = Table(header_table_data, colWidths=[55 * mm, 75 * mm, 135 * mm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#dbe4f0')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fbff')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8 * mm))

    data = [headers] + [[_clean(c) for c in row] for row in rows[:250]]
    if len(data) == 1:
        data.append(['Aucune donnée disponible'] + [''] * (len(headers) - 1))

    usable_width = landscape(A4)[0] - 28 * mm
    col_width = usable_width / max(len(headers), 1)
    table = Table(data, repeatRows=1, colWidths=[col_width] * len(headers))
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#243a8f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#dbe4f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fbff')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    story.append(Spacer(1, 7 * mm))
    story.append(Paragraph('<font size="8">HACINT ERP — Rapport exporté en français. Signature électronique non requise.</font>', styles['Normal']))
    doc.build(story)
    return buffer.getvalue()


def _professional_xlsx(resource, title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rapport Installation'

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=max(len(headers), 6))
    ws['A1'] = f'HACINT ERP - {title}'
    ws['A1'].font = Font(bold=True, size=16, color='243A8F')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    logo = _logo_path()
    if logo:
        try:
            img = ExcelImage(logo)
            img.width = 160
            img.height = 45
            ws.add_image(img, 'A1')
        except Exception:
            pass

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(headers), 6))
    ws['A3'] = f'Date de génération : {timezone.localtime().strftime("%d/%m/%Y %H:%M")}  |  Module : Installation  |  Export : {resource}'
    ws['A3'].font = Font(size=10, color='64748B')
    ws['A3'].alignment = Alignment(horizontal='center')

    start_row = 5
    blue = PatternFill('solid', fgColor='243A8F')
    light = PatternFill('solid', fgColor='F8FBFF')
    white_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='DBE4F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = blue
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for r, row in enumerate(rows, start_row + 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=_clean(value))
            cell.border = border
            if r % 2 == 0:
                cell.fill = light
            cell.alignment = Alignment(vertical='top')

    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=start_row, column=col_idx).column_letter
        max_len = max([len(_clean(ws.cell(row=r, column=col_idx).value)) for r in range(start_row, ws.max_row + 1)] + [12])
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 35)

    ws.freeze_panes = f'A{start_row + 1}'
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_installation(request, resource, fmt):
    if resource not in ['projects', 'products', 'tasks', 'documents', 'reports']:
        return Response({'detail': 'Ressource inconnue'}, status=404)

    headers, rows = _rows_for(resource, request)
    filename = f'hacint_installation_{resource}.{fmt}'
    title = _report_title(resource)

    if fmt == 'csv':
        output = StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(['HACINT ERP', title, f'Date de génération: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}'])
        writer.writerow([])
        writer.writerow(headers)
        writer.writerows(rows)
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    elif fmt == 'xlsx':
        response = HttpResponse(_professional_xlsx(resource, title, headers, rows), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        response = HttpResponse(_professional_pdf(title, headers, rows), content_type='application/pdf')
    else:
        return Response({'detail': 'Format inconnu'}, status=404)

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
