import csv
import os
import shutil
import time
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
import csv as csv_module
import io

from rest_framework import filters, status, viewsets
from rest_framework.parsers import MultiPartParser
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from rest_framework.decorators import api_view, permission_classes as drf_permission_classes
from rest_framework.permissions import BasePermission, IsAdminUser
from accounting.models import Tiers


class IsTechStudyOrAdmin(BasePermission):
    """Allow access to admin (staff/superuser) or users in the 'Etude Technique' group."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return (
            request.user.is_staff or
            request.user.is_superuser or
            request.user.groups.filter(name='Etude Technique').exists()
        )

class IsSalesOrAdmin(BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return (
            request.user.is_staff
            or request.user.is_superuser
            or request.user.groups.filter(name__in=['Sales Manager', 'Sales Employee']).exists()
        )


from .models import (
    CLIENT_CHOICES, CONNECTOR_FILL_CHOICES, PLACEMENT_RE, STATUS_CHOICES,
    AuditLog, BomItem, JimideDxfFile, MatrixEntry, ProjectValidation, ProgrammerFile, Sample,
    ProjectDocument, SalesRecord, SalesAuditLog, SalesProjectHistory, SalesTarget,
)
from .serializers import (
    AuditLogSerializer, BomItemSerializer, MatrixEntrySerializer, ProjectValidationSerializer,
    SampleDetailSerializer, SampleListSerializer, SalesRecordSerializer,
)


class SamplePagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 200


def _parse_bom_excel(file, sample):
    """
    Parse an Excel BOM file, wipe existing BomItems for the sample, and
    re-create them from the sheet.  Returns (count_imported, error_message).
    """
    import openpyxl

    REF_KEYS  = {'reference', 'ref', 'code', 'p/n', 'part', 'part number', 'part_number',
                 'num', 'numéro', 'repère', 'repere', 'référence'}
    DES_KEYS  = {'designation', 'description', 'name', 'désignation', 'designation',
                 'libellé', 'libelle', 'nom', 'article', 'desc'}
    QTY_KEYS  = {'quantity', 'qty', 'quantité', 'quantite', 'qté', 'qte',
                 'nbr', 'nombre', 'nbre', 'q', 'qté.'}
    UNIT_KEYS = {'unit', 'unité', 'unite', 'uom', 'u', 'unités', 'u.', 'um', 'unité'}
    UNIT_MAP  = {
        'pcs': 'pcs', 'pièces': 'pcs', 'pieces': 'pcs', 'piece': 'pcs', 'pc': 'pcs',
        'un': 'pcs', 'u': 'pcs', 'ea': 'pcs',
        'm': 'm', 'mètre': 'm', 'metre': 'm', 'mètres': 'm', 'metres': 'm',
        'm2': 'm2', 'm²': 'm2', 'mètre²': 'm2', 'mètre carré': 'm2',
        'kg': 'kg', 'kilogramme': 'kg', 'kilogrammes': 'kg',
        'g': 'g', 'gramme': 'g', 'grammes': 'g',
        'l': 'l', 'litre': 'l', 'litres': 'l',
        'mm': 'mm', 'millimètre': 'mm', 'millimètres': 'mm', 'millimetre': 'mm',
    }

    try:
        file.seek(0)
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return 0, f'Impossible de lire le fichier Excel : {exc}'

    if not rows:
        return 0, 'Le fichier Excel est vide.'

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    ref_col = des_col = qty_col = unit_col = None
    for i, h in enumerate(headers):
        if h in REF_KEYS  and ref_col  is None: ref_col  = i
        if h in DES_KEYS  and des_col  is None: des_col  = i
        if h in QTY_KEYS  and qty_col  is None: qty_col  = i
        if h in UNIT_KEYS and unit_col is None: unit_col = i

    if ref_col is None:
        return 0, (
            'Colonne "Reference" introuvable. '
            'Nommez la première colonne : Reference (ou Ref, Code, P/N).'
        )

    BomItem.objects.filter(sample=sample).delete()

    count = 0
    for row in rows[1:]:
        raw_ref = row[ref_col] if ref_col is not None else None
        if raw_ref is None:
            continue
        ref = str(raw_ref).strip()
        if not ref or ref.lower() == 'none':
            continue

        des = ''
        if des_col is not None and row[des_col] is not None:
            des = str(row[des_col]).strip()

        qty = 1.0
        if qty_col is not None and row[qty_col] is not None:
            try:
                qty = float(row[qty_col])
            except (TypeError, ValueError):
                qty = 1.0

        unit = 'pcs'
        if unit_col is not None and row[unit_col] is not None:
            unit = UNIT_MAP.get(str(row[unit_col]).strip().lower(), 'pcs')

        BomItem.objects.create(
            sample=sample, reference=ref[:100],
            designation=des[:200], quantity=qty, unit=unit,
        )
        count += 1

    return count, None


class SampleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = SamplePagination
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['created_at', 'received_date', 'apn', 'project', 'client', 'status']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = Sample.objects.select_related('created_by', 'updated_by')

        # Full-text search (includes serial number)
        search = self.request.query_params.get('search', '').strip()
        if search:
            q = (
                Q(apn__icontains=search) |
                Q(project__icontains=search) |
                Q(placement__icontains=search) |
                Q(client__icontains=search) |
                Q(description__icontains=search)
            )
            try:
                sn = int(search.lstrip('#'))
                q |= Q(serial_number=sn)
            except ValueError:
                pass
            qs = qs.filter(q)

        client = self.request.query_params.get('client')
        if client:
            qs = qs.filter(client=client)

        project = self.request.query_params.get('project')
        if project:
            qs = qs.filter(project=project)

        approved_only = self.request.query_params.get('approved_only')
        if approved_only == 'true':
            # Only admin-approved projects (approved_at set), regardless of validation result
            approved_names = list(
                ProjectValidation.objects.filter(approved_at__isnull=False)
                .values_list('project_name', flat=True)
            )
            # Only show samples explicitly approved within their project
            qs = qs.filter(project__in=approved_names, study_approved=True)

        status_param = self.request.query_params.get('status')
        if status_param:
            qs = qs.filter(status=status_param)

        date_from = self.request.query_params.get('date_from')
        if date_from:
            qs = qs.filter(received_date__gte=date_from)

        date_to = self.request.query_params.get('date_to')
        if date_to:
            qs = qs.filter(received_date__lte=date_to)

        # Designer filter: ?is_done=true|false
        is_done_param = self.request.query_params.get('is_done')
        if is_done_param is not None:
            qs = qs.filter(is_done=(is_done_param.lower() in ('true', '1', 'yes')))

        # Programmer view: show designer-done + rework items, rework pinned first
        programmer_view = self.request.query_params.get('programmer_view')
        if programmer_view:
            from django.db.models import Q as _Q
            qs = qs.filter(_Q(is_done=True) | _Q(is_rework=True))
            qs = qs.order_by('-is_rework', '-is_cnc_rework', '-created_at')

        # Programmer done filter: ?programmer_done=true|false
        prog_done_param = self.request.query_params.get('programmer_done')
        if prog_done_param is not None:
            qs = qs.filter(programmer_done=(prog_done_param.lower() in ('true', '1', 'yes')))

        # CNC view: show programmer-done + cnc rework items, rework pinned first
        cnc_view = self.request.query_params.get('cnc_view')
        if cnc_view:
            from django.db.models import Q as _Q
            qs = qs.filter(_Q(programmer_done=True) | _Q(is_cnc_rework=True))
            qs = qs.order_by('-is_cnc_rework', '-is_assembly_rework', '-created_at')

        # CNC done filter: ?cnc_done=true|false
        cnc_done_param = self.request.query_params.get('cnc_done')
        if cnc_done_param is not None:
            qs = qs.filter(cnc_done=(cnc_done_param.lower() in ('true', '1', 'yes')))

        # Assembly view: show cnc-started (count>=1) + cnc-done + assembly/quality rework
        assembly_view = self.request.query_params.get('assembly_view')
        if assembly_view:
            from django.db.models import Q as _Q
            qs = qs.filter(
                _Q(cnc_done=True) | _Q(cnc_produced_count__gte=1) |
                _Q(is_assembly_rework=True) | _Q(is_quality_rework=True)
            )
            qs = qs.order_by('-is_assembly_rework', '-is_quality_rework', '-created_at')

        # Assembly done filter: ?assembly_done=true|false
        assembly_done_param = self.request.query_params.get('assembly_done')
        if assembly_done_param is not None:
            qs = qs.filter(assembly_done=(assembly_done_param.lower() in ('true', '1', 'yes')))

        # Quality view: show assembly-done + quality rework, rework pinned first
        quality_view = self.request.query_params.get('quality_view')
        if quality_view:
            from django.db.models import Q as _Q
            qs = qs.filter(_Q(assembly_done=True) | _Q(is_quality_rework=True))
            qs = qs.order_by('-is_quality_rework', '-created_at')

        # Quality done filter: ?quality_done=true|false
        quality_done_param = self.request.query_params.get('quality_done')
        if quality_done_param is not None:
            qs = qs.filter(quality_done=(quality_done_param.lower() in ('true', '1', 'yes')))

        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return SampleListSerializer
        return SampleDetailSerializer

    def perform_create(self, serializer):
        sample = serializer.save(created_by=self.request.user)
        AuditLog.objects.create(sample=sample, user=self.request.user, action='create')

    def perform_update(self, serializer):
        sample = serializer.save(updated_by=self.request.user)
        AuditLog.objects.create(sample=sample, user=self.request.user, action='update')

    def destroy(self, request, *args, **kwargs):
        sample = self.get_object()
        sample.soft_delete(user=request.user)
        AuditLog.objects.create(sample=sample, user=request.user, action='delete')
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='export')
    def export_csv(self, request):
        qs = self.get_queryset()

        # Apply search same as list (includes serial number)
        search = request.query_params.get('search', '').strip()
        if search:
            q = (
                Q(apn__icontains=search) |
                Q(project__icontains=search) |
                Q(placement__icontains=search) |
                Q(client__icontains=search) |
                Q(description__icontains=search)
            )
            try:
                sn = int(search.lstrip('#'))
                q |= Q(serial_number=sn)
            except ValueError:
                pass
            qs = qs.filter(q)

        filename = f"samples_{date.today().strftime('%Y-%m-%d')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            'apn', 'project', 'placement', 'received_date',
            'client', 'status', 'quantity', 'connector_fill',
            'description', 'image_filename',
        ])

        for s in qs:
            writer.writerow([
                s.apn,
                s.project,
                s.placement,
                s.received_date.strftime('%Y-%m-%d') if s.received_date else '',
                s.client,
                s.status,
                s.quantity,
                s.connector_fill,
                s.description,
                os.path.basename(s.image.name) if s.image else '',
            ])

        AuditLog.objects.create(
            user=request.user,
            action='export',
            changes={'count': qs.count()},
        )
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_csv(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)

        import pandas as pd

        file = request.FILES['file']
        try:
            df = pd.read_csv(file, encoding='utf-8-sig')
        except Exception:
            try:
                file.seek(0)
                df = pd.read_csv(file, encoding='latin-1')
            except Exception as e:
                return Response(
                    {'error': f'Impossible de lire le fichier CSV: {e}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        required_cols = {'apn', 'project', 'placement', 'client'}
        missing = required_cols - set(df.columns)
        if missing:
            return Response(
                {'error': f'Colonnes manquantes: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        valid_clients = {c[0] for c in CLIENT_CHOICES}
        valid_statuses = {s[0] for s in STATUS_CHOICES}
        valid_fills = {f[0] for f in CONNECTOR_FILL_CHOICES}
        success_count = 0
        errors = []

        for idx, row in df.iterrows():
            row_num = int(idx) + 2

            apn = str(row.get('apn', '')).strip()
            project = str(row.get('project', '')).strip()
            placement = str(row.get('placement', '')).strip()
            client = str(row.get('client', '')).strip()

            if not all([apn, project, placement, client]) or any(
                v == 'nan' for v in [apn, project, placement, client]
            ):
                errors.append({'row': row_num, 'message': 'Champs obligatoires manquants (apn, project, placement, client).'})
                continue

            if not PLACEMENT_RE.match(placement):
                errors.append({'row': row_num, 'message': f'Format de placement invalide: "{placement}". Attendu: lettre + 1-2 chiffres.'})
                continue

            if client not in valid_clients:
                errors.append({'row': row_num, 'message': f'Client invalide: "{client}". Valeurs acceptées: {", ".join(sorted(valid_clients))}.'})
                continue

            if Sample.objects.filter(apn=apn, project=project).exists():
                errors.append({'row': row_num, 'message': f'Doublon: APN "{apn}" + Projet "{project}" existe déjà.'})
                continue

            raw_status = str(row.get('status', 'pending')).strip()
            sample_status = raw_status if raw_status in valid_statuses else 'pending'

            received_date = timezone.now().date()
            raw_date = str(row.get('received_date', '')).strip()
            if raw_date and raw_date != 'nan':
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        received_date = datetime.strptime(raw_date, fmt).date()
                        break
                    except ValueError:
                        continue

            description = str(row.get('description', '')).strip()
            if description == 'nan':
                description = ''

            # quantity
            raw_qty = str(row.get('quantity', '1')).strip()
            try:
                quantity = max(1, int(float(raw_qty))) if raw_qty and raw_qty != 'nan' else 1
            except (ValueError, TypeError):
                quantity = 1

            # connector_fill
            raw_fill = str(row.get('connector_fill', 'empty')).strip().lower()
            connector_fill = raw_fill if raw_fill in valid_fills else 'empty'

            Sample.objects.create(
                apn=apn,
                project=project,
                placement=placement,
                client=client,
                status=sample_status,
                quantity=quantity,
                connector_fill=connector_fill,
                received_date=received_date,
                description=description,
                created_by=request.user,
            )
            success_count += 1

        AuditLog.objects.create(
            user=request.user,
            action='import',
            changes={'success': success_count, 'errors': len(errors)},
        )

        return Response({'success': success_count, 'errors': errors, 'total': len(df)})

    @action(detail=False, methods=['get'], url_path='projects')
    def list_projects(self, request):
        """Return sorted unique approved project names for the filter dropdown."""
        approved_names = list(
            ProjectValidation.objects.filter(approved_at__isnull=False)
            .values_list('project_name', flat=True)
        )
        projects = (
            Sample.objects
            .filter(project__in=approved_names)
            .values_list('project', flat=True)
            .distinct()
            .order_by('project')
        )
        return Response(list(projects))

    @action(detail=False, methods=['get'], url_path='search')
    def search_samples(self, request):
        q = request.query_params.get('q', '').strip()
        if not q:
            return Response([])

        qs = Sample.objects.filter(
            Q(apn__icontains=q) |
            Q(project__icontains=q) |
            Q(placement__icontains=q) |
            Q(client__icontains=q)
        )[:20]

        serializer = SampleListSerializer(qs, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='mark_done')
    def mark_done(self, request, pk=None):
        """Designer endpoint: mark a sample as done (or undo it)."""
        sample = self.get_object()
        is_done = str(request.data.get('is_done', 'true')).lower() in ('true', '1', 'yes')

        sample.is_done    = is_done
        sample.done_date  = request.data.get('done_date') or (timezone.now().date() if is_done else None)
        sample.done_by    = request.user if is_done else None
        sample.updated_by = request.user
        # When undoing, fully reset designer-status so timer stops and counters are correct
        if not is_done:
            sample.designer_status     = None
            sample.time_started        = None
            sample.time_spent_minutes  = 0
            sample.designer_locked_by  = None   # release lock on undo
        sample.save(update_fields=[
            'is_done', 'done_date', 'done_by', 'updated_by', 'updated_at',
            'designer_status', 'time_started', 'time_spent_minutes', 'designer_locked_by',
        ])

        AuditLog.objects.create(
            sample=sample,
            user=request.user,
            action='update',
            changes={
                'is_done': is_done,
                'done_date': str(sample.done_date) if sample.done_date else None,
            },
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_designer_status')
    def set_designer_status(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_designer_status/
        body: { "designer_status": "ongoing" | "standby" | "done" }

        - ongoing  → record time_started = now (start / resume timer)
        - standby  → accumulate delta into time_spent_minutes, clear time_started
        - done     → finalize time, set is_done = True, done_date = today
        """
        sample = self.get_object()
        new_status = request.data.get('designer_status', '').lower()
        valid = {'ongoing', 'standby', 'done'}
        if new_status not in valid:
            return Response(
                {'error': f'designer_status must be one of {sorted(valid)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Lock check ───────────────────────────────────────────────────────
        if (sample.designer_locked_by_id
                and sample.designer_locked_by != request.user
                and sample.designer_status in ('ongoing', 'standby')):
            name = (sample.designer_locked_by.get_full_name()
                    or sample.designer_locked_by.username)
            return Response(
                {'error': f'Verrouillé par {name}.', 'locked_by': name},
                status=status.HTTP_403_FORBIDDEN,
            )

        now = timezone.now()

        if new_status == 'ongoing':
            if sample.designer_status != 'ongoing':
                sample.time_started = now
            sample.pause_reason      = None
            sample.designer_locked_by = request.user   # acquire / keep lock

        elif new_status == 'standby':
            if sample.designer_status == 'ongoing' and sample.time_started:
                delta = now - sample.time_started
                sample.time_spent_minutes += int(delta.total_seconds() // 60)
            sample.time_started       = None
            sample.pause_reason       = request.data.get('pause_reason') or None
            sample.designer_locked_by = request.user   # keep lock while paused

        elif new_status == 'done':
            if sample.designer_status == 'ongoing' and sample.time_started:
                delta = now - sample.time_started
                sample.time_spent_minutes += int(delta.total_seconds() // 60)
            sample.time_started       = None
            sample.pause_reason       = None
            sample.is_done            = True
            sample.done_date          = now.date()
            sample.done_by            = request.user
            sample.is_rework          = False
            sample.designer_locked_by = None            # release lock on completion

        sample.designer_status = new_status
        sample.updated_by = request.user
        save_fields = [
            'designer_status', 'time_started', 'time_spent_minutes',
            'pause_reason', 'designer_locked_by',
            'is_done', 'done_date', 'done_by',
            'is_rework', 'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample,
            user=request.user,
            action='update',
            changes={
                'designer_status': new_status,
                'time_spent_minutes': sample.time_spent_minutes,
            },
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_programmer_status')
    def set_programmer_status(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_programmer_status/
        body: { "programmer_status": "ongoing"|"standby"|"done", "pause_reason": "..." }
        """
        sample = self.get_object()
        new_status = request.data.get('programmer_status', '').lower()
        valid = {'ongoing', 'standby', 'done'}
        if new_status not in valid:
            return Response(
                {'error': f'programmer_status must be one of {sorted(valid)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Lock check ───────────────────────────────────────────────────────
        if (sample.programmer_locked_by_id
                and sample.programmer_locked_by != request.user
                and sample.programmer_status in ('ongoing', 'standby')):
            name = (sample.programmer_locked_by.get_full_name()
                    or sample.programmer_locked_by.username)
            return Response(
                {'error': f'Verrouillé par {name}.', 'locked_by': name},
                status=status.HTTP_403_FORBIDDEN,
            )

        now = timezone.now()

        if new_status == 'ongoing':
            if sample.programmer_status != 'ongoing':
                sample.programmer_time_started = now
            sample.programmer_pause_reason  = None
            sample.programmer_locked_by     = request.user

        elif new_status == 'standby':
            if sample.programmer_status == 'ongoing' and sample.programmer_time_started:
                delta = now - sample.programmer_time_started
                sample.programmer_time_spent_minutes += int(delta.total_seconds() // 60)
            sample.programmer_time_started  = None
            sample.programmer_pause_reason  = request.data.get('pause_reason') or None
            sample.programmer_locked_by     = request.user

        elif new_status == 'done':
            if sample.programmer_status == 'ongoing' and sample.programmer_time_started:
                delta = now - sample.programmer_time_started
                sample.programmer_time_spent_minutes += int(delta.total_seconds() // 60)
            sample.programmer_time_started  = None
            sample.programmer_pause_reason  = None
            sample.programmer_done          = True
            sample.programmer_done_date     = now.date()
            sample.programmer_done_by       = request.user
            sample.programmer_locked_by     = None
            sample.is_cnc_rework            = False   # clear CNC rework flag — ready for CNC

        sample.programmer_status = new_status
        sample.updated_by = request.user
        save_fields = [
            'programmer_status', 'programmer_time_started', 'programmer_time_spent_minutes',
            'programmer_pause_reason', 'programmer_locked_by',
            'programmer_done', 'programmer_done_date',
            'programmer_done_by', 'is_cnc_rework', 'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'programmer_status': new_status,
                     'programmer_time_spent_minutes': sample.programmer_time_spent_minutes},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_rework')
    def set_rework(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_rework/
        body: { "is_rework": true }   → flag for designer rework, resets designer done
              { "is_rework": false }  → clear rework flag
        """
        sample = self.get_object()
        is_rework = str(request.data.get('is_rework', 'true')).lower() in ('true', '1', 'yes')

        sample.is_rework = is_rework
        if is_rework:
            # Send back to designer — reset their completion and chronometer
            sample.is_done          = False
            sample.designer_status  = None
            sample.time_started     = None
            # Reset programmer tracking so programmer can re-run when designer is done again
            sample.programmer_status           = None
            sample.programmer_time_started     = None
            sample.programmer_time_spent_minutes = 0
            sample.programmer_done             = False
            sample.programmer_done_date        = None
            sample.programmer_done_by          = None
        else:
            # Programmer manually clears rework flag
            pass

        sample.updated_by = request.user
        save_fields = [
            'is_rework', 'is_done', 'designer_status', 'time_started',
            'programmer_status', 'programmer_time_started', 'programmer_time_spent_minutes',
            'programmer_done', 'programmer_done_date', 'programmer_done_by',
            'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'is_rework': is_rework},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_cnc_status')
    def set_cnc_status(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_cnc_status/
        body: { "cnc_status": "ongoing"|"standby"|"done", "pause_reason": "..." }

        Multi-worker: multiple users can work simultaneously.
        Timer starts when the FIRST worker joins, accumulates when the LAST worker pauses.
        """
        sample = self.get_object()
        new_status = request.data.get('cnc_status', '').lower()
        valid = {'ongoing', 'standby', 'done'}
        if new_status not in valid:
            return Response(
                {'error': f'cnc_status must be one of {sorted(valid)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()
        workers = list(sample.cnc_active_workers or [])
        user_entry = {
            'id': request.user.pk,
            'name': request.user.get_full_name() or request.user.username,
        }
        is_active = any(w['id'] == request.user.pk for w in workers)

        if new_status == 'ongoing':
            # Add user to workers list if not already there
            if not is_active:
                workers.append(user_entry)
            # Start timer on the first worker joining
            if not sample.cnc_time_started:
                sample.cnc_time_started = now
            sample.cnc_active_workers = workers
            sample.cnc_status = 'ongoing'
            sample.cnc_pause_reason = None

        elif new_status == 'standby':
            # Only the requesting user can leave
            workers = [w for w in workers if w['id'] != request.user.pk]
            sample.cnc_active_workers = workers
            if not workers:
                # Last worker paused — accumulate time and mark standby
                if sample.cnc_time_started:
                    delta = now - sample.cnc_time_started
                    sample.cnc_time_spent_minutes += int(delta.total_seconds() // 60)
                sample.cnc_time_started = None
                sample.cnc_status = 'standby'
                sample.cnc_pause_reason = request.data.get('pause_reason') or None
            # else: other workers still active, keep 'ongoing' status

        elif new_status == 'done':
            # Finalize — any worker can mark done
            if sample.cnc_time_started:
                delta = now - sample.cnc_time_started
                sample.cnc_time_spent_minutes += int(delta.total_seconds() // 60)
            sample.cnc_time_started = None
            sample.cnc_pause_reason = None
            sample.cnc_active_workers = []
            sample.cnc_done = True
            sample.cnc_done_date = now.date()
            sample.cnc_done_by = request.user
            sample.cnc_locked_by = None
            sample.is_cnc_rework = False
            sample.is_assembly_rework = False
            sample.cnc_status = 'done'

        sample.updated_by = request.user
        save_fields = [
            'cnc_status', 'cnc_time_started', 'cnc_time_spent_minutes',
            'cnc_pause_reason', 'cnc_locked_by', 'cnc_active_workers',
            'cnc_done', 'cnc_done_date', 'cnc_done_by',
            'is_cnc_rework', 'is_assembly_rework', 'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'cnc_status': sample.cnc_status,
                     'cnc_time_spent_minutes': sample.cnc_time_spent_minutes},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_cnc_rework')
    def set_cnc_rework(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_cnc_rework/
        body: { "is_cnc_rework": true }  → send back to programmer for correction
        """
        sample = self.get_object()
        is_cnc_rework = str(request.data.get('is_cnc_rework', 'true')).lower() in ('true', '1', 'yes')

        sample.is_cnc_rework = is_cnc_rework
        if is_cnc_rework:
            # Send back to programmer — reset their completion and CNC chronometer
            sample.programmer_done               = False
            sample.programmer_done_date          = None
            sample.programmer_status             = None
            sample.programmer_time_started       = None
            sample.programmer_time_spent_minutes = 0
            sample.programmer_done_by            = None
            sample.programmer_locked_by          = None
            # Reset CNC tracking
            sample.cnc_status             = None
            sample.cnc_time_started       = None
            sample.cnc_time_spent_minutes = 0
            sample.cnc_done               = False
            sample.cnc_done_date          = None
            sample.cnc_done_by            = None
            sample.cnc_locked_by          = None
            sample.cnc_active_workers     = []

        sample.updated_by = request.user
        save_fields = [
            'is_cnc_rework',
            'programmer_done', 'programmer_done_date', 'programmer_status',
            'programmer_time_started', 'programmer_time_spent_minutes',
            'programmer_done_by', 'programmer_locked_by',
            'cnc_status', 'cnc_time_started', 'cnc_time_spent_minutes',
            'cnc_done', 'cnc_done_date', 'cnc_done_by', 'cnc_locked_by',
            'cnc_active_workers',
            'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'is_cnc_rework': is_cnc_rework},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_cnc_count')
    def set_cnc_count(self, request, pk=None):
        """PATCH /api/samples/{id}/set_cnc_count/  body: { "cnc_produced_count": N }"""
        sample = self.get_object()
        try:
            count = max(0, int(request.data.get('cnc_produced_count', 0)))
        except (TypeError, ValueError):
            count = 0
        sample.cnc_produced_count = count
        sample.updated_by = request.user
        sample.save(update_fields=['cnc_produced_count', 'updated_by', 'updated_at'])
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_assembly_status')
    def set_assembly_status(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_assembly_status/
        body: { "assembly_status": "ongoing"|"standby"|"done", "pause_reason": "..." }

        Multi-worker: multiple users can work simultaneously.
        Timer starts when the FIRST worker joins, accumulates when the LAST worker pauses.
        """
        sample = self.get_object()
        new_status = request.data.get('assembly_status', '').lower()
        valid = {'ongoing', 'standby', 'done'}
        if new_status not in valid:
            return Response(
                {'error': f'assembly_status must be one of {sorted(valid)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()
        workers = list(sample.assembly_active_workers or [])
        user_entry = {
            'id': request.user.pk,
            'name': request.user.get_full_name() or request.user.username,
        }
        is_active = any(w['id'] == request.user.pk for w in workers)

        if new_status == 'ongoing':
            # Add user to workers list if not already there
            if not is_active:
                workers.append(user_entry)
            # Start timer on the first worker joining
            if not sample.assembly_time_started:
                sample.assembly_time_started = now
            sample.assembly_active_workers = workers
            sample.assembly_status = 'ongoing'
            sample.assembly_pause_reason = None

        elif new_status == 'standby':
            # Only the requesting user can leave
            workers = [w for w in workers if w['id'] != request.user.pk]
            sample.assembly_active_workers = workers
            if not workers:
                # Last worker paused — accumulate time and mark standby
                if sample.assembly_time_started:
                    delta = now - sample.assembly_time_started
                    sample.assembly_time_spent_minutes += int(delta.total_seconds() // 60)
                sample.assembly_time_started = None
                sample.assembly_status = 'standby'
                sample.assembly_pause_reason = request.data.get('pause_reason') or None
            # else: other workers still active, keep 'ongoing' status

        elif new_status == 'done':
            # Finalize — any worker can mark done
            if sample.assembly_time_started:
                delta = now - sample.assembly_time_started
                sample.assembly_time_spent_minutes += int(delta.total_seconds() // 60)
            sample.assembly_time_started = None
            sample.assembly_pause_reason = None
            sample.assembly_active_workers = []
            sample.assembly_done = True
            sample.assembly_done_date = now.date()
            sample.assembly_done_by = request.user
            sample.assembly_locked_by = None
            sample.is_assembly_rework = False
            sample.is_quality_rework = False   # clear quality rework — ready for quality
            sample.assembly_status = 'done'

        sample.updated_by = request.user
        save_fields = [
            'assembly_status', 'assembly_time_started', 'assembly_time_spent_minutes',
            'assembly_pause_reason', 'assembly_locked_by', 'assembly_active_workers',
            'assembly_done', 'assembly_done_date', 'assembly_done_by',
            'is_assembly_rework', 'is_quality_rework', 'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'assembly_status': sample.assembly_status,
                     'assembly_time_spent_minutes': sample.assembly_time_spent_minutes},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_assembly_rework')
    def set_assembly_rework(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_assembly_rework/
        body: { "is_assembly_rework": true }  → send back to CNC for correction
        """
        sample = self.get_object()
        is_assembly_rework = str(request.data.get('is_assembly_rework', 'true')).lower() in ('true', '1', 'yes')

        sample.is_assembly_rework = is_assembly_rework
        if is_assembly_rework:
            # Send back to CNC — reset their completion and assembly chronometer
            sample.cnc_done                  = False
            sample.cnc_done_date             = None
            sample.cnc_status                = None
            sample.cnc_time_started          = None
            sample.cnc_time_spent_minutes    = 0
            sample.cnc_done_by               = None
            sample.cnc_locked_by             = None
            # Reset assembly tracking
            sample.assembly_status             = None
            sample.assembly_time_started       = None
            sample.assembly_time_spent_minutes = 0
            sample.assembly_done               = False
            sample.assembly_done_date          = None
            sample.assembly_done_by            = None
            sample.assembly_locked_by          = None
            sample.assembly_active_workers     = []
            # Also reset CNC active workers
            sample.cnc_active_workers          = []

        sample.updated_by = request.user
        save_fields = [
            'is_assembly_rework',
            'cnc_done', 'cnc_done_date', 'cnc_status',
            'cnc_time_started', 'cnc_time_spent_minutes',
            'cnc_done_by', 'cnc_locked_by', 'cnc_active_workers',
            'assembly_status', 'assembly_time_started', 'assembly_time_spent_minutes',
            'assembly_done', 'assembly_done_date', 'assembly_done_by',
            'assembly_locked_by', 'assembly_active_workers',
            'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'is_assembly_rework': is_assembly_rework},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_assembly_count')
    def set_assembly_count(self, request, pk=None):
        """PATCH /api/samples/{id}/set_assembly_count/  body: { "assembly_produced_count": N }"""
        sample = self.get_object()
        try:
            count = max(0, int(request.data.get('assembly_produced_count', 0)))
        except (TypeError, ValueError):
            count = 0
        sample.assembly_produced_count = count
        sample.updated_by = request.user
        sample.save(update_fields=['assembly_produced_count', 'updated_by', 'updated_at'])
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='reset_programmer')
    def reset_programmer(self, request, pk=None):
        """Undo a programmer 'done' mark — keeps accumulated time, clears done state."""
        sample = self.get_object()
        sample.programmer_done          = False
        sample.programmer_done_date     = None
        sample.programmer_done_by       = None
        sample.programmer_status        = None
        sample.programmer_time_started  = None
        sample.updated_by = request.user
        sample.save(update_fields=[
            'programmer_done', 'programmer_done_date', 'programmer_done_by',
            'programmer_status', 'programmer_time_started',
            'updated_by', 'updated_at',
        ])
        AuditLog.objects.create(sample=sample, user=request.user, action='update',
                                changes={'reset': 'programmer_done'})
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='reset_cnc')
    def reset_cnc(self, request, pk=None):
        """Undo a CNC 'done' mark — keeps accumulated time, clears done state."""
        sample = self.get_object()
        sample.cnc_done             = False
        sample.cnc_done_date        = None
        sample.cnc_done_by          = None
        sample.cnc_status           = None
        sample.cnc_time_started     = None
        sample.cnc_active_workers   = []
        sample.updated_by = request.user
        sample.save(update_fields=[
            'cnc_done', 'cnc_done_date', 'cnc_done_by',
            'cnc_status', 'cnc_time_started', 'cnc_active_workers',
            'updated_by', 'updated_at',
        ])
        AuditLog.objects.create(sample=sample, user=request.user, action='update',
                                changes={'reset': 'cnc_done'})
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='reset_assembly')
    def reset_assembly(self, request, pk=None):
        """Undo an assembly 'done' mark — keeps accumulated time, clears done state."""
        sample = self.get_object()
        sample.assembly_done             = False
        sample.assembly_done_date        = None
        sample.assembly_done_by          = None
        sample.assembly_status           = None
        sample.assembly_time_started     = None
        sample.assembly_active_workers   = []
        sample.updated_by = request.user
        sample.save(update_fields=[
            'assembly_done', 'assembly_done_date', 'assembly_done_by',
            'assembly_status', 'assembly_time_started', 'assembly_active_workers',
            'updated_by', 'updated_at',
        ])
        AuditLog.objects.create(sample=sample, user=request.user, action='update',
                                changes={'reset': 'assembly_done'})
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_quality_status')
    def set_quality_status(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_quality_status/
        body: { "quality_status": "ongoing"|"standby"|"done", "pause_reason": "..." }

        Multi-worker: multiple quality controllers can work simultaneously.
        """
        sample = self.get_object()
        new_status = request.data.get('quality_status', '').lower()
        valid = {'ongoing', 'standby', 'done'}
        if new_status not in valid:
            return Response(
                {'error': f'quality_status must be one of {sorted(valid)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()
        workers = list(sample.quality_active_workers or [])
        user_entry = {
            'id': request.user.pk,
            'name': request.user.get_full_name() or request.user.username,
        }
        is_active = any(w['id'] == request.user.pk for w in workers)

        if new_status == 'ongoing':
            if not is_active:
                workers.append(user_entry)
            if not sample.quality_time_started:
                sample.quality_time_started = now
            sample.quality_active_workers = workers
            sample.quality_status = 'ongoing'
            sample.quality_pause_reason = None

        elif new_status == 'standby':
            workers = [w for w in workers if w['id'] != request.user.pk]
            sample.quality_active_workers = workers
            if not workers:
                if sample.quality_time_started:
                    delta = now - sample.quality_time_started
                    sample.quality_time_spent_minutes += int(delta.total_seconds() // 60)
                sample.quality_time_started = None
                sample.quality_status = 'standby'
                sample.quality_pause_reason = request.data.get('pause_reason') or None

        elif new_status == 'done':
            if sample.quality_time_started:
                delta = now - sample.quality_time_started
                sample.quality_time_spent_minutes += int(delta.total_seconds() // 60)
            sample.quality_time_started = None
            sample.quality_pause_reason = None
            sample.quality_active_workers = []
            sample.quality_done = True
            sample.quality_done_date = now.date()
            sample.quality_done_by = request.user
            sample.is_quality_rework = False
            sample.quality_status = 'done'

        sample.updated_by = request.user
        save_fields = [
            'quality_status', 'quality_time_started', 'quality_time_spent_minutes',
            'quality_pause_reason', 'quality_active_workers',
            'quality_done', 'quality_done_date', 'quality_done_by',
            'is_quality_rework', 'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'quality_status': sample.quality_status,
                     'quality_time_spent_minutes': sample.quality_time_spent_minutes},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='set_quality_rework')
    def set_quality_rework(self, request, pk=None):
        """
        PATCH /api/samples/{id}/set_quality_rework/
        body: { "is_quality_rework": true }  → send back to assembly for correction
        """
        sample = self.get_object()
        is_quality_rework = str(request.data.get('is_quality_rework', 'true')).lower() in ('true', '1', 'yes')

        sample.is_quality_rework = is_quality_rework
        if is_quality_rework:
            # Send back to assembly — reset assembly completion and quality chronometer
            sample.assembly_done               = False
            sample.assembly_done_date          = None
            sample.assembly_status             = None
            sample.assembly_time_started       = None
            sample.assembly_time_spent_minutes = 0
            sample.assembly_done_by            = None
            sample.assembly_locked_by          = None
            sample.assembly_active_workers     = []
            # Reset quality tracking
            sample.quality_status             = None
            sample.quality_time_started       = None
            sample.quality_time_spent_minutes = 0
            sample.quality_done               = False
            sample.quality_done_date          = None
            sample.quality_done_by            = None
            sample.quality_active_workers     = []

        sample.updated_by = request.user
        save_fields = [
            'is_quality_rework',
            'assembly_done', 'assembly_done_date', 'assembly_status',
            'assembly_time_started', 'assembly_time_spent_minutes',
            'assembly_done_by', 'assembly_locked_by', 'assembly_active_workers',
            'quality_status', 'quality_time_started', 'quality_time_spent_minutes',
            'quality_done', 'quality_done_date', 'quality_done_by', 'quality_active_workers',
            'updated_by', 'updated_at',
        ]
        sample.save(update_fields=save_fields)

        AuditLog.objects.create(
            sample=sample, user=request.user, action='update',
            changes={'is_quality_rework': is_quality_rework},
        )
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='reset_quality')
    def reset_quality(self, request, pk=None):
        """Undo a quality 'done' mark — keeps accumulated time, clears done state."""
        sample = self.get_object()
        sample.quality_done             = False
        sample.quality_done_date        = None
        sample.quality_done_by          = None
        sample.quality_status           = None
        sample.quality_time_started     = None
        sample.quality_active_workers   = []
        sample.updated_by = request.user
        sample.save(update_fields=[
            'quality_done', 'quality_done_date', 'quality_done_by',
            'quality_status', 'quality_time_started', 'quality_active_workers',
            'updated_by', 'updated_at',
        ])
        AuditLog.objects.create(sample=sample, user=request.user, action='update',
                                changes={'reset': 'quality_done'})
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='audit')
    def audit_log(self, request, pk=None):
        sample = self.get_object()
        logs = sample.audit_logs.select_related('user').all()
        serializer = AuditLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='upload_design')
    def upload_design(self, request, pk=None):
        """
        POST /api/samples/{id}/upload_design/
        Multipart form with optional fields:
          - design_file : CAD file (.sldprt .sldasm .slddrw .step .stp .iges .igs .dxf .toppkg .top .ens)
          - design_pdf  : PDF drawing
        Only designer or admin can upload.
        """
        sample = self.get_object()

        # Permission: designer group or staff/superuser
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Designer').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé. Seuls les designers peuvent uploader des fichiers.'},
                            status=status.HTTP_403_FORBIDDEN)

        ALLOWED_3D = {'.sldprt', '.sldasm', '.slddrw', '.step', '.stp', '.iges', '.igs', '.dxf', '.toppkg', '.top', '.ens'}
        updated = False

        # ── CAD file ──────────────────────────────────────────────────────────
        if 'design_file' in request.FILES:
            f = request.FILES['design_file']
            ext = os.path.splitext(f.name)[1].lower()
            if ext not in ALLOWED_3D:
                return Response(
                    {'error': f'Format 3D non supporté: {ext}. Acceptés: {", ".join(sorted(ALLOWED_3D))}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Delete old file from disk
            if sample.design_file:
                try:
                    sample.design_file.delete(save=False)
                except Exception:
                    pass
            sample.design_file = f
            updated = True

        # ── PDF file ──────────────────────────────────────────────────────────
        if 'design_pdf' in request.FILES:
            f = request.FILES['design_pdf']
            ext = os.path.splitext(f.name)[1].lower()
            if ext != '.pdf':
                return Response(
                    {'error': 'Le plan doit être un fichier PDF.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if sample.design_pdf:
                try:
                    sample.design_pdf.delete(save=False)
                except Exception:
                    pass
            sample.design_pdf = f
            updated = True

        if not updated:
            return Response({'error': 'Aucun fichier fourni (design_file ou design_pdf).'},
                            status=status.HTTP_400_BAD_REQUEST)

        sample.design_uploaded_at = timezone.now()
        sample.design_uploaded_by = request.user
        sample.updated_by = request.user
        sample.save(update_fields=[
            'design_file', 'design_pdf',
            'design_uploaded_at', 'design_uploaded_by',
            'updated_by', 'updated_at',
        ])

        AuditLog.objects.create(
            sample=sample,
            user=request.user,
            action='update',
            changes={'design_upload': True},
        )

        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    # ── Machine list ─────────────────────────────────────────────────────────────
    MACHINES = [
        {'id': 'T600S1', 'name': 'T600S-1', 'ip': '192.168.16.61',  'type': 'mitsubishi_m80a', 'ftp_user': '',  'ftp_pass': '', 'ftp_path': '/'},
        {'id': 'T600S2', 'name': 'T600S-2', 'ip': '192.168.16.71',  'type': 'mitsubishi_m80a', 'ftp_user': '',  'ftp_pass': '', 'ftp_path': '/'},
        {'id': 'T600S3', 'name': 'T600S-3', 'ip': '192.168.16.91',  'type': 'mitsubishi_m80a', 'ftp_user': '',  'ftp_pass': '', 'ftp_path': '/'},
        {'id': 'TX500',  'name': 'TX500',   'ip': '192.168.16.201', 'type': 'fanuc',            'ftp_user': '',  'ftp_pass': '', 'ftp_path': '/'},
    ]

    @action(detail=False, methods=['get'], url_path='machines')
    def machines(self, request):
        """GET /api/samples/machines/ — list all CNC machines"""
        return Response(self.MACHINES)

    @action(detail=True, methods=['post'], url_path='upload_gcode')
    def upload_gcode(self, request, pk=None):
        """
        POST /api/samples/{id}/upload_gcode/
        Multipart: one or more files under key 'gcode_files'
        Only programmer or admin.
        """
        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Programmateur').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)

        files = request.FILES.getlist('gcode_files')
        if not files:
            # fallback: try single file key
            single = request.FILES.get('gcode_file')
            if single:
                files = [single]
        if not files:
            return Response({'error': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        for f in files:
            pf = ProgrammerFile.objects.create(
                sample=sample,
                file=f,
                file_name=f.name,
                uploaded_by=user,
            )
            created.append(pf)

        sample.updated_by = user
        sample.save(update_fields=['updated_by', 'updated_at'])
        AuditLog.objects.create(sample=sample, user=user, action='update',
                                changes={'gcode_upload': [f.file_name for f in created]})

        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['delete'], url_path='delete_programmer_file/(?P<file_id>[0-9]+)')
    def delete_programmer_file(self, request, pk=None, file_id=None):
        """DELETE /api/samples/{id}/delete_programmer_file/{file_id}/"""
        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Programmateur').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            pf = ProgrammerFile.objects.get(id=file_id, sample=sample)
        except ProgrammerFile.DoesNotExist:
            return Response({'error': 'Fichier introuvable.'}, status=status.HTTP_404_NOT_FOUND)
        try:
            pf.file.delete(save=False)
        except Exception:
            pass
        pf.delete()
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['delete'], url_path='delete_design_file')
    def delete_design_file(self, request, pk=None):
        """DELETE /api/samples/{id}/delete_design_file/ — remove CAD file"""
        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Designer').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        if sample.design_file:
            try:
                sample.design_file.delete(save=False)
            except Exception:
                pass
            sample.design_file = None
            sample.updated_by = user
            sample.save(update_fields=['design_file', 'updated_by', 'updated_at'])
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['delete'], url_path='delete_design_pdf')
    def delete_design_pdf(self, request, pk=None):
        """DELETE /api/samples/{id}/delete_design_pdf/ — remove PDF"""
        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Designer').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        if sample.design_pdf:
            try:
                sample.design_pdf.delete(save=False)
            except Exception:
                pass
            sample.design_pdf = None
            sample.updated_by = user
            sample.save(update_fields=['design_pdf', 'updated_by', 'updated_at'])
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='upload_bom')
    def upload_bom(self, request, pk=None):
        """POST /api/samples/{id}/upload_bom/ — upload BOM Excel (.xlsx/.xls) and auto-import line items."""
        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Designer').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        f = request.FILES.get('bom_file')
        if not f:
            return Response({'error': 'Aucun fichier fourni (bom_file).'}, status=status.HTTP_400_BAD_REQUEST)
        ext = os.path.splitext(f.name)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return Response({'error': 'Le BOM doit être un fichier Excel (.xlsx ou .xls).'}, status=status.HTTP_400_BAD_REQUEST)
        if sample.bom_pdf:
            try:
                sample.bom_pdf.delete(save=False)
            except Exception:
                pass
        sample.bom_pdf = f
        sample.bom_uploaded_at = timezone.now()
        sample.updated_by = user
        sample.save(update_fields=['bom_pdf', 'bom_uploaded_at', 'updated_by', 'updated_at'])

        # ── Parse Excel and auto-import BOM items ──────────────────────────────
        imported, err = _parse_bom_excel(f, sample)

        serializer = SampleDetailSerializer(sample, context={'request': request})
        data = serializer.data
        data['bom_imported_count'] = imported
        if err:
            data['bom_import_warning'] = err
        return Response(data)

    @action(detail=True, methods=['delete'], url_path='delete_bom')
    def delete_bom(self, request, pk=None):
        """DELETE /api/samples/{id}/delete_bom/ — remove BOM PDF"""
        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Designer').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        if sample.bom_pdf:
            try:
                sample.bom_pdf.delete(save=False)
            except Exception:
                pass
            sample.bom_pdf = None
            sample.bom_uploaded_at = None
            sample.updated_by = user
            sample.save(update_fields=['bom_pdf', 'bom_uploaded_at', 'updated_by', 'updated_at'])
        serializer = SampleDetailSerializer(sample, context={'request': request})
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='bom_aggregate')
    def bom_aggregate(self, request):
        """GET /api/samples/bom_aggregate/?project=XX — aggregate BOM items across samples."""
        project = request.query_params.get('project', '').strip()
        qs = BomItem.objects.select_related('sample')
        if project:
            qs = qs.filter(sample__project=project)

        aggregated = {}
        for item in qs:
            key = (item.reference, item.unit)
            if key not in aggregated:
                aggregated[key] = {
                    'reference':   item.reference,
                    'designation': item.designation,
                    'unit':        item.unit,
                    'total_qty':   0,
                    'samples':     [],
                }
            aggregated[key]['total_qty'] += float(item.quantity)
            aggregated[key]['samples'].append(item.sample.apn)

        result = sorted(aggregated.values(), key=lambda x: x['reference'])
        return Response(result)

    @action(detail=True, methods=['post'], url_path='send_to_machine')
    def send_to_machine(self, request, pk=None):
        """
        POST /api/samples/{id}/send_to_machine/
        Body: { "machine_id": "T600S1" }
        Sends the gcode_file to the CNC machine via FTP.
        """
        import ftplib
        import socket

        sample = self.get_object()
        user = request.user
        is_allowed = (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='CNC').exists()
        )
        if not is_allowed:
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)

        if not sample.gcode_file:
            return Response({'error': 'Aucun fichier G-code uploadé pour cet échantillon.'}, status=status.HTTP_400_BAD_REQUEST)

        machine_id = request.data.get('machine_id')
        machine = next((m for m in self.MACHINES if m['id'] == machine_id), None)
        if not machine:
            return Response({'error': f'Machine inconnue: {machine_id}'}, status=status.HTTP_400_BAD_REQUEST)

        file_path = sample.gcode_file.path
        file_name = os.path.basename(file_path)

        try:
            ftp = ftplib.FTP()
            ftp.connect(machine['ip'], 21, timeout=10)
            ftp.login(machine['ftp_user'] or 'anonymous', machine['ftp_pass'] or '')

            # Navigate to target folder
            if machine['ftp_path'] and machine['ftp_path'] != '/':
                try:
                    ftp.cwd(machine['ftp_path'])
                except ftplib.error_perm:
                    pass  # Stay at root if folder doesn't exist

            with open(file_path, 'rb') as f:
                ftp.storbinary(f'STOR {file_name}', f)

            ftp.quit()

            AuditLog.objects.create(
                sample=sample, user=user, action='update',
                changes={'send_to_machine': machine_id, 'file': file_name}
            )
            return Response({
                'success': True,
                'message': f'✅ {file_name} envoyé à {machine["name"]} ({machine["ip"]})',
                'machine': machine['name'],
                'file': file_name,
            })

        except socket.timeout:
            return Response({'error': f'⏱ Timeout: machine {machine["name"]} ({machine["ip"]}) ne répond pas. Vérifiez qu\'elle est allumée et connectée au réseau.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except ConnectionRefusedError:
            return Response({'error': f'🔌 Connexion refusée par {machine["name"]} ({machine["ip"]}). FTP non activé sur la machine?'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except ftplib.error_perm as e:
            return Response({'error': f'🔐 Erreur FTP (permission): {str(e)}'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as e:
            return Response({'error': f'❌ Erreur: {str(e)}'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)


# ── JIMIDE-4030 DXF file management ──────────────────────────────────────────

def _serialize_dxf(obj, request):
    return {
        'id': obj.id,
        'file_name': obj.file_name,
        'description': obj.description,
        'uploaded_at': obj.uploaded_at.isoformat(),
        'uploaded_by': (
            f"{obj.uploaded_by.first_name} {obj.uploaded_by.last_name}".strip()
            or obj.uploaded_by.username
        ) if obj.uploaded_by else None,
        'file_url': request.build_absolute_uri(f'/media/{obj.file.name}') if obj.file else None,
    }


class BomItemViewSet(viewsets.ModelViewSet):
    """CRUD for BOM items linked to a sample."""
    permission_classes = [IsAuthenticated]
    serializer_class   = BomItemSerializer

    def get_queryset(self):
        qs = BomItem.objects.select_related('sample')
        sample_id = self.request.query_params.get('sample')
        if sample_id:
            qs = qs.filter(sample_id=sample_id)
        return qs


class JimideDxfViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def _is_allowed(self, user):
        return (
            user.is_staff or user.is_superuser or
            user.groups.filter(name='Designer').exists()
        )

    def list(self, request):
        """GET /api/jimide/ — list all DXF files"""
        if not self._is_allowed(request.user):
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        qs = JimideDxfFile.objects.select_related('uploaded_by').all()
        return Response([_serialize_dxf(obj, request) for obj in qs])

    def create(self, request):
        """POST /api/jimide/ — upload a DXF file"""
        from django.conf import settings as django_settings

        if not self._is_allowed(request.user):
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)

        f = request.FILES.get('dxf_file')
        if not f:
            return Response({'error': 'Aucun fichier fourni (clé: dxf_file).'}, status=status.HTTP_400_BAD_REQUEST)

        ext = os.path.splitext(f.name)[1].lower()
        if ext != '.dxf':
            return Response({'error': 'Seuls les fichiers .dxf sont acceptés.'}, status=status.HTTP_400_BAD_REQUEST)

        description = request.data.get('description', '')

        obj = JimideDxfFile.objects.create(
            file=f,
            file_name=f.name,
            description=description,
            uploaded_by=request.user,
        )

        # Also copy to the laptop DXF folder
        try:
            dxf_folder = Path(django_settings.JIMIDE_DXF_FOLDER)
            dxf_folder.mkdir(parents=True, exist_ok=True)
            dest = dxf_folder / f.name
            if dest.exists():
                base, ext2 = os.path.splitext(f.name)
                dest = dxf_folder / f'{base}_{int(time.time())}{ext2}'
            shutil.copy2(obj.file.path, dest)
        except Exception:
            pass  # Don't fail the upload if the local copy fails

        return Response(_serialize_dxf(obj, request), status=status.HTTP_201_CREATED)

    def destroy(self, request, pk=None):
        """DELETE /api/jimide/{id}/ — delete a DXF file"""
        if not self._is_allowed(request.user):
            return Response({'error': 'Accès refusé.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            obj = JimideDxfFile.objects.get(pk=pk)
        except JimideDxfFile.DoesNotExist:
            return Response({'error': 'Fichier introuvable.'}, status=status.HTTP_404_NOT_FOUND)
        try:
            obj.file.delete(save=False)
        except Exception:
            pass
        obj.delete()
        return Response({'ok': True})


# ── Technical Study Validation ────────────────────────────────────────────────

class MatrixEntryViewSet(viewsets.ModelViewSet):
    """CRUD for the reference matrix entries. Write access: admin or Etude Technique."""

    queryset = MatrixEntry.objects.all()
    serializer_class = MatrixEntrySerializer
    pagination_class = None

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy', 'import_csv'):
            return [IsTechStudyOrAdmin()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser])
    def import_csv(self, request):
        """POST /api/matrix/import/ — bulk upsert from CSV. Columns: reference, designation, quantity, sample_type, notes."""
        # Permission enforced by get_permissions; no need to check again here.

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.lower().endswith('.csv'):
            return Response({'error': 'Le fichier doit être au format CSV.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            text = file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                file.seek(0)
                text = file.read().decode('latin-1')
            except Exception:
                return Response({'error': 'Encodage du fichier non supporté (utilisez UTF-8).'}, status=status.HTTP_400_BAD_REQUEST)

        valid_types = {c[0] for c in CONNECTOR_FILL_CHOICES}
        reader = csv_module.DictReader(io.StringIO(text))
        created, updated, errors = 0, 0, []

        for i, row in enumerate(reader, start=2):
            ref = (row.get('reference') or row.get('Reference') or '').strip()
            if not ref:
                errors.append(f'Ligne {i}: référence manquante — ignorée.')
                continue

            try:
                qty = max(1, int((row.get('quantity') or row.get('Quantity') or '1').strip()))
            except ValueError:
                qty = 1

            stype = (row.get('sample_type') or row.get('type') or row.get('Type') or '').strip().lower()
            if stype not in valid_types:
                stype = ''

            designation = (row.get('designation') or row.get('Designation') or '').strip()
            notes = (row.get('notes') or row.get('Notes') or '').strip()

            # Les références peuvent exister en plusieurs exemplaires (variantes
            # importées d'Excel) : le CSV met à jour la première, sinon crée.
            entry = MatrixEntry.objects.filter(reference=ref).first()
            if entry:
                entry.designation = designation
                entry.quantity = qty
                entry.sample_type = stype
                entry.notes = notes
                entry.save()
                updated += 1
            else:
                MatrixEntry.objects.create(
                    reference=ref, designation=designation, quantity=qty,
                    sample_type=stype, notes=notes, created_by=request.user,
                )
                created += 1

        return Response({'created': created, 'updated': updated, 'errors': errors})


def _run_comparison(project_name):
    """Compare actual project samples against the reference matrix.

    La matrice peut contenir plusieurs lignes pour la même référence
    (variantes importées d'Excel) : elles sont agrégées par référence —
    quantité attendue = somme des quantités. Le contrôle de type n'est
    appliqué que si le type matrice est un type connecteur connu
    (full/empty/partial) ; un texte libre (Equipment) est informatif.
    """
    from collections import defaultdict

    # Aggregate project samples by APN; track approval status
    project_map = defaultdict(lambda: {'quantity': 0, 'types': set(), 'approved': False})
    for s in Sample.objects.filter(project=project_name):
        project_map[s.apn]['quantity'] += s.quantity
        project_map[s.apn]['types'].add(s.connector_fill or '')
        if s.study_approved:
            project_map[s.apn]['approved'] = True

    # Aggregate matrix entries by reference (duplicates = variantes)
    matrix = defaultdict(lambda: {'quantity': 0, 'designation': '', 'types': set()})
    for e in MatrixEntry.objects.all():
        m = matrix[e.reference]
        m['quantity'] += e.quantity
        if e.designation and not m['designation']:
            m['designation'] = e.designation
        if e.sample_type:
            m['types'].add(e.sample_type)

    fill_values = {c[0] for c in CONNECTOR_FILL_CHOICES}
    matched, missing, mismatched, extra = [], [], [], []

    for ref, entry in matrix.items():
        types_matrice = entry['types']
        matrix_type = (list(types_matrice)[0] if len(types_matrice) == 1
                       else ('mixed' if types_matrice else ''))
        ps = project_map.get(ref)
        if ps is None:
            missing.append({
                'reference': ref,
                'designation': entry['designation'],
                'matrix_quantity': entry['quantity'],
                'matrix_type': matrix_type,
                'project_quantity': None,
                'project_type': None,
                'status': 'missing',
                'approved': False,
            })
        else:
            types = ps['types']
            proj_type = list(types)[0] if len(types) == 1 else ('mixed' if types else '')
            qty_match = ps['quantity'] == entry['quantity']
            # Type imposé uniquement s'il s'agit d'un type connecteur connu
            type_match = (matrix_type not in fill_values) or (proj_type == matrix_type)

            if qty_match and type_match:
                matched.append({
                    'reference': ref,
                    'designation': entry['designation'],
                    'matrix_quantity': entry['quantity'],
                    'matrix_type': matrix_type,
                    'project_quantity': ps['quantity'],
                    'project_type': proj_type,
                    'status': 'matched',
                    'approved': ps['approved'],
                })
            else:
                mismatched.append({
                    'reference': ref,
                    'designation': entry['designation'],
                    'matrix_quantity': entry['quantity'],
                    'matrix_type': matrix_type,
                    'project_quantity': ps['quantity'],
                    'project_type': proj_type,
                    'status': 'mismatched',
                    'approved': ps['approved'],
                })

    for apn, ps in project_map.items():
        if apn not in matrix:
            types = ps['types']
            proj_type = list(types)[0] if len(types) == 1 else ('mixed' if types else '')
            extra.append({
                'reference': apn,
                'designation': '',
                'matrix_quantity': None,
                'matrix_type': None,
                'project_quantity': ps['quantity'],
                'project_type': proj_type,
                'status': 'extra',
                'approved': ps['approved'],
            })

    result_status = 'approved' if not missing and not mismatched and not extra else 'rejected'
    return {
        'validation_status': result_status,
        'matched': matched,
        'missing': missing,
        'mismatched': mismatched,
        'extra': extra,
        'summary': {
            'total_matrix': len(matrix),
            'total_project': len(project_map),
            'matched': len(matched),
            'missing': len(missing),
            'mismatched': len(mismatched),
            'extra': len(extra),
        },
    }


@api_view(['GET'])
@drf_permission_classes([IsAuthenticated])
def validation_list(request):
    """GET /api/validation/ — all projects (from samples OR ProjectValidation records)."""
    sample_projects = set(
        Sample.objects.values_list('project', flat=True).distinct()
    )
    pv_qs = ProjectValidation.objects.select_related('validated_by', 'approved_by').all()
    pv_map = {v.project_name: v for v in pv_qs}

    all_names = sorted(sample_projects | set(pv_map.keys()))

    # Sample counts per project in one query
    from django.db.models import Count
    counts = dict(
        Sample.objects.filter(project__in=all_names)
        .values('project').annotate(n=Count('id')).values_list('project', 'n')
    )

    result = []
    for name in all_names:
        v = pv_map.get(name)
        result.append({
            'project_name': name,
            'validation_status': v.validation_status if v else 'pending',
            'sample_count': counts.get(name, 0),
            'validated_at': v.validated_at.isoformat() if v and v.validated_at else None,
            'validated_by': (
                v.validated_by.get_full_name() or v.validated_by.username
            ) if v and v.validated_by else None,
            'approved_at': v.approved_at.isoformat() if v and v.approved_at else None,
            'approved_by': (
                v.approved_by.get_full_name() or v.approved_by.username
            ) if v and v.approved_by else None,
        })
    return Response(result)


@api_view(['POST'])
@drf_permission_classes([IsAuthenticated])
def validation_run(request):
    """POST /api/validation/run/  body: { project_name }"""
    project_name = request.data.get('project_name', '').strip()
    if not project_name:
        return Response({'error': 'project_name est requis.'}, status=status.HTTP_400_BAD_REQUEST)

    result = _run_comparison(project_name)
    new_status = result['validation_status']

    pv, _ = ProjectValidation.objects.update_or_create(
        project_name=project_name,
        defaults={
            'validation_status': new_status,
            'validated_at': timezone.now(),
            'validated_by': request.user,
            'approved_at': None,
            'approved_by': None,
            'result': result,
        },
    )
    return Response({**result, 'validation': ProjectValidationSerializer(pv).data})


@api_view(['POST'])
@drf_permission_classes([IsTechStudyOrAdmin])
def validation_approve(request):
    """POST /api/validation/approve/  body: { project_name }"""
    project_name = request.data.get('project_name', '').strip()
    if not project_name:
        return Response({'error': 'project_name est requis.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        pv = ProjectValidation.objects.get(project_name=project_name)
    except ProjectValidation.DoesNotExist:
        return Response({'error': 'Aucune validation trouvée. Lancez d\'abord la vérification.'}, status=status.HTTP_404_NOT_FOUND)

    # Project can be approved at any time, regardless of validation_status.
    # Individual APN approval can be adjusted afterward via the comparison table.
    pv.approved_at = timezone.now()
    pv.approved_by = request.user
    pv.save(update_fields=['approved_at', 'approved_by', 'updated_at'])

    # Auto-approve all samples in this project so they appear in the dashboard.
    # Admin can individually un-approve specific APNs afterward via the comparison table.
    Sample.objects.filter(project=project_name).update(study_approved=True)

    return Response(ProjectValidationSerializer(pv).data)


# ── Project management ────────────────────────────────────────────────────────

_CLIENT_VALUES = [c[0] for c in CLIENT_CHOICES]
_CONNECTOR_VALUES = [c[0] for c in CONNECTOR_FILL_CHOICES]


@api_view(['POST'])
@drf_permission_classes([IsAuthenticated])
def project_create(request):
    """POST /api/validation/projects/  body: { project_name }"""
    project_name = request.data.get('project_name', '').strip()
    if not project_name:
        return Response({'error': 'project_name est requis.'}, status=status.HTTP_400_BAD_REQUEST)
    if ProjectValidation.objects.filter(project_name=project_name).exists():
        return Response({'error': 'Un projet avec ce nom existe déjà.'}, status=status.HTTP_409_CONFLICT)
    pv = ProjectValidation.objects.create(project_name=project_name)
    return Response({
        'project_name': pv.project_name,
        'validation_status': pv.validation_status,
        'sample_count': 0,
        'validated_at': None,
        'validated_by': None,
        'approved_at': None,
        'approved_by': None,
    }, status=status.HTTP_201_CREATED)


# ── Import Excel client (Board Specification) ────────────────────────────────
# Feuille « Board Specification » : en-têtes ligne 11, données à partir de la
# ligne 13. Colonnes de base : I=Status, J=Item, K=Equipment, L=Kit name,
# M=Component APN, N=Customer ID, O=Holder APN-ID.
# Colonnes de variante : Z=Holder type (Wifi / Mechanical / Electrified),
# AE=Aluminium Sliding Part, AF=Metallic Inlet, AG=Metallic Coding,
# AH=POM part (quantités par item) — utilisées pour dériver un libellé
# d'item parlant : « Wi-Fi detection », « Body clip detection »,
# « Metallic Coding »…

EXCEL_SHEET_NAME = 'Board Specification'
EXCEL_DATA_MIN_ROW = 13
EXCEL_COLS = {'status': 9, 'item': 10, 'equipment': 11, 'kit': 12,
              'component_apn': 13, 'customer_id': 14, 'holder_apn': 15}
EXCEL_VARIANT_COLS = {'component_type': 16, 'colour': 20, 'board_location': 21,
                      'comments': 24, 'qty_item': 25, 'holder_type': 26,
                      'support': 27, 'lock_type': 29,
                      'alu_sliding': 31, 'metallic_inlet': 32,
                      'metallic_coding': 33, 'pom_part': 34}
EXCEL_MAX_COL = 34


def _cell_str(value):
    """Valeur de cellule → texte propre ('' si vide, 123.0 → '123')."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _qty_positive(value):
    try:
        return float(str(value).replace(',', '.')) > 0
    except (TypeError, ValueError):
        return False


def _item_label(cells):
    """Libellé de variante dérivé d'Equipment / Holder type / options.

    Exemples : « Wi-Fi detection », « Body clip detection (Wi-Fi), Metallic
    Coding », « Electrified detection », « Mechanical »."""
    equipment = cells['equipment'].lower()
    holder_type = cells['holder_type']
    detection = holder_type.lower().replace('-', '').replace(' ', '')
    parts = []
    if 'body clip' in equipment:
        base = 'Body clip detection'
        if 'wifi' in detection:
            base += ' (Wi-Fi)'
        elif 'electrified' in detection:
            base += ' (Electrified)'
        parts.append(base)
    elif 'wifi' in detection:
        parts.append('Wi-Fi detection')
    elif 'electrified' in detection:
        parts.append('Electrified detection')
    elif holder_type:
        parts.append(holder_type)
    if _qty_positive(cells['metallic_coding']):
        parts.append('Metallic Coding')
    if _qty_positive(cells['metallic_inlet']):
        parts.append('Metallic Inlet')
    if _qty_positive(cells['alu_sliding']):
        parts.append('Aluminium Sliding Part')
    if _qty_positive(cells['pom_part']):
        parts.append('POM part 90°')
    label = ', '.join(parts)
    # Préfixe « Component type » (P) : Connector, Ring / Battery Terminal…
    # (redondant pour les body clips, déjà dans « Body clip detection »)
    component_type = cells['component_type']
    if component_type and 'body clip' not in equipment \
            and component_type.lower() != cells['equipment'].lower():
        label = f'{component_type} — {label}' if label else component_type
    return label


def _parse_board_specification(file):
    """Parse le fichier Excel → (rows, ignored_details). Lève ValueError."""
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
    except Exception:
        raise ValueError('Fichier illisible — utilisez un classeur Excel (.xlsm / .xlsx).')

    if EXCEL_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f'Feuille « {EXCEL_SHEET_NAME} » introuvable. '
            f'Feuilles disponibles : {", ".join(workbook.sheetnames)}.')
    sheet = workbook[EXCEL_SHEET_NAME]

    rows, ignored = [], []
    for index, values in enumerate(
            sheet.iter_rows(min_row=EXCEL_DATA_MIN_ROW, min_col=9,
                            max_col=EXCEL_MAX_COL, values_only=True),
            start=EXCEL_DATA_MIN_ROW):
        cells = {key: _cell_str(values[col - 9])
                 for key, col in {**EXCEL_COLS, **EXCEL_VARIANT_COLS}.items()}
        if not any(cells[key] for key in EXCEL_COLS):
            continue   # ligne (de base) entièrement vide — ignorée silencieusement
        if not cells['holder_apn'] and not cells['component_apn']:
            # Lignes de gabarit : seuls Status / n° d'item préremplis → silencieux.
            if not any(cells[k] for k in ('equipment', 'kit', 'customer_id')):
                continue
            ignored.append({
                'row': index,
                'reason': 'Holder APN-ID et Component APN vides.',
            })
            continue
        # Item enrichi : libellé de variante + n° d'origine du fichier client
        label = _item_label(cells)
        if label:
            numero_item = cells['item']
            cells['item'] = f'{label} (n°{numero_item})' if numero_item else label
        apn = (cells['holder_apn'] or cells['component_apn'])[:50]
        description = ' | '.join(filter(None, [
            f"Component APN: {cells['component_apn']}" if cells['component_apn'] else '',
            f"Customer ID: {cells['customer_id']}" if cells['customer_id'] else '',
            f"Kit: {cells['kit']}" if cells['kit'] else '',
            f"Item: {cells['item']}" if cells['item'] else '',
            f"Equipment: {cells['equipment']}" if cells['equipment'] else '',
            f"Location: {cells['board_location']}" if cells['board_location'] else '',
            f"Colour: {cells['colour']}" if cells['colour'] else '',
            f"Q-ty/item: {cells['qty_item']}" if cells['qty_item'] else '',
            f"Support: {cells['support']}" if cells['support'] else '',
            f"Lock: {cells['lock_type']}" if cells['lock_type'] else '',
        ]))
        rows.append({'row': index, 'apn': apn, 'description': description, **cells})
    workbook.close()
    return rows, ignored


@api_view(['POST'])
@drf_permission_classes([IsTechStudyOrAdmin])
def project_import_excel(request):
    """POST /api/projects/import-excel/ — multipart : file, project_name, client,
    mode ('preview' | 'commit').

    Crée le projet + un échantillon par ligne (les doublons d'APN sont des
    variantes : chacune garde son propre n° de série et sa description) + une
    entrée de matrice de référence par ligne (non fusionnée, notes = détails).
    """
    from django.db import transaction

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'Aucun fichier fourni.'},
                        status=status.HTTP_400_BAD_REQUEST)
    if not file.name.lower().endswith(('.xlsm', '.xlsx')):
        return Response({'error': 'Le fichier doit être au format Excel (.xlsm ou .xlsx).'},
                        status=status.HTTP_400_BAD_REQUEST)

    project_name = (request.data.get('project_name') or '').strip()
    if not project_name:
        return Response({'error': 'Le nom du projet est requis.'},
                        status=status.HTTP_400_BAD_REQUEST)

    client = (request.data.get('client') or 'Other').strip()
    if client not in _CLIENT_VALUES:
        client = 'Other'

    mode = (request.data.get('mode') or 'preview').strip()
    commentaire_global = (request.data.get('comment') or '').strip()

    try:
        rows, ignored = _parse_board_specification(file)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    if mode != 'commit':
        return Response({
            'project_name': project_name,
            'client': client,
            'total': len(rows),
            'ignored': len(ignored),
            'ignored_details': ignored[:50],
            'rows': rows[:100],          # aperçu limité à 100 lignes
            'truncated': len(rows) > 100,
        })

    # ── Commit ────────────────────────────────────────────────────────────────
    if not rows:
        return Response({'error': 'Aucune ligne valide dans le fichier.'},
                        status=status.HTTP_400_BAD_REQUEST)
    if ProjectValidation.objects.filter(project_name=project_name).exists() \
            or Sample.objects.filter(project=project_name).exists():
        return Response({'error': 'Un projet avec ce nom existe déjà.'},
                        status=status.HTTP_409_CONFLICT)

    with transaction.atomic():
        ProjectValidation.objects.create(project_name=project_name)
        for row in rows:
            # Commentaire = saisie du modal + colonne X « Description / Comments »
            commentaire = '\n'.join(filter(None, [commentaire_global, row['comments']]))
            sample = Sample.objects.create(
                apn=row['apn'],
                project=project_name,
                placement='A1',
                client=client,
                status='pending',
                quantity=1,
                received_date=timezone.localdate(),
                description=row['description'],
                commentaire=commentaire,
                created_by=request.user,
            )
            AuditLog.objects.create(sample=sample, user=request.user, action='import')
            MatrixEntry.objects.create(
                reference=row['apn'],
                designation=row['equipment'][:200],
                quantity=1,
                sample_type=row['equipment'][:100],
                notes=row['description'],
                created_by=request.user,
            )

    return Response({
        'project_name': project_name,
        'client': client,
        'created_samples': len(rows),
        'created_matrix': len(rows),
        'ignored': len(ignored),
        'ignored_details': ignored[:50],
    }, status=status.HTTP_201_CREATED)


@api_view(['GET', 'POST'])
@drf_permission_classes([IsAuthenticated])
def project_samples(request, project_name):
    """GET/POST /api/validation/projects/<project_name>/samples/"""
    if request.method == 'GET':
        samples = Sample.objects.filter(project=project_name).order_by('apn', '-created_at')
        return Response([
            {
                'id': s.id,
                'apn': s.apn,
                'quantity': s.quantity,
                'connector_fill': s.connector_fill,
                'placement': s.placement,
                'client': s.client,
                'serial_number': s.serial_number,
                'description': s.description,
                'commentaire': s.commentaire,
                'study_approved': s.study_approved,
                'created_at': s.created_at.isoformat(),
            }
            for s in samples
        ])

    apn = request.data.get('apn', '').strip()
    if not apn:
        return Response({'error': 'apn est requis.'}, status=status.HTTP_400_BAD_REQUEST)

    placement = request.data.get('placement', 'A1').strip().upper()
    if not PLACEMENT_RE.match(placement):
        return Response(
            {'error': 'Format placement invalide (ex: A1, B12).'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    connector_fill = request.data.get('connector_fill', 'empty')
    if connector_fill not in _CONNECTOR_VALUES:
        connector_fill = 'empty'

    client = request.data.get('client', 'Other')
    if client not in _CLIENT_VALUES:
        client = 'Other'

    try:
        quantity = max(1, int(request.data.get('quantity', 1)))
    except (ValueError, TypeError):
        quantity = 1

    sample = Sample.objects.create(
        apn=apn,
        project=project_name,
        placement=placement,
        client=client,
        connector_fill=connector_fill,
        quantity=quantity,
        created_by=request.user,
    )
    AuditLog.objects.create(sample=sample, user=request.user, action='create')
    return Response({
        'id': sample.id,
        'apn': sample.apn,
        'quantity': sample.quantity,
        'connector_fill': sample.connector_fill,
        'placement': sample.placement,
        'client': sample.client,
        'serial_number': sample.serial_number,
        'created_at': sample.created_at.isoformat(),
    }, status=status.HTTP_201_CREATED)


@api_view(['PATCH'])
@drf_permission_classes([IsTechStudyOrAdmin])
def project_update_status(request, project_name):
    """PATCH /api/validation/projects/<project_name>/status/ — admin manually sets status."""
    try:
        pv = ProjectValidation.objects.get(project_name=project_name)
    except ProjectValidation.DoesNotExist:
        return Response({'error': 'Projet introuvable.'}, status=status.HTTP_404_NOT_FOUND)

    new_status = request.data.get('validation_status', '').strip()
    if new_status not in ('pending', 'rejected'):
        return Response(
            {'error': "validation_status doit être 'pending' ou 'rejected'."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    pv.validation_status = new_status
    pv.approved_at = None
    pv.approved_by = None
    if new_status == 'pending':
        pv.validated_at = None
        pv.validated_by = None
        pv.result = None
    pv.save()

    # Un-approve all samples in this project so they disappear from the dashboard
    Sample.objects.filter(project=project_name).update(study_approved=False)

    return Response({
        'project_name': pv.project_name,
        'validation_status': pv.validation_status,
        'sample_count': Sample.objects.filter(project=pv.project_name).count(),
        'validated_at': pv.validated_at.isoformat() if pv.validated_at else None,
        'validated_by': (pv.validated_by.get_full_name() or pv.validated_by.username) if pv.validated_by else None,
        'approved_at': None,
        'approved_by': None,
    })


@api_view(['PATCH'])
@drf_permission_classes([IsTechStudyOrAdmin])
def project_approve_apn(request, project_name):
    """PATCH /api/validation/projects/<name>/approve-apn/
    body: { apn, approved }  — toggle study_approved on all samples with that APN in the project.
    """
    apn = request.data.get('apn', '').strip()
    if not apn:
        return Response({'error': 'apn est requis.'}, status=status.HTTP_400_BAD_REQUEST)
    approved = bool(request.data.get('approved', True))
    count = Sample.objects.filter(project=project_name, apn=apn).update(study_approved=approved)
    return Response({'apn': apn, 'approved': approved, 'updated': count})


@api_view(['DELETE'])
@drf_permission_classes([IsAuthenticated])
def project_sample_delete(request, project_name, sample_id):
    """DELETE /api/validation/projects/<project_name>/samples/<sample_id>/"""
    try:
        sample = Sample.objects.get(pk=sample_id, project=project_name)
    except Sample.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    sample.soft_delete(user=request.user)
    AuditLog.objects.create(sample=sample, user=request.user, action='delete')
    return Response(status=status.HTTP_204_NO_CONTENT)


# ── Sales module ──────────────────────────────────────────────────────────────

SALES_DEPARTMENT_NAME    = 'Sales'
SALES_EMPLOYEE_POSITION  = 'Sales Employee'
SALES_ASSIGNMENT_ERROR   = (
    'Seuls les employés du département Sales ayant le rôle '
    '« Sales Employee » peuvent être affectés à un projet de vente.'
)


def _sales_employees():
    from hr.models import Employee
    return Employee.objects.select_related('user', 'department').filter(
        department__name__iexact=SALES_DEPARTMENT_NAME,
        position__iexact=SALES_EMPLOYEE_POSITION,
        status=Employee.Status.ACTIVE,
    )


def _validate_sales_owner(value):
    if value in (None, ''):
        return None, 'Un commercial Sales Employee doit être affecté au projet.'
    try:
        employee_id = int(value)
    except (TypeError, ValueError):
        return None, SALES_ASSIGNMENT_ERROR
    employee = _sales_employees().filter(pk=employee_id).first()
    if not employee:
        return None, SALES_ASSIGNMENT_ERROR
    return employee.id, None


def _validate_accounting_client(value):
    if not value:
        return None, 'Le client Comptabilité est obligatoire.'
    try:
        client_id = int(value)
        Tiers.objects.get(pk=client_id)
        return client_id, None
    except (TypeError, ValueError, Tiers.DoesNotExist):
        return None, 'Client introuvable.'


def _attach_project_documents(project, files, user):
    allowed = {'.pdf', '.ppt', '.pptx'}
    for file in files:
        ext = os.path.splitext(file.name)[1].lower()
        if ext not in allowed:
            return f'Format non autorisé pour {file.name}. Utilisez PDF, PPT ou PPTX.'
    for file in files:
        ProjectDocument.objects.create(project=project, file=file, file_name=file.name, uploaded_by=user)
    return None


def _can_manage_sales_projects(user):
    return bool(
        user.is_staff or user.is_superuser
        or user.groups.filter(name__in=['Sales Manager', 'Sales Employee']).exists()
    )


def _can_manage_sales_targets(user):
    return bool(
        user.is_staff or user.is_superuser
        or user.groups.filter(name='Sales Manager').exists()
    )


def _current_sales_employee(user):
    if not user or not user.is_authenticated:
        return None
    return _sales_employees().filter(user=user).first()


def _ensure_project_for_won_opportunity(record, actor):
    if record.record_type != 'opportunity' or record.status != 'won' or record.project_id:
        return record
    project_name = (record.title or '').strip()[:100] or f'Projet {record.code}'
    project, created = ProjectValidation.objects.get_or_create(
        project_name=project_name,
        defaults={
            'sales_reference':    f'PRJ-{timezone.now():%Y%m%d%H%M%S%f}',
            'accounting_client_id': record.accounting_client_id,
            'customer_name': record.accounting_client.raison_sociale if record.accounting_client else record.company_name,
            'sales_owner_id':     record.assigned_employee_id,
            'sales_stage':        'won',
            'estimated_value':    record.value,
            'probability':        record.probability,
            'actual_close_date':  record.due_date or timezone.localdate(),
            'expected_close_date': record.due_date,
            'sales_notes':        record.notes,
        },
    )
    if not created:
        updated = False
        if not project.sales_owner_id and record.assigned_employee_id:
            project.sales_owner_id = record.assigned_employee_id; updated = True
        if not project.accounting_client_id and record.accounting_client_id:
            project.accounting_client_id = record.accounting_client_id; updated = True
        if record.value and not project.estimated_value:
            project.estimated_value = record.value; updated = True
        if updated:
            project.save()
    record.project = project
    record.save(update_fields=['project', 'updated_at'])
    SalesProjectHistory.objects.create(project=project, actor=actor, new_status=project.sales_stage)
    return record


def _sales_project_payload(project, request=None):
    samples = list(Sample.objects.filter(project=project.project_name).order_by('apn'))
    total   = len(samples)

    def done_count(field):
        return sum(bool(getattr(s, field)) for s in samples)

    stages = {
        'design':      done_count('is_done'),
        'programming': done_count('programmer_done'),
        'cnc':         done_count('cnc_done'),
        'assembly':    done_count('assembly_done'),
        'quality':     done_count('quality_done'),
    }
    completed = stages['quality']
    progress  = round((completed / total) * 100) if total else 0
    clients   = sorted({s.client for s in samples if s.client})
    today     = timezone.localdate()
    return {
        'id':                  project.id,
        'sales_reference':     project.sales_reference,
        'project_name':        project.project_name,
        'accounting_client':   project.accounting_client_id,
        'customer_name':       project.customer_name or ', '.join(clients),
        'customer_contact':    project.customer_contact,
        'customer_email':      project.customer_email,
        'customer_phone':      project.customer_phone,
        'sales_stage':         project.sales_stage,
        'sales_stage_display': project.get_sales_stage_display(),
        'sales_priority':      project.sales_priority,
        'sales_priority_display': project.get_sales_priority_display(),
        'sales_owner':         project.sales_owner_id,
        'sales_owner_name':    project.sales_owner.full_name if project.sales_owner else None,
        'estimated_value':     project.estimated_value,
        'probability':         project.probability,
        'weighted_value':      project.estimated_value * project.probability / 100,
        'expected_close_date': project.expected_close_date,
        'actual_close_date':   project.actual_close_date,
        'delivery_target_date': project.delivery_target_date,
        'next_action':         project.next_action,
        'next_action_date':    project.next_action_date,
        'sales_notes':         project.sales_notes,
        'validation_status':   project.validation_status,
        'approved_at':         project.approved_at,
        'sample_count':        total,
        'completed_samples':   completed,
        'progress':            progress,
        'stage_progress':      stages,
        'overdue_follow_up':   bool(project.next_action_date and project.next_action_date < today and project.sales_stage not in ('won', 'lost')),
        'delivery_at_risk':    bool(project.delivery_target_date and project.delivery_target_date < today and completed < total),
        'samples': [
            {'id': s.id, 'apn': s.apn, 'placement': s.placement, 'quantity': s.quantity,
             'status': s.status, 'design_done': s.is_done, 'programming_done': s.programmer_done,
             'cnc_done': s.cnc_done, 'assembly_done': s.assembly_done, 'quality_done': s.quality_done}
            for s in samples
        ],
        'documents': [
            {'id': d.id, 'file_name': d.file_name,
             'url': request.build_absolute_uri(d.file.url) if request and d.file else (d.file.url if d.file else None),
             'uploaded_at': d.uploaded_at,
             'uploaded_by': d.uploaded_by.get_full_name() or d.uploaded_by.username if d.uploaded_by else None}
            for d in project.documents.all()
        ],
        'created_at': project.created_at,
        'updated_at': project.updated_at,
        'history': [
            {'id': h.id, 'old_status': h.old_status, 'new_status': h.new_status,
             'actor': h.actor.get_full_name() or h.actor.username if h.actor else None,
             'created_at': h.created_at}
            for h in project.sales_history.all()
        ],
    }


def _sales_period_bounds(period='month', year=None, month=None, quarter=None, start_date=None, end_date=None):
    today  = timezone.localdate()
    year   = int(year or today.year)
    month  = int(month or today.month)
    if period == 'year':
        return date(year, 1, 1), date(year, 12, 31)
    if period == 'quarter':
        quarter     = int(quarter or ((today.month - 1) // 3 + 1))
        first_month = (quarter - 1) * 3 + 1
        start       = date(year, first_month, 1)
        return (start, date(year, 12, 31)) if quarter == 4 else (start, date(year, first_month + 3, 1) - timedelta(days=1))
    if period == 'custom' and start_date and end_date:
        return start_date, end_date
    start = date(year, month, 1)
    end   = date(year, 12, 31) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def _sales_target_scope_from_request(request):
    period    = request.query_params.get('period') or request.data.get('period') or 'month'
    year      = request.query_params.get('year') or request.data.get('year')
    month     = request.query_params.get('month') or request.data.get('month')
    quarter   = request.query_params.get('quarter') or request.data.get('quarter')
    start_raw = request.query_params.get('start_date') or request.data.get('start_date')
    end_raw   = request.query_params.get('end_date') or request.data.get('end_date')
    start_date = date.fromisoformat(start_raw) if start_raw else None
    end_date   = date.fromisoformat(end_raw) if end_raw else None
    start, end = _sales_period_bounds(period, year, month, quarter, start_date, end_date)
    return {
        'period':  period,
        'year':    start.year if period == 'custom' else int(year or start.year),
        'month':   int(month) if month else (start.month if period == 'month' else None),
        'quarter': int(quarter) if quarter else (((start.month - 1) // 3 + 1) if period == 'quarter' else None),
        'start_date': start,
        'end_date':   end,
    }


def _sales_target_label(target):
    month_names = {
        1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
        5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
        9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre',
    }
    if target.period_type == SalesTarget.PERIOD_YEAR:
        return f'Année {target.year}'
    if target.period_type == SalesTarget.PERIOD_QUARTER:
        return f'T{target.quarter} {target.year}'
    if target.period_type == SalesTarget.PERIOD_CUSTOM and target.start_date and target.end_date:
        return f'{target.start_date:%d/%m/%Y} - {target.end_date:%d/%m/%Y}'
    return f'{month_names.get(target.month, target.month)} {target.year}'


def _sales_target_queryset(scope, employee_ids=None):
    qs = SalesTarget.objects.select_related('employee', 'employee__department', 'employee__user')
    period = scope['period']
    if period == SalesTarget.PERIOD_CUSTOM:
        qs = qs.filter(period_type=SalesTarget.PERIOD_CUSTOM, start_date=scope['start_date'], end_date=scope['end_date'])
    else:
        qs = qs.filter(period_type=period, year=scope['year'])
        if period == SalesTarget.PERIOD_MONTH:
            qs = qs.filter(month=scope['month'])
        elif period == SalesTarget.PERIOD_QUARTER:
            qs = qs.filter(quarter=scope['quarter'])
    if employee_ids is not None:
        qs = qs.filter(employee_id__in=employee_ids)
    return qs


def _sales_target_summary(scope, request_user):
    employees = list(_sales_employees().annotate(
        ongoing_projects=Count('sales_projects', filter=~Q(sales_projects__sales_stage__in=['won', 'lost'])),
    ).order_by('last_name', 'first_name'))
    if request_user.groups.filter(name='Sales Employee').exists():
        employees = [e for e in employees if e.user_id == request_user.id]
    employee_ids = [e.id for e in employees]
    target_map   = {t.employee_id: t for t in _sales_target_queryset(scope, employee_ids)}
    opportunity_qs = SalesRecord.objects.filter(record_type='opportunity', assigned_employee_id__in=employee_ids)
    won_qs = opportunity_qs.filter(status='won').filter(
        updated_at__date__gte=scope['start_date'], updated_at__date__lte=scope['end_date'],
    )
    projects_qs = ProjectValidation.objects.filter(
        sales_owner_id__in=employee_ids,
        created_at__date__gte=scope['start_date'],
        created_at__date__lte=scope['end_date'],
    )
    projects_counts = {item['sales_owner_id']: item['count'] for item in projects_qs.values('sales_owner_id').annotate(count=Count('id'))}
    won_counts  = {item['assigned_employee_id']: item['count'] for item in won_qs.values('assigned_employee_id').annotate(count=Count('id'))}
    won_revenue = {item['assigned_employee_id']: float(item['total'] or 0) for item in won_qs.values('assigned_employee_id').annotate(total=Sum('value'))}
    rows = []
    for employee in employees:
        target        = target_map.get(employee.id)
        target_amount = float(target.target_amount) if target else 0
        achieved      = won_revenue.get(employee.id, 0)
        rate          = round((achieved / target_amount) * 100, 2) if target_amount else 0
        rows.append({
            'employee_id':      employee.id,
            'employee_name':    employee.full_name,
            'department':       employee.department.name if employee.department else '',
            'target_id':        target.id if target else None,
            'target_label':     _sales_target_label(target) if target else None,
            'target_amount':    target_amount,
            'achieved_amount':  achieved,
            'achievement_rate': rate,
            'won_opportunities': won_counts.get(employee.id, 0),
            'projects_created': projects_counts.get(employee.id, 0),
            'active_projects':  employee.ongoing_projects,
        })
    return rows


def _sales_export_filters(request, extra=None):
    parts = []
    for key, label in (
        ('search', 'Recherche'), ('status', 'Statut'), ('stage', 'Etape'),
        ('period', 'Periode'), ('year', 'Annee'), ('month', 'Mois'), ('quarter', 'Trimestre'),
    ):
        value = request.query_params.get(key)
        if value:
            parts.append(f'{label}: {value}')
    if extra:
        parts.extend(extra)
    return ' | '.join(parts) if parts else 'Aucun'


def _sales_excel(filename, sheet_name, headers, rows, filters=None):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if filters:
        ws.append([f'Filtres: {filters}'])
        ws.append([])
    header_row = ws.max_row + 1
    ws.append(headers)
    for cell in ws[header_row]:
        cell.fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(['GET', 'POST'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_projects(request):
    if request.method == 'POST':
        if not _can_manage_sales_projects(request.user):
            return Response({'error': 'Seuls les utilisateurs Sales peuvent créer des projets.'}, status=403)
        name = str(request.data.get('project_name', '')).strip()
        if not name:
            return Response({'project_name': 'Project title is required.'}, status=400)
        if ProjectValidation.objects.filter(project_name=name).exists():
            return Response({'project_name': 'A project with this title already exists.'}, status=400)
        sales_employee = _current_sales_employee(request.user)
        sales_owner_value = request.data.get('sales_owner')
        if sales_employee and not sales_owner_value:
            sales_owner_id = sales_employee.id
        else:
            sales_owner_id, assignment_error = _validate_sales_owner(sales_owner_value)
            if assignment_error:
                return Response({'sales_owner': assignment_error}, status=400)
        accounting_client_id, client_error = _validate_accounting_client(request.data.get('accounting_client'))
        if client_error:
            return Response({'accounting_client': client_error}, status=400)
        project = ProjectValidation.objects.create(
            project_name=name,
            sales_reference=f'PRJ-{timezone.now():%Y%m%d%H%M%S%f}',
            accounting_client_id=accounting_client_id,
            customer_name=str(request.data.get('customer_name', '')).strip(),
            sales_owner_id=sales_owner_id,
            sales_stage=request.data.get('sales_stage', 'prospect'),
            estimated_value=request.data.get('estimated_value') or 0,
            expected_close_date=request.data.get('expected_close_date') or None,
            actual_close_date=request.data.get('actual_close_date') or None,
            sales_notes=str(request.data.get('description', '')).strip(),
        )
        attachment_error = _attach_project_documents(project, request.FILES.getlist('documents'), request.user)
        if attachment_error:
            project.delete()
            return Response({'documents': attachment_error}, status=400)
        SalesProjectHistory.objects.create(project=project, actor=request.user, new_status=project.sales_stage)
        return Response(_sales_project_payload(project, request), status=201)

    sample_names = Sample.objects.values_list('project', flat=True).distinct()
    existing = set(ProjectValidation.objects.filter(project_name__in=sample_names).values_list('project_name', flat=True))
    ProjectValidation.objects.bulk_create(
        [ProjectValidation(project_name=n) for n in sample_names if n and n not in existing],
        ignore_conflicts=True,
    )
    projects = ProjectValidation.objects.select_related('sales_owner', 'accounting_client').prefetch_related('sales_history__actor', 'documents__uploaded_by').all()
    if request.user.groups.filter(name='Sales Employee').exists():
        projects = projects.filter(sales_owner__user=request.user)
    search = request.query_params.get('search', '').strip()
    stage  = request.query_params.get('stage', '').strip()
    if search:
        projects = projects.filter(Q(project_name__icontains=search) | Q(customer_name__icontains=search) | Q(customer_contact__icontains=search))
    if stage:
        projects = projects.filter(sales_stage=stage)
    return Response([_sales_project_payload(p, request) for p in projects])


@api_view(['PATCH', 'DELETE'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_project_update(request, pk):
    try:
        project = ProjectValidation.objects.select_related('sales_owner').prefetch_related('documents__uploaded_by').get(pk=pk)
    except ProjectValidation.DoesNotExist:
        return Response({'error': 'Project not found.'}, status=404)
    if not _can_manage_sales_projects(request.user):
        return Response({'error': 'Seuls les utilisateurs Sales peuvent modifier ou supprimer des projets.'}, status=403)
    if request.method == 'DELETE':
        if project.sales_records.exists() or Sample.objects.filter(project=project.project_name).exists():
            return Response({'error': 'Ce projet ne peut pas être supprimé car il contient des données liées.'}, status=400)
        project.delete()
        return Response(status=204)
    old_status = project.sales_stage
    editable = {
        'customer_name', 'customer_contact', 'customer_email', 'customer_phone',
        'sales_stage', 'sales_priority', 'sales_owner', 'estimated_value', 'probability',
        'expected_close_date', 'delivery_target_date', 'next_action', 'next_action_date',
        'sales_notes', 'accounting_client', 'actual_close_date',
    }
    nullable_fields = {'sales_owner', 'expected_close_date', 'delivery_target_date', 'next_action_date', 'accounting_client', 'actual_close_date'}
    sales_owner_id = None
    accounting_client_id = None
    if 'sales_owner' in request.data:
        sales_owner_id, assignment_error = _validate_sales_owner(request.data.get('sales_owner'))
        if assignment_error:
            return Response({'sales_owner': assignment_error}, status=400)
    if 'accounting_client' in request.data:
        accounting_client_id, client_error = _validate_accounting_client(request.data.get('accounting_client'))
        if client_error:
            return Response({'accounting_client': client_error}, status=400)
    for field in editable:
        if field in request.data:
            value = request.data[field]
            if field == 'sales_owner':
                project.sales_owner_id = sales_owner_id
            elif field == 'accounting_client':
                project.accounting_client_id = accounting_client_id
            else:
                setattr(project, field, None if field in nullable_fields and value == '' else value)
    if not project.accounting_client_id:
        return Response({'accounting_client': 'Le client Comptabilité est obligatoire.'}, status=400)
    project.save()
    attachment_error = _attach_project_documents(project, request.FILES.getlist('documents'), request.user)
    if attachment_error:
        return Response({'documents': attachment_error}, status=400)
    if old_status != project.sales_stage:
        SalesProjectHistory.objects.create(project=project, actor=request.user, old_status=old_status, new_status=project.sales_stage)
    project.refresh_from_db()
    return Response(_sales_project_payload(project, request))


@api_view(['DELETE'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_project_document_delete(request, pk, document_id):
    if not _can_manage_sales_projects(request.user):
        return Response({'error': 'Seuls les utilisateurs Sales peuvent supprimer des documents.'}, status=403)
    try:
        document = ProjectDocument.objects.get(pk=document_id, project_id=pk)
    except ProjectDocument.DoesNotExist:
        return Response({'error': 'Document introuvable.'}, status=404)
    document.delete()
    return Response(status=204)


@api_view(['GET'])
@drf_permission_classes([IsSalesOrAdmin])
def salespeople(request):
    employees = _sales_employees().annotate(
        ongoing_projects=Count('sales_projects', filter=~Q(sales_projects__sales_stage__in=['won', 'lost'])),
    ).order_by('ongoing_projects', 'last_name', 'first_name')
    return Response([
        {'id': e.id, 'employee_id': e.id, 'name': e.full_name,
         'department': e.department.name if e.department else None,
         'position': e.position, 'ongoing_projects': e.ongoing_projects}
        for e in employees
    ])


@api_view(['GET', 'POST'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_targets(request):
    scope = _sales_target_scope_from_request(request)
    if request.method == 'POST':
        if not _can_manage_sales_targets(request.user):
            return Response({'error': 'Seuls les managers Sales peuvent définir des objectifs.'}, status=403)
        target_amount = Decimal(str(request.data.get('target_amount') or '0'))
        employee_id   = request.data.get('employee_id')
        apply_all     = str(request.data.get('apply_all', '')).lower() in ('1', 'true', 'yes')
        if str(employee_id) == 'all':
            apply_all  = True
            employee_id = ''
        employees = list(_sales_employees())
        if not apply_all and employee_id:
            employees = [e for e in employees if str(e.id) == str(employee_id)]
        if not apply_all and not employee_id:
            return Response({'employee_id': "Sélectionnez un employé ou activez l'application globale."}, status=400)
        if not employees:
            return Response({'employee_id': 'Aucun employé valide sélectionné.'}, status=400)
        created = []
        for employee in employees:
            target, _ = SalesTarget.objects.update_or_create(
                employee=employee,
                period_type=scope['period'],
                year=scope['year'],
                month=scope['month'],
                quarter=scope['quarter'],
                start_date=scope['start_date'] if scope['period'] == SalesTarget.PERIOD_CUSTOM else None,
                end_date=scope['end_date'] if scope['period'] == SalesTarget.PERIOD_CUSTOM else None,
                defaults={'target_amount': target_amount, 'created_by': request.user, 'updated_by': request.user},
            )
            created.append(target.id)
        return Response({'updated_targets': created, 'count': len(created)})
    rows = _sales_target_summary(scope, request.user)
    return Response({
        'period': scope['period'], 'year': scope['year'], 'month': scope['month'],
        'quarter': scope['quarter'], 'start_date': scope['start_date'], 'end_date': scope['end_date'],
        'rows': rows,
    })


@api_view(['GET'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_targets_export(request):
    scope = _sales_target_scope_from_request(request)
    rows  = _sales_target_summary(scope, request.user)
    return _sales_excel(
        'objectifs_commerciaux.xlsx', 'Objectifs',
        ['Employé', 'Objectif EUR', 'Réalisé EUR', "Taux d'atteinte", 'Opportunités gagnées', 'Projets créés'],
        [(r['employee_name'], r['target_amount'], r['achieved_amount'], f"{r['achievement_rate']}%", r['won_opportunities'], r['projects_created']) for r in rows],
        filters=_sales_export_filters(request),
    )


@api_view(['GET'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_projects_export(request):
    projects = ProjectValidation.objects.select_related('accounting_client', 'sales_owner')
    search = request.query_params.get('search', '').strip()
    if search:
        projects = projects.filter(Q(project_name__icontains=search) | Q(customer_name__icontains=search) | Q(sales_reference__icontains=search))
    rows = [
        (p.sales_reference, p.project_name,
         p.accounting_client.raison_sociale if p.accounting_client else '',
         p.sales_owner.full_name if p.sales_owner else '',
         f'EUR {p.estimated_value}', p.expected_close_date, p.actual_close_date,
         p.get_sales_stage_display())
        for p in projects
    ]
    return _sales_excel(
        'projets_ventes.xlsx', 'Projets',
        ['Référence', 'Projet', 'Client', 'Commercial', 'Valeur estimée', 'Clôture prévue', 'Clôture réelle', 'Statut'],
        rows, filters=_sales_export_filters(request),
    )


@api_view(['GET'])
@drf_permission_classes([IsSalesOrAdmin])
def sales_opportunities_export(request):
    opportunities = SalesRecord.objects.filter(record_type='opportunity').select_related('accounting_client', 'assigned_employee', 'project')
    search = request.query_params.get('search', '').strip()
    if search:
        opportunities = opportunities.filter(Q(code__icontains=search) | Q(title__icontains=search) | Q(accounting_client__raison_sociale__icontains=search))
    rows = [
        (x.code, x.title,
         x.accounting_client.raison_sociale if x.accounting_client else '',
         x.project.project_name if x.project else '',
         x.assigned_employee.full_name if x.assigned_employee else '',
         f'EUR {x.value}', f'{x.probability}%', x.due_date, x.status)
        for x in opportunities
    ]
    return _sales_excel(
        'opportunites_ventes.xlsx', 'Opportunités',
        ['Référence', 'Titre', 'Client', 'Projet', 'Commercial', 'Valeur', 'Probabilité', 'Échéance', 'Statut'],
        rows, filters=_sales_export_filters(request),
    )


class SalesRecordPagination(PageNumberPagination):
    page_size             = 25
    page_size_query_param = 'page_size'
    max_page_size         = 200


class SalesRecordViewSet(viewsets.ModelViewSet):
    permission_classes = [IsSalesOrAdmin]
    serializer_class   = SalesRecordSerializer
    pagination_class   = SalesRecordPagination

    def get_queryset(self):
        qs = SalesRecord.objects.select_related(
            'assigned_to', 'assigned_employee', 'accounting_client', 'lead', 'opportunity', 'project',
        ).prefetch_related('history')
        for field in ('record_type', 'status', 'industry', 'assigned_to', 'assigned_employee', 'accounting_client', 'project'):
            value = self.request.query_params.get(field)
            if value:
                qs = qs.filter(**{field: value})
        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(code__icontains=search) | Q(title__icontains=search) | Q(company_name__icontains=search) | Q(contact_person__icontains=search))
        date_from = self.request.query_params.get('date_from')
        date_to   = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if self.request.user.groups.filter(name='Sales Employee').exists():
            qs = qs.filter(assigned_employee__user=self.request.user)
        return qs

    def perform_create(self, serializer):
        prefix = serializer.validated_data['record_type'][:3].upper()
        code   = f'{prefix}-{timezone.now():%Y%m%d%H%M%S%f}'
        sales_employee = _current_sales_employee(self.request.user)
        with transaction.atomic():
            obj = serializer.save(
                code=code, created_by=self.request.user, updated_by=self.request.user,
                assigned_employee=serializer.validated_data.get('assigned_employee') or sales_employee,
            )
            obj = _ensure_project_for_won_opportunity(obj, self.request.user)
            SalesAuditLog.objects.create(record=obj, actor=self.request.user, action='create', changes=dict(SalesRecordSerializer(obj).data))

    def perform_update(self, serializer):
        sales_employee    = _current_sales_employee(self.request.user)
        assigned_employee = serializer.validated_data.get('assigned_employee')
        extra = {'updated_by': self.request.user}
        if self.request.user.groups.filter(name='Sales Employee').exists() and sales_employee:
            extra['assigned_employee'] = assigned_employee or sales_employee
        obj = serializer.save(**extra)
        obj = _ensure_project_for_won_opportunity(obj, self.request.user)
        SalesAuditLog.objects.create(record=obj, actor=self.request.user, action='update', changes=dict(SalesRecordSerializer(obj).data))

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.downstream_records.exists() or obj.project_id:
            return Response({'error': 'Ce client contient des données liées et ne peut pas être supprimé.'}, status=400)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='reports')
    def reports(self, request):
        qs       = self.get_queryset()
        client   = request.query_params.get('client')
        employee = request.query_params.get('employee')
        if client:
            qs = qs.filter(accounting_client_id=client)
        if employee:
            qs = qs.filter(assigned_employee_id=employee)
        projects = ProjectValidation.objects.all()
        if employee:
            projects = projects.filter(sales_owner_id=employee)
        target_scope = _sales_target_scope_from_request(request)
        target_rows  = _sales_target_summary(target_scope, request.user)
        return Response({
            'revenue_by_client': list(qs.filter(record_type='opportunity', status='won').values('accounting_client__raison_sociale').annotate(value=Sum('value')).order_by('-value')[:20]),
            'revenue_by_employee': list(qs.filter(record_type='opportunity', status='won').values('assigned_employee__first_name', 'assigned_employee__last_name').annotate(value=Sum('value')).order_by('-value')[:20]),
            'project_status': list(projects.values('sales_stage').annotate(count=Count('id'))),
            'won':           projects.filter(sales_stage='won').count(),
            'lost':          projects.filter(sales_stage='lost').count(),
            'total_revenue': sum(float(r.value) for r in qs.filter(record_type='opportunity', status='won')),
            'pipeline_value': sum(float(r.value) for r in qs.filter(record_type='opportunity').exclude(status__in=['won', 'lost'])),
            'targets': target_rows,
        })

    @action(detail=False, methods=['get'], url_path='dashboard')
    def dashboard(self, request):
        qs            = self.get_queryset()
        opportunities = qs.filter(record_type='opportunity')
        projects_qs   = ProjectValidation.objects.all()
        if request.user.groups.filter(name='Sales Employee').exists():
            projects_qs = projects_qs.filter(sales_owner__user=request.user)
        won         = opportunities.filter(status='won')
        total_closed = won.count() + opportunities.filter(status='lost').count()
        scope        = _sales_target_scope_from_request(request)
        target_rows  = _sales_target_summary(scope, request.user)
        target_total  = sum(r['target_amount'] for r in target_rows)
        achieved_total = sum(r['achieved_amount'] for r in target_rows)
        meetings = list(qs.filter(record_type='activity').exclude(due_date__isnull=True).order_by('due_date', '-updated_at').values('id', 'title', 'due_date', 'status')[:8])
        followups = list(projects_qs.exclude(next_action_date__isnull=True).exclude(sales_stage__in=['won', 'lost']).filter(next_action_date__gte=scope['start_date'], next_action_date__lte=scope['end_date']).order_by('next_action_date').values('id', 'project_name', 'next_action', 'next_action_date')[:8])
        notifications = []
        overdue_count = qs.exclude(due_date__isnull=True).exclude(status__in=['won', 'lost', 'completed']).filter(due_date__lt=timezone.localdate()).count()
        if overdue_count:
            notifications.append({'type': 'warning', 'message': f'{overdue_count} activités ou opportunités en retard.'})
        due_projects = projects_qs.exclude(delivery_target_date__isnull=True).exclude(sales_stage__in=['won', 'lost']).filter(delivery_target_date__lte=timezone.localdate() + timedelta(days=7)).count()
        if due_projects:
            notifications.append({'type': 'info', 'message': f'{due_projects} projets approchent de leur échéance.'})
        return Response({
            'active_projects':    projects_qs.exclude(sales_stage__in=['won', 'lost']).count(),
            'open_opportunities': opportunities.exclude(status__in=['won', 'lost']).count(),
            'conversion_rate':    round(won.count() * 100 / total_closed, 2) if total_closed else 0,
            'revenue_forecast':   sum(float(x.value) * x.probability / 100 for x in opportunities.exclude(status__in=['won', 'lost'])),
            'pipeline':           list(opportunities.values('status').annotate(count=Count('id'), value=Sum('value'))),
            'at_risk':            qs.filter(due_date__lt=timezone.localdate()).exclude(status__in=['completed', 'won', 'lost']).count(),
            'total_sales_this_month': won.filter(updated_at__month=timezone.localdate().month, updated_at__year=timezone.localdate().year).count(),
            'revenue_generated':  sum(float(r.value) for r in won.filter(updated_at__month=timezone.localdate().month, updated_at__year=timezone.localdate().year)),
            'target_amount':      target_total,
            'target_achieved':    achieved_total,
            'target_rate':        round((achieved_total / target_total) * 100, 2) if target_total else 0,
            'target_rows':        target_rows,
            'upcoming_meetings':  meetings,
            'upcoming_followups': followups,
            'notifications':      notifications,
        })
