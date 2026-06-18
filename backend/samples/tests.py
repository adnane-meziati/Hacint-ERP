from io import BytesIO

from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from rest_framework.test import APIClient

from .models import MatrixEntry, ProjectValidation, Sample
from .views import _run_comparison

SHEET = 'Board Specification'
HEADERS = ['Status', 'Item', 'Equipment', 'Kit name',
           'Component APN', 'Customer ID', 'Holder APN-ID']


def _make_excel(rows, sheet_name=SHEET, filename='board.xlsm'):
    """Construit un classeur de test : en-têtes ligne 11, données ligne 13+."""
    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for col, header in enumerate(HEADERS, start=9):     # I..O
        sheet.cell(row=11, column=col, value=header)
    for index, row in enumerate(rows, start=13):
        for col, value in enumerate(row, start=9):
            sheet.cell(row=index, column=col, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(filename, buffer.getvalue())


# (Status, Item, Equipment, Kit, Component APN, Customer ID, Holder APN-ID)
LIGNES_DEMO = [
    ('OK', 'Metallic Inlet',   'Holder',    'KIT-A', 'CAPN-100', 'CUST-1', 'HOLD-001'),
    ('OK', 'Wi-Fi detection',  'Holder',    'KIT-A', 'CAPN-101', 'CUST-2', 'HOLD-001'),
    ('',   '',                 'Connector', 'KIT-B', 'CAPN-200', '',       'HOLD-002'),
    ('',   '',                 '',          '',      '',         '',       ''),          # vide → silencieuse
    ('NOK', 'Sans référence',  'Holder',    'KIT-C', '',         'CUST-9', ''),          # M+O vides → ignorée
    ('',   '',                 'Inlet',     '',      'CAPN-300', 'CUST-3', ''),          # O vide → APN = M
]


class ImportExcelTests(TestCase):
    def setUp(self):
        self.client_api = APIClient()
        self.etude = User.objects.create_user('etude', password='x')
        self.etude.groups.add(Group.objects.get_or_create(name='Etude Technique')[0])
        self.client_api.force_authenticate(self.etude)
        self.url = '/api/projects/import-excel/'

    def _post(self, mode, rows=LIGNES_DEMO, project='PRJ-IMPORT', **extra):
        return self.client_api.post(self.url, {
            'file': _make_excel(rows), 'project_name': project,
            'client': 'Aptiv', 'mode': mode, **extra,
        }, format='multipart')

    def test_preview_compte_et_lignes(self):
        response = self._post('preview')
        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.data['total'], 4)        # 2 HOLD-001 + HOLD-002 + CAPN-300
        self.assertEqual(response.data['ignored'], 1)      # ligne M+O vides
        self.assertEqual(Sample.objects.count(), 0)        # aucun écrit en preview
        self.assertEqual(ProjectValidation.objects.count(), 0)
        premiere = response.data['rows'][0]
        self.assertEqual(premiere['apn'], 'HOLD-001')
        self.assertIn('Component APN: CAPN-100', premiere['description'])
        self.assertIn('Item: Metallic Inlet', premiere['description'])

    def test_commit_cree_projet_echantillons_et_matrice(self):
        response = self._post('commit')
        self.assertEqual(response.status_code, 201, response.content)
        self.assertEqual(response.data['created_samples'], 4)
        self.assertEqual(response.data['created_matrix'], 4)
        self.assertEqual(response.data['ignored'], 1)

        self.assertTrue(ProjectValidation.objects.filter(project_name='PRJ-IMPORT').exists())
        samples = Sample.objects.filter(project='PRJ-IMPORT')
        self.assertEqual(samples.count(), 4)
        self.assertTrue(all(s.status == 'pending' and s.quantity == 1 for s in samples))
        self.assertTrue(all(s.client == 'Aptiv' for s in samples))

        # Doublons HOLD-001 : 2 échantillons distincts, numéros de série uniques
        variantes = list(samples.filter(apn='HOLD-001'))
        self.assertEqual(len(variantes), 2)
        self.assertNotEqual(variantes[0].serial_number, variantes[1].serial_number)
        descriptions = {v.description for v in variantes}
        self.assertEqual(len(descriptions), 2)             # variantes distinguables
        self.assertTrue(any('Metallic Inlet' in d for d in descriptions))
        self.assertTrue(any('Wi-Fi detection' in d for d in descriptions))

        # APN de repli = Component APN quand Holder APN-ID est vide
        self.assertTrue(samples.filter(apn='CAPN-300').exists())

        # Matrice : une entrée par ligne, non fusionnée, notes = détails
        entrees = MatrixEntry.objects.filter(reference='HOLD-001')
        self.assertEqual(entrees.count(), 2)
        self.assertTrue(all(e.quantity == 1 for e in entrees))
        self.assertTrue(all(e.sample_type == 'Holder' for e in entrees))
        self.assertTrue(any('Wi-Fi detection' in e.notes for e in entrees))

    def test_projet_existant_refuse(self):
        ProjectValidation.objects.create(project_name='PRJ-IMPORT')
        response = self._post('commit')
        self.assertEqual(response.status_code, 409)
        self.assertEqual(Sample.objects.count(), 0)

    def test_feuille_manquante(self):
        response = self.client_api.post(self.url, {
            'file': _make_excel(LIGNES_DEMO, sheet_name='Autre feuille'),
            'project_name': 'PRJ-X', 'mode': 'preview',
        }, format='multipart')
        self.assertEqual(response.status_code, 400)
        self.assertIn('Board Specification', response.data['error'])

    def test_extension_invalide(self):
        response = self.client_api.post(self.url, {
            'file': SimpleUploadedFile('donnees.csv', b'a,b,c'),
            'project_name': 'PRJ-X', 'mode': 'preview',
        }, format='multipart')
        self.assertEqual(response.status_code, 400)

    def test_permission_refusee_aux_autres_roles(self):
        designer = User.objects.create_user('designer1', password='x')
        designer.groups.add(Group.objects.get_or_create(name='Designer')[0])
        self.client_api.force_authenticate(designer)
        response = self._post('preview')
        self.assertEqual(response.status_code, 403)

    def test_recherche_par_apn_et_serie(self):
        """La recherche renvoie toutes les variantes d'un APN, et par n° série."""
        self._post('commit')
        response = self.client_api.get('/api/samples/', {'search': 'HOLD-001'})
        self.assertEqual(response.data['count'], 2)
        serial = Sample.objects.filter(apn='HOLD-001').first().serial_number
        response = self.client_api.get('/api/samples/', {'search': str(serial)})
        self.assertEqual(response.data['count'], 1)
        # La description (détail de variante) est aussi cherchable
        response = self.client_api.get('/api/samples/', {'search': 'Wi-Fi detection'})
        self.assertEqual(response.data['count'], 1)


def _ligne_reelle(status, item, equipment, kit, comp, cust, holder,
                  holder_type='', alu='', inlet='', coding='', pom='',
                  component_type='', colour='', location='', comments='',
                  qty_item='', support='', lock=''):
    """Ligne au format du fichier client réel : colonnes I..AH (26 valeurs).

    P=7 Component type, T=11 Colour, U=12 Location, X=15 Comments, Y=16 Q-ty,
    Z=17 Holder type, AA=18 Support, AC=20 Lock, AE-AH=22-25 options."""
    row = [None] * 26
    row[0:7] = [status, item, equipment, kit, comp, cust, holder]
    row[7]  = component_type
    row[11] = colour
    row[12] = location
    row[15] = comments
    row[16] = qty_item
    row[17] = holder_type
    row[18] = support
    row[20] = lock
    row[22] = alu
    row[23] = inlet
    row[24] = coding
    row[25] = pom
    return tuple(row)


class ImportExcelVariantesTests(TestCase):
    """Libellés d'item dérivés des colonnes Holder type / options (fichier réel)."""

    def setUp(self):
        self.client_api = APIClient()
        etude = User.objects.create_user('etude2', password='x')
        etude.groups.add(Group.objects.get_or_create(name='Etude Technique')[0])
        self.client_api.force_authenticate(etude)

    def _preview(self, rows):
        response = self.client_api.post('/api/projects/import-excel/', {
            'file': _make_excel(rows), 'project_name': 'PRJ-REEL',
            'client': 'Other', 'mode': 'preview',
        }, format='multipart')
        self.assertEqual(response.status_code, 200, response.content)
        return response.data

    def test_libelles_derives(self):
        data = self._preview([
            _ligne_reelle('Active', 1, 'Holder', '2F4_971_008', '10810760', '', '160100000',
                          holder_type='Wifi'),
            _ligne_reelle('Active', 2, 'Body clip', '2F4_971_008', '13798037', '5M0 971 838',
                          '12W111022', holder_type='WiFi', coding=1),
            _ligne_reelle('Active', 3, 'Holder', '', '15327006', '', '160000000',
                          holder_type='Mechanical'),
            _ligne_reelle('Active', 4, 'Body clip', '', '35102109', '', '12W101010',
                          holder_type='Electrified', coding=1),
            _ligne_reelle('Active', 5, 'Holder', '', '33155189', '', '140100000',
                          holder_type='Electrified', inlet=2, alu=1),
        ])
        items = [r['item'] for r in data['rows']]
        self.assertEqual(items, [
            'Wi-Fi detection (n°1)',
            'Body clip detection (Wi-Fi), Metallic Coding (n°2)',
            'Mechanical (n°3)',
            'Body clip detection (Electrified), Metallic Coding (n°4)',
            'Electrified detection, Metallic Inlet, Aluminium Sliding Part (n°5)',
        ])

    def test_description_et_echantillons_avec_libelles(self):
        response = self.client_api.post('/api/projects/import-excel/', {
            'file': _make_excel([
                _ligne_reelle('Active', 2, 'Body clip', 'KIT-A', '13798037', '5M0',
                              'APN-DUP', holder_type='WiFi', coding=1),
                _ligne_reelle('Active', 7, 'Body clip', 'KIT-A', '13798045', 'N106',
                              'APN-DUP', holder_type='Electrified'),
            ]),
            'project_name': 'PRJ-REEL', 'client': 'Aptiv', 'mode': 'commit',
        }, format='multipart')
        self.assertEqual(response.status_code, 201, response.content)
        variantes = Sample.objects.filter(apn='APN-DUP').order_by('id')
        self.assertEqual(variantes.count(), 2)
        self.assertIn('Item: Body clip detection (Wi-Fi), Metallic Coding (n°2)',
                      variantes[0].description)
        self.assertIn('Item: Body clip detection (Electrified) (n°7)',
                      variantes[1].description)
        # Recherche par libellé de variante
        response = self.client_api.get('/api/samples/', {'search': 'Metallic Coding'})
        self.assertEqual(response.data['count'], 1)

    def test_sans_colonnes_variante_conserve_item_brut(self):
        """Rétro-compatibilité : sans Holder type ni options, Item reste tel quel."""
        data = self._preview([
            ('OK', 'Texte item libre', 'Holder', 'KIT', 'CAPN-1', '', 'H-1'),
        ])
        self.assertEqual(data['rows'][0]['item'], 'Texte item libre')

    def test_component_type_et_details_enrichis(self):
        """Component type en préfixe d'item ; Location/Support/Lock dans la description."""
        data = self._preview([
            _ligne_reelle('Active', 3, 'Holder', '', '15327006', '', '160000000',
                          holder_type='Mechanical',
                          component_type='Ring / Battery Terminal',
                          location='MB62A', qty_item=4,
                          support='Holder support - L', lock='Mechanical - metal'),
            _ligne_reelle('Active', 7, 'Holder', 'TAB_016', '35102458', '3G0 973 708',
                          '151150014', holder_type='Electrified',
                          component_type='Connector', colour='black'),
        ])
        self.assertEqual(data['rows'][0]['item'],
                         'Ring / Battery Terminal — Mechanical (n°3)')
        self.assertEqual(data['rows'][1]['item'],
                         'Connector — Electrified detection (n°7)')
        description = data['rows'][0]['description']
        self.assertIn('Location: MB62A', description)
        self.assertIn('Q-ty/item: 4', description)
        self.assertIn('Support: Holder support - L', description)
        self.assertIn('Lock: Mechanical - metal', description)
        self.assertIn('Colour: black', data['rows'][1]['description'])

    def test_commentaire_import(self):
        """Colonne X « Description / Comments » + commentaire global du modal."""
        response = self.client_api.post('/api/projects/import-excel/', {
            'file': _make_excel([
                _ligne_reelle('Active', 1, 'Body clip', 'KIT', '13798037', '5M0',
                              'APN-C1', holder_type='WiFi',
                              comments="the holder shouldn't damage the BODY CLIP"),
                _ligne_reelle('Active', 2, 'Holder', '', '15327006', '', 'APN-C2'),
            ]),
            'project_name': 'PRJ-COM', 'client': 'Aptiv', 'mode': 'commit',
            'comment': 'Import client ID1 — révision B',
        }, format='multipart')
        self.assertEqual(response.status_code, 201, response.content)
        s1 = Sample.objects.get(apn='APN-C1')
        self.assertIn('Import client ID1 — révision B', s1.commentaire)
        self.assertIn("the holder shouldn't damage the BODY CLIP", s1.commentaire)
        # Le commentaire ne pollue pas la description
        self.assertNotIn('BODY CLIP', s1.description)
        s2 = Sample.objects.get(apn='APN-C2')
        self.assertEqual(s2.commentaire, 'Import client ID1 — révision B')

    def test_lignes_gabarit_silencieuses(self):
        """Les lignes de gabarit (seul le n° d'item prérempli) ne comptent pas
        comme ignorées — contrairement aux lignes avec du contenu sans APN."""
        data = self._preview([
            ('Active', 1, 'Holder', '', 'CAPN-1', '', 'H-1'),   # valide
            ('', 20, '', '', '', '', ''),                       # gabarit → silencieux
            ('Active', 21, '', '', '', '', ''),                 # gabarit → silencieux
            ('', 22, 'Body clip', 'KIT-X', '', 'CUST', ''),     # contenu sans APN → ignorée
        ])
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['ignored'], 1)
        self.assertEqual(data['ignored_details'][0]['row'], 16)


class ValidationAvecDoublonsTests(TestCase):
    """La matrice agrégée par référence doit valider les variantes importées."""

    def test_comparaison_agregee(self):
        for item in ('Metallic Inlet', 'Wi-Fi detection'):
            MatrixEntry.objects.create(
                reference='HOLD-001', designation='Holder', quantity=1,
                sample_type='Holder', notes=f'Item: {item}')
            Sample.objects.create(
                apn='HOLD-001', project='PRJ-V', placement='A1',
                client='Aptiv', quantity=1, description=f'Item: {item}')
        resultat = _run_comparison('PRJ-V')
        self.assertEqual(resultat['validation_status'], 'approved')
        self.assertEqual(len(resultat['matched']), 1)
        self.assertEqual(resultat['matched'][0]['matrix_quantity'], 2)
        # Le type libre « Holder » est informatif — pas de blocage de type
        self.assertEqual(resultat['matched'][0]['matrix_type'], 'Holder')
