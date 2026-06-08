"""
Export live data to legacy OP Excel format.
Usage:  python manage.py export_op <output.xlsx>
        python manage.py export_op op_export.xlsx --from 2026-01-01 --to 2026-12-31

Produces a file with the same column layout as the legacy tracking sheet
(header at row 6, data from row 7).
"""

import io
from datetime import date

from django.core.management.base import BaseCommand, CommandError

from apps.orders.models import OrderLine
from apps.production.models import StageEventStatus


STAGE_SEQUENCE = ["ECH", "CAD", "CAM", "CNC", "MTG", "QF", "AQC"]

HEADERS = [
    "N° OP", "Dt Création", "Dt Livraison", "N° Série",
    "Client", "Réf Article", "Famille", "Désignation",
    "Priorité", "Statut",
    "ECH Dt", "ECH Op", "ECH Obs",
    "CAD Dt", "CAD Op", "CAD Obs",
    "CAM Dt", "CAM Op", "CAM Obs",
    "CNC Dt", "CNC Op", "CNC Obs",
    "MTG Dt", "MTG Op", "MTG Obs",
    "QF Dt",  "QF Op",  "QF Obs",
    "AQC Dt", "AQC Op", "AQC Obs",
]


class Command(BaseCommand):
    help = "Export live OP data to legacy Excel format."

    def add_arguments(self, parser):
        parser.add_argument("output", type=str, help="Output .xlsx file path")
        parser.add_argument("--from", dest="from_date", type=str, help="Delivery date from (YYYY-MM-DD)")
        parser.add_argument("--to", dest="to_date", type=str, help="Delivery date to (YYYY-MM-DD)")

    def handle(self, *args: object, **options: object) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise CommandError("openpyxl is required: pip install openpyxl")

        output_path = options["output"]
        from_date = options.get("from_date")
        to_date = options.get("to_date")

        qs = (
            OrderLine.objects.select_related("order__client", "article__family")
            .prefetch_related("events__stage", "events__completed_by")
            .filter(deleted_at__isnull=True, order__deleted_at__isnull=True)
            .order_by("order__n_ordre", "n_serie")
        )
        if from_date:
            qs = qs.filter(order__delivery_date__gte=from_date)
        if to_date:
            qs = qs.filter(order__delivery_date__lte=to_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OP"

        # Rows 1-5: title block
        ws.cell(row=1, column=1, value="HACINT — Suivi Ordres de Production")
        ws.cell(row=2, column=1, value=f"Exporté le {date.today()}")

        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[6].height = 30

        row_num = 7
        for line in qs:
            events_by_code = {ev.stage.code: ev for ev in line.events.all()}
            order = line.order

            row_data = [
                order.n_ordre,
                order.creation_date,
                order.delivery_date,
                line.n_serie,
                order.client.code,
                line.article.ref_client,
                line.article.family.code,
                line.article.description,
                line.get_priority_display(),
                line.get_status_display(),
            ]

            for stage_code in STAGE_SEQUENCE:
                ev = events_by_code.get(stage_code)
                if ev and ev.status == StageEventStatus.DONE:
                    op_name = ev.completed_by.get_full_name() if ev.completed_by else ""
                    row_data += [
                        ev.completed_at.date() if ev.completed_at else None,
                        op_name,
                        ev.comment,
                    ]
                else:
                    row_data += [None, None, None]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=value)
            row_num += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(
            f"✓ Exported {row_num - 7} lines → {output_path}"
        ))
