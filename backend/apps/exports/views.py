import io
import tempfile
import os
from datetime import date

from django.http import StreamingHttpResponse
from common.permissions import IsPlannerOrAbove, IsAdmin
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.orders.models import Order, OrderLine
from apps.production.models import StageEventStatus

STAGE_SEQUENCE = ["ECH", "CAD", "CAM", "CNC", "MTG", "QF", "AQC"]

# Column positions matching the legacy OP Excel (1-indexed)
# Row 6 = header row.  Columns: A=n_ordre, B=client, C=creation, D=delivery,
# E=n_serie, F=article, G=family, H=description, I=priority, J=status,
# K=ECH_date, L=ECH_op, M=ECH_comment, N=CAD_date, … repeating 3 cols per stage
HEADERS = [
    "N° OP", "Client", "Dt Création", "Dt Livraison",
    "N° Série", "Réf Article", "Famille", "Désignation",
    "Priorité", "Statut",
    "ECH Dt", "ECH Op", "ECH Obs",
    "CAD Dt", "CAD Op", "CAD Obs",
    "CAM Dt", "CAM Op", "CAM Obs",
    "CNC Dt", "CNC Op", "CNC Obs",
    "MTG Dt", "MTG Op", "MTG Obs",
    "QF Dt",  "QF Op",  "QF Obs",
    "AQC Dt", "AQC Op", "AQC Obs",
]


def _build_workbook(queryset) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OP"

    # Header row 6 (rows 1-5 left blank for the legacy title block)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[6].height = 30

    row_num = 7
    for line in queryset:
        events_by_code = {ev.stage.code: ev for ev in line.events.all()}
        order = line.order

        row = [
            order.n_ordre,
            order.client.code,
            order.creation_date,
            order.delivery_date,
            line.n_serie,
            line.article.ref_client,
            line.article.family.code,
            line.article.description,
            line.get_priority_display(),
            line.get_status_display(),
        ]

        for stage_code in STAGE_SEQUENCE:
            ev = events_by_code.get(stage_code)
            if ev and ev.status == StageEventStatus.DONE:
                op_name = ev.completed_by.get_full_name() if ev.completed_by else ""
                row += [ev.completed_at.date() if ev.completed_at else None, op_name, ev.comment]
            else:
                row += [None, None, None]

        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)

        row_num += 1

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExportOPView(APIView):
    """GET /api/exports/op.xlsx?from=YYYY-MM-DD&to=YYYY-MM-DD"""

    permission_classes = [IsPlannerOrAbove]

    def get(self, request: Request) -> StreamingHttpResponse:
        qs = (
            OrderLine.objects.select_related(
                "order__client", "article__family"
            )
            .prefetch_related("events__stage", "events__completed_by")
            .filter(deleted_at__isnull=True, order__deleted_at__isnull=True)
            .order_by("order__n_ordre", "n_serie")
        )

        from_date = request.query_params.get("from")
        to_date = request.query_params.get("to")
        if from_date:
            qs = qs.filter(order__delivery_date__gte=from_date)
        if to_date:
            qs = qs.filter(order__delivery_date__lte=to_date)

        try:
            content = _build_workbook(qs)
        except RuntimeError as exc:
            from django.http import HttpResponse
            return HttpResponse(str(exc), status=501)

        filename = f"OP_export_{date.today()}.xlsx"
        response = StreamingHttpResponse(
            iter([content]),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(content)
        return response


class ImportOPView(APIView):
    """POST /api/imports/op/  — multipart: file=<xlsx>  (admin only)"""

    permission_classes = [IsAdmin]

    def post(self, request: Request) -> Response:
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "No file provided. Send multipart field 'file'."}, status=400)

        if not uploaded.name.lower().endswith((".xlsx", ".xls")):
            return Response({"detail": "Only .xlsx / .xls files are accepted."}, status=400)

        # Save to a temp file so pandas can read it
        suffix = ".xlsx" if uploaded.name.lower().endswith(".xlsx") else ".xls"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            from django.core.management import call_command
            from io import StringIO
            out = StringIO()
            call_command("import_legacy_op", tmp_path, stdout=out, stderr=out)
            output = out.getvalue()
        except Exception as exc:
            return Response({"detail": str(exc)}, status=500)
        finally:
            os.unlink(tmp_path)

        return Response({"detail": "Import completed.", "log": output})
