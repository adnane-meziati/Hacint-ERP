"""Rapport de TVA périodique + exports (relevé des déductions Simpl-TVA, Excel).

Règles appliquées :
- TVA collectée :
    * régime « débits »       → factures de vente validées dans la période
      (les avoirs validés viennent en déduction) ;
    * régime « encaissement » → paiements clients reçus dans la période,
      ventilés au prorata des taux du document (les avoirs validés dans la
      période viennent en déduction à leur date d'émission — approximation
      usuelle des petits outils de gestion).
- TVA déductible : le droit à déduction naît au paiement de la facture
  fournisseur (art. 101-3° du CGI) → paiements fournisseurs de la période,
  ventilés au prorata des taux.
"""
import calendar
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from xml.etree import ElementTree as ET

from .models import MODE_PAIEMENT_DGI, Document, Paiement

TWO_PLACES = Decimal('0.01')


def _round2(value):
    return Decimal(value).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)


def periode_bounds(annee, periode, periodicite):
    """Bornes (incluses) d'une période mensuelle ou trimestrielle."""
    annee, periode = int(annee), int(periode)
    if periodicite == 'mensuelle':
        if not 1 <= periode <= 12:
            raise ValueError('Le mois doit être compris entre 1 et 12.')
        last = calendar.monthrange(annee, periode)[1]
        return date(annee, periode, 1), date(annee, periode, last)
    if not 1 <= periode <= 4:
        raise ValueError('Le trimestre doit être compris entre 1 et 4.')
    premier_mois = (periode - 1) * 3 + 1
    dernier_mois = premier_mois + 2
    last = calendar.monthrange(annee, dernier_mois)[1]
    return date(annee, premier_mois, 1), date(annee, dernier_mois, last)


def _ventilation_paiement(paiement):
    """Ventile un paiement par taux de TVA au prorata du TTC du document.

    Retourne [(taux, base_ht, tva, ttc_part)].
    """
    document = paiement.document
    if document.total_ttc <= 0:
        return []
    facteur = paiement.montant / document.total_ttc
    lignes = []
    for taux, montants in sorted(document.tva_par_taux().items(), reverse=True):
        base = _round2(montants['base'] * facteur)
        tva  = _round2(montants['tva'] * facteur)
        lignes.append((taux, base, tva, _round2(base + tva)))
    return lignes


def _ajouter(detail, taux, base, tva):
    entry = detail.setdefault(taux, {'base': Decimal('0'), 'tva': Decimal('0')})
    entry['base'] += base
    entry['tva']  += tva


def _detail_to_list(detail):
    return [
        {'taux': float(taux), 'base': float(_round2(d['base'])), 'tva': float(_round2(d['tva']))}
        for taux, d in sorted(detail.items(), reverse=True)
    ]


def _paiements_periode(doc_types, date_from, date_to):
    return (
        Paiement.objects
        .filter(document__doc_type__in=doc_types,
                date_paiement__gte=date_from, date_paiement__lte=date_to)
        .select_related('document', 'document__tiers')
        .prefetch_related('document__lignes__tax_code')
        .order_by('date_paiement', 'id')
    )


def _avoirs_periode(doc_type, date_from, date_to):
    return (
        Document.objects
        .filter(doc_type=doc_type, statut='validee',
                date_emission__gte=date_from, date_emission__lte=date_to)
        .prefetch_related('lignes__tax_code')
    )


def rapport_tva(annee, periode, periodicite, regime):
    date_from, date_to = periode_bounds(annee, periode, periodicite)

    # ── TVA collectée ─────────────────────────────────────────────────────────
    collectee = {}
    if regime == 'debit':
        factures = (
            Document.objects
            .filter(doc_type='facture',
                    statut__in=['validee', 'partiellement_payee', 'payee'],
                    date_emission__gte=date_from, date_emission__lte=date_to)
            .prefetch_related('lignes__tax_code')
        )
        for facture in factures:
            for taux, montants in facture.tva_par_taux().items():
                _ajouter(collectee, taux, montants['base'], montants['tva'])
    else:  # encaissement
        for paiement in _paiements_periode(['facture'], date_from, date_to):
            for taux, base, tva, _ in _ventilation_paiement(paiement):
                _ajouter(collectee, taux, base, tva)
    for avoir in _avoirs_periode('avoir', date_from, date_to):
        for taux, montants in avoir.tva_par_taux().items():
            _ajouter(collectee, taux, -montants['base'], -montants['tva'])

    # ── TVA déductible (au paiement — art. 101-3° CGI) ───────────────────────
    deductible = {}
    for paiement in _paiements_periode(['facture_achat'], date_from, date_to):
        for taux, base, tva, _ in _ventilation_paiement(paiement):
            _ajouter(deductible, taux, base, tva)
    for avoir in _avoirs_periode('avoir_achat', date_from, date_to):
        for taux, montants in avoir.tva_par_taux().items():
            _ajouter(deductible, taux, -montants['base'], -montants['tva'])

    total_collectee  = sum((d['tva'] for d in collectee.values()), Decimal('0'))
    total_deductible = sum((d['tva'] for d in deductible.values()), Decimal('0'))
    solde = _round2(total_collectee - total_deductible)

    return {
        'annee': int(annee),
        'periode': int(periode),
        'periodicite': periodicite,
        'regime': regime,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'collectee': _detail_to_list(collectee),
        'deductible': _detail_to_list(deductible),
        'total_collectee': float(_round2(total_collectee)),
        'total_deductible': float(_round2(total_deductible)),
        'tva_due': float(solde) if solde > 0 else 0.0,
        'credit_tva': float(-solde) if solde < 0 else 0.0,
    }


# ─── Relevé des déductions Simpl-TVA (EDI/XML DGI) ───────────────────────────
# Structure conforme au schéma DGI « ReleveDeduction » (une entrée <rd> par
# taux de TVA présent sur chaque paiement fournisseur de la période).
# À valider contre le XSD en vigueur sur simpl.tax.gov.ma avant télédéclaration.

def releve_deductions_xml(company, annee, periode, periodicite):
    date_from, date_to = periode_bounds(annee, periode, periodicite)

    root = ET.Element('DeclarationReleveDeduction')
    ET.SubElement(root, 'identifiantFiscal').text = company.if_fiscal or ''
    ET.SubElement(root, 'annee').text = str(int(annee))
    ET.SubElement(root, 'periode').text = str(int(periode))
    ET.SubElement(root, 'regime').text = '1' if periodicite == 'mensuelle' else '2'
    releve = ET.SubElement(root, 'releveDeductions')

    ordre = 0
    for paiement in _paiements_periode(['facture_achat'], date_from, date_to):
        document = paiement.document
        tiers = document.tiers
        for taux, base, tva, ttc in _ventilation_paiement(paiement):
            ordre += 1
            rd = ET.SubElement(releve, 'rd')
            ET.SubElement(rd, 'ord').text = str(ordre)
            ET.SubElement(rd, 'num').text = document.reference_externe or document.numero or ''
            ET.SubElement(rd, 'des').text = document.objet or (
                document.lignes.first().designation if document.lignes.exists() else '')
            ET.SubElement(rd, 'mht').text = f'{base:.2f}'
            ET.SubElement(rd, 'tva').text = f'{tva:.2f}'
            ET.SubElement(rd, 'ttc').text = f'{ttc:.2f}'
            ref = ET.SubElement(rd, 'refF')
            ET.SubElement(ref, 'if').text = tiers.if_fiscal or ''
            ET.SubElement(ref, 'nom').text = tiers.raison_sociale
            ET.SubElement(ref, 'ice').text = tiers.ice or ''
            ET.SubElement(rd, 'taux').text = f'{taux:.2f}'
            ET.SubElement(rd, 'mp').text = str(MODE_PAIEMENT_DGI.get(paiement.mode, 7))
            ET.SubElement(rd, 'dpai').text = paiement.date_paiement.isoformat()
            ET.SubElement(rd, 'dfac').text = document.date_emission.isoformat()

    ET.indent(root)
    return ET.tostring(root, encoding='UTF-8', xml_declaration=True)


# ─── Export Excel du rapport TVA ─────────────────────────────────────────────

def export_tva_xlsx(company, annee, periode, periodicite, regime):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    date_from, date_to = periode_bounds(annee, periode, periodicite)
    rapport = rapport_tva(annee, periode, periodicite, regime)
    bold = Font(bold=True)

    wb = Workbook()

    # ── Feuille 1 : synthèse ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Synthèse'
    ws.append(['Déclaration TVA', company.raison_sociale])
    ws.append(['IF', company.if_fiscal, 'ICE', company.ice])
    ws.append(['Période', f"{rapport['date_from']} → {rapport['date_to']}",
               'Régime', regime])
    ws.append([])
    ws.append(['TVA collectée', 'Base HT', 'TVA'])
    for row in rapport['collectee']:
        ws.append([f"Taux {row['taux']:g} %", row['base'], row['tva']])
    ws.append(['Total collectée', '', rapport['total_collectee']])
    ws.append([])
    ws.append(['TVA déductible', 'Base HT', 'TVA'])
    for row in rapport['deductible']:
        ws.append([f"Taux {row['taux']:g} %", row['base'], row['tva']])
    ws.append(['Total déductible', '', rapport['total_deductible']])
    ws.append([])
    ws.append(['TVA due', '', rapport['tva_due']])
    ws.append(['Crédit de TVA', '', rapport['credit_tva']])
    for cell in ('A1', 'A5', 'A9'):
        ws[cell].font = bold

    # ── Feuille 2 : détail des encaissements (collectée) ─────────────────────
    ws2 = wb.create_sheet('Collectée — détail')
    ws2.append(['Date', 'Facture', 'Client', 'ICE client', 'Mode',
                'Taux %', 'Base HT', 'TVA', 'TTC'])
    for cell in ws2[1]:
        cell.font = bold
    if regime == 'debit':
        factures = (
            Document.objects
            .filter(doc_type='facture',
                    statut__in=['validee', 'partiellement_payee', 'payee'],
                    date_emission__gte=date_from, date_emission__lte=date_to)
            .select_related('tiers').prefetch_related('lignes__tax_code')
        )
        for facture in factures:
            for taux, montants in sorted(facture.tva_par_taux().items(), reverse=True):
                ws2.append([
                    facture.date_emission.isoformat(), facture.numero,
                    facture.tiers.raison_sociale, facture.tiers.ice, '—',
                    float(taux), float(_round2(montants['base'])),
                    float(_round2(montants['tva'])),
                    float(_round2(montants['base'] + montants['tva'])),
                ])
    else:
        for paiement in _paiements_periode(['facture'], date_from, date_to):
            for taux, base, tva, ttc in _ventilation_paiement(paiement):
                ws2.append([
                    paiement.date_paiement.isoformat(), paiement.document.numero,
                    paiement.document.tiers.raison_sociale, paiement.document.tiers.ice,
                    paiement.get_mode_display(),
                    float(taux), float(base), float(tva), float(ttc),
                ])

    # ── Feuille 3 : détail des déductions ─────────────────────────────────────
    ws3 = wb.create_sheet('Déductible — détail')
    ws3.append(['Date paiement', 'Réf. fournisseur', 'N° interne', 'Fournisseur',
                'IF', 'ICE', 'Mode', 'Date facture',
                'Taux %', 'Base HT', 'TVA', 'TTC'])
    for cell in ws3[1]:
        cell.font = bold
    for paiement in _paiements_periode(['facture_achat'], date_from, date_to):
        document = paiement.document
        for taux, base, tva, ttc in _ventilation_paiement(paiement):
            ws3.append([
                paiement.date_paiement.isoformat(), document.reference_externe,
                document.numero, document.tiers.raison_sociale,
                document.tiers.if_fiscal, document.tiers.ice,
                paiement.get_mode_display(), document.date_emission.isoformat(),
                float(taux), float(base), float(tva), float(ttc),
            ])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
