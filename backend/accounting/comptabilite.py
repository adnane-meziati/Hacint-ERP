"""Comptabilisation automatique (CGNC) + rapports comptables.

Schémas d'écritures générées :

Facture de vente (journal VT) :
    D 3421 Clients ................. net à payer
    D 3458 État — RAS subie ........ retenue à la source (si RAS)
        C 7121 Ventes (ou compte de contrepartie du document) ... total HT
        C 4455 État — TVA facturée ............................. total TVA
Avoir de vente : écriture inverse.

Facture d'achat (journal AC) :
    D 6121 Achats (ou compte de contrepartie) ... total HT
    D 34552 État — TVA récupérable .............. total TVA
        C 4411 Fournisseurs ..................... net à payer
        C 4458 État — RAS à reverser ............ retenue à la source (si RAS)
Avoir d'achat : écriture inverse.

Encaissement client (journal BQ, ou CS si espèces) :
    D 5141 Banques / 5161 Caisse ... montant
        C 3421 Clients ............. montant
    + droit de timbre espèces : D 6167 / C 4458 (à reverser à l'État)

Décaissement fournisseur :
    D 4411 Fournisseurs ............ montant
        C 5141 / 5161 .............. montant
    + droit de timbre espèces : D 6167 / C 5161
"""
from decimal import Decimal

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Sum

from .models import (
    CompteComptable, Document, EcritureComptable, ExerciceComptable, Journal,
    JournalSequence, LigneEcriture, Paiement,
)

# Comptes PCGE utilisés par la comptabilisation automatique
COMPTES_AUTO = {
    'clients':          '3421',
    'fournisseurs':     '4411',
    'tva_facturee':     '4455',
    'tva_recuperable':  '34552',
    'etat_debiteur':    '3458',   # RAS subie sur ventes (créance sur l'État)
    'etat_crediteur':   '4458',   # RAS à reverser / timbre dû
    'banque':           '5141',
    'caisse':           '5161',
    'impots_taxes':     '6167',   # droits de timbre
    'vente_defaut':     '7121',
    'achat_defaut':     '6121',
}


def _compte(cle_ou_numero):
    numero = COMPTES_AUTO.get(cle_ou_numero, cle_ou_numero)
    try:
        return CompteComptable.objects.get(numero=numero)
    except CompteComptable.DoesNotExist:
        raise ValidationError(
            f"Le compte PCGE {numero} est introuvable — "
            "vérifiez le plan comptable (onglet Plan comptable).")


def _journal(code):
    try:
        return Journal.objects.get(code=code)
    except Journal.DoesNotExist:
        raise ValidationError(f'Le journal {code} est introuvable.')


def _creer_ecriture(journal, date_ecriture, libelle, lignes, user=None,
                    document=None, paiement=None):
    """Crée une écriture équilibrée. `lignes` = [(compte, libelle, debit, credit, tiers)]."""
    lignes = [l for l in lignes if l[2] or l[3]]   # ignorer les lignes à zéro
    total_debit  = sum((l[2] for l in lignes), Decimal('0'))
    total_credit = sum((l[3] for l in lignes), Decimal('0'))
    if total_debit != total_credit:
        raise ValidationError(
            f'Écriture déséquilibrée : débit {total_debit} ≠ crédit {total_credit}.')
    if not lignes:
        return None
    ExerciceComptable.verifier_ouvert(date_ecriture)
    with transaction.atomic():
        ecriture = EcritureComptable.objects.create(
            journal=journal,
            numero=JournalSequence.next_numero(journal, date_ecriture.year),
            date_ecriture=date_ecriture,
            libelle=libelle,
            document=document,
            paiement=paiement,
            created_by=user,
        )
        for ordre, (compte, lib, debit, credit, tiers) in enumerate(lignes):
            LigneEcriture.objects.create(
                ecriture=ecriture, ordre=ordre, compte=compte,
                libelle=lib or libelle, debit=debit, credit=credit, tiers=tiers)
    return ecriture


# ─── Comptabilisation d'un document ───────────────────────────────────────────

def comptabiliser_document(document, user=None):
    """Génère l'écriture d'une facture/avoir validé. Idempotent (None si déjà fait)."""
    if document.doc_type == 'devis' or document.statut == 'brouillon':
        return None
    if EcritureComptable.objects.filter(document=document).exists():
        return None

    tiers = document.tiers
    est_vente = document.doc_type in ('facture', 'avoir')
    est_avoir = document.doc_type in ('avoir', 'avoir_achat')

    if est_vente:
        compte_tiers = _compte('clients')
        compte_contrepartie = document.compte_contrepartie or _compte('vente_defaut')
        compte_tva = _compte('tva_facturee')
        compte_ras = _compte('etat_debiteur')
        journal = _journal('VT')
    else:
        compte_tiers = _compte('fournisseurs')
        compte_contrepartie = document.compte_contrepartie or _compte('achat_defaut')
        compte_tva = _compte('tva_recuperable')
        compte_ras = _compte('etat_crediteur')
        journal = _journal('AC')

    zero = Decimal('0')
    net, ht, tva, ras = (document.net_a_payer, document.total_ht,
                         document.total_tva, document.ras_montant)

    if est_vente:
        lignes = [
            (compte_tiers,        '', net, zero, tiers),
            (compte_ras,          'Retenue à la source subie', ras, zero, None),
            (compte_contrepartie, '', zero, ht, None),
            (compte_tva,          '', zero, tva, None),
        ]
    else:
        lignes = [
            (compte_contrepartie, '', ht, zero, None),
            (compte_tva,          '', tva, zero, None),
            (compte_tiers,        '', zero, net, tiers),
            (compte_ras,          'Retenue à la source à reverser', zero, ras, None),
        ]
    if est_avoir:   # écriture inverse
        lignes = [(c, l, credit, debit, t) for (c, l, debit, credit, t) in lignes]

    libelle = f'{document.get_doc_type_display()} {document.numero} — {tiers.raison_sociale}'
    return _creer_ecriture(journal, document.date_emission, libelle, lignes,
                           user=user, document=document)


# ─── Comptabilisation d'un paiement ───────────────────────────────────────────

def comptabiliser_paiement(paiement, user=None):
    """Génère l'écriture de trésorerie d'un paiement. Idempotent."""
    if EcritureComptable.objects.filter(paiement=paiement).exists():
        return None

    document = paiement.document
    tiers = document.tiers
    est_encaissement = document.doc_type == 'facture'
    en_especes = paiement.mode == 'espece'

    compte_treso = _compte('caisse' if en_especes else 'banque')
    journal = _journal('CS' if en_especes else 'BQ')
    zero = Decimal('0')
    montant = paiement.montant
    timbre = paiement.timbre_montant or zero

    if est_encaissement:
        lignes = [
            (compte_treso,        '', montant, zero, None),
            (_compte('clients'),  '', zero, montant, tiers),
        ]
        if timbre:   # timbre sur quittance espèces, dû à l'État
            lignes += [
                (_compte('impots_taxes'),   'Droit de timbre espèces', timbre, zero, None),
                (_compte('etat_crediteur'), 'Droit de timbre à reverser', zero, timbre, None),
            ]
        sens = 'Encaissement'
    else:
        lignes = [
            (_compte('fournisseurs'), '', montant, zero, tiers),
            (compte_treso,            '', zero, montant, None),
        ]
        if timbre:   # timbre décaissé avec le règlement espèces
            lignes += [
                (_compte('impots_taxes'), 'Droit de timbre espèces', timbre, zero, None),
                (compte_treso,            '', zero, timbre, None),
            ]
        sens = 'Règlement'

    libelle = (f'{sens} {paiement.get_mode_display().lower()} — '
               f'{document.numero} — {tiers.raison_sociale}')
    return _creer_ecriture(journal, paiement.date_paiement, libelle, lignes,
                           user=user, paiement=paiement)


def comptabiliser_tout(user=None):
    """Comptabilise tous les documents validés et paiements sans écriture."""
    nb_docs = 0
    documents = (
        Document.objects
        .exclude(doc_type='devis').exclude(statut__in=['brouillon', 'annulee'])
        .filter(ecriture__isnull=True)
        .order_by('date_emission', 'id')
    )
    for document in documents:
        if comptabiliser_document(document, user=user):
            nb_docs += 1
    nb_paiements = 0
    for paiement in Paiement.objects.filter(ecriture__isnull=True) \
                                    .order_by('date_paiement', 'id'):
        if comptabiliser_paiement(paiement, user=user):
            nb_paiements += 1
    return {'documents': nb_docs, 'paiements': nb_paiements}


# ─── Rapports : grand livre, balance, bilan, CPC ─────────────────────────────

def _lignes_periode(date_from=None, date_to=None):
    qs = LigneEcriture.objects.select_related('ecriture', 'compte')
    if date_from:
        qs = qs.filter(ecriture__date_ecriture__gte=date_from)
    if date_to:
        qs = qs.filter(ecriture__date_ecriture__lte=date_to)
    return qs


def grand_livre(compte, date_from=None, date_to=None):
    """Mouvements d'un compte avec solde progressif."""
    report = Decimal('0')
    if date_from:
        avant = _lignes_periode(None, None).filter(
            compte=compte, ecriture__date_ecriture__lt=date_from,
        ).aggregate(d=Sum('debit'), c=Sum('credit'))
        report = (avant['d'] or Decimal('0')) - (avant['c'] or Decimal('0'))

    lignes = (
        _lignes_periode(date_from, date_to)
        .filter(compte=compte)
        .order_by('ecriture__date_ecriture', 'ecriture_id', 'ordre')
    )
    solde = report
    resultat = []
    total_debit = total_credit = Decimal('0')
    for ligne in lignes:
        solde += ligne.debit - ligne.credit
        total_debit += ligne.debit
        total_credit += ligne.credit
        resultat.append({
            'date': ligne.ecriture.date_ecriture.isoformat(),
            'ecriture': ligne.ecriture.numero,
            'journal': ligne.ecriture.journal.code,
            'libelle': ligne.libelle or ligne.ecriture.libelle,
            'tiers': ligne.tiers.raison_sociale if ligne.tiers else None,
            'debit': float(ligne.debit),
            'credit': float(ligne.credit),
            'solde': float(solde),
        })
    return {
        'compte': {'numero': compte.numero, 'intitule': compte.intitule},
        'report': float(report),
        'lignes': resultat,
        'total_debit': float(total_debit),
        'total_credit': float(total_credit),
        'solde_final': float(solde),
    }


def balance(date_from=None, date_to=None):
    """Balance générale : totaux débit/crédit et soldes par compte."""
    rows = (
        _lignes_periode(date_from, date_to)
        .values('compte__numero', 'compte__intitule')
        .annotate(debit=Sum('debit'), credit=Sum('credit'))
        .order_by('compte__numero')
    )
    comptes = []
    totaux = {'debit': Decimal('0'), 'credit': Decimal('0'),
              'solde_debiteur': Decimal('0'), 'solde_crediteur': Decimal('0')}
    for row in rows:
        debit, credit = row['debit'] or Decimal('0'), row['credit'] or Decimal('0')
        solde = debit - credit
        comptes.append({
            'numero': row['compte__numero'],
            'intitule': row['compte__intitule'],
            'classe': row['compte__numero'][0],
            'debit': float(debit),
            'credit': float(credit),
            'solde_debiteur': float(solde) if solde > 0 else 0.0,
            'solde_crediteur': float(-solde) if solde < 0 else 0.0,
        })
        totaux['debit'] += debit
        totaux['credit'] += credit
        totaux['solde_debiteur'] += max(solde, Decimal('0'))
        totaux['solde_crediteur'] += max(-solde, Decimal('0'))
    return {
        'comptes': comptes,
        'totaux': {k: float(v) for k, v in totaux.items()},
        'equilibree': totaux['debit'] == totaux['credit'],
    }


def compte_de_produits_et_charges(annee):
    """CPC simplifié : produits (classe 7) − charges (classe 6) de l'exercice."""
    from datetime import date
    data = balance(date(annee, 1, 1), date(annee, 12, 31))
    charges  = [c for c in data['comptes'] if c['classe'] == '6']
    produits = [c for c in data['comptes'] if c['classe'] == '7']
    total_charges  = sum(c['debit'] - c['credit'] for c in charges)
    total_produits = sum(c['credit'] - c['debit'] for c in produits)
    return {
        'annee': annee,
        'produits': produits,
        'charges': charges,
        'total_produits': round(total_produits, 2),
        'total_charges': round(total_charges, 2),
        'resultat': round(total_produits - total_charges, 2),
    }


# ─── Exports Excel (livre-journal, grand livre, balance) ─────────────────────

def _nouveau_classeur(titre, entetes):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = titre[:31]
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return wb, ws, BytesIO


def export_journal_xlsx(date_from=None, date_to=None, journal_code=None):
    """Livre-journal : toutes les lignes d'écritures de la période."""
    wb, ws, BytesIO = _nouveau_classeur('Livre-journal', [
        'Date', 'Journal', 'N° écriture', 'Compte', 'Intitulé compte',
        'Libellé', 'Tiers', 'Débit', 'Crédit'])
    lignes = (
        _lignes_periode(date_from, date_to)
        .select_related('ecriture__journal', 'tiers')
        .order_by('ecriture__date_ecriture', 'ecriture_id', 'ordre')
    )
    if journal_code:
        lignes = lignes.filter(ecriture__journal__code=journal_code)
    total_debit = total_credit = Decimal('0')
    for ligne in lignes:
        total_debit += ligne.debit
        total_credit += ligne.credit
        ws.append([
            ligne.ecriture.date_ecriture.isoformat(),
            ligne.ecriture.journal.code,
            ligne.ecriture.numero,
            ligne.compte.numero,
            ligne.compte.intitule,
            ligne.libelle or ligne.ecriture.libelle,
            ligne.tiers.raison_sociale if ligne.tiers else '',
            float(ligne.debit), float(ligne.credit),
        ])
    ws.append(['', '', '', '', '', 'TOTAL', '', float(total_debit), float(total_credit)])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_grand_livre_xlsx(compte, date_from=None, date_to=None):
    donnees = grand_livre(compte, date_from, date_to)
    wb, ws, BytesIO = _nouveau_classeur(
        f'GL {compte.numero}', ['Date', 'Écriture', 'Journal', 'Libellé',
                                'Tiers', 'Débit', 'Crédit', 'Solde'])
    ws.insert_rows(1)
    ws['A1'] = f"Grand livre — {compte.numero} {compte.intitule}"
    if donnees['report']:
        ws.append(['', '', '', 'Report à nouveau', '', '', '', donnees['report']])
    for ligne in donnees['lignes']:
        ws.append([ligne['date'], ligne['ecriture'], ligne['journal'],
                   ligne['libelle'], ligne['tiers'] or '',
                   ligne['debit'], ligne['credit'], ligne['solde']])
    ws.append(['', '', '', 'TOTAL', '', donnees['total_debit'],
               donnees['total_credit'], donnees['solde_final']])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_balance_xlsx(date_from=None, date_to=None):
    donnees = balance(date_from, date_to)
    wb, ws, BytesIO = _nouveau_classeur('Balance', [
        'Compte', 'Intitulé', 'Débit', 'Crédit',
        'Solde débiteur', 'Solde créditeur'])
    for compte in donnees['comptes']:
        ws.append([compte['numero'], compte['intitule'],
                   compte['debit'], compte['credit'],
                   compte['solde_debiteur'], compte['solde_crediteur']])
    totaux = donnees['totaux']
    ws.append(['', 'TOTAL', totaux['debit'], totaux['credit'],
               totaux['solde_debiteur'], totaux['solde_crediteur']])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def bilan(date_to):
    """Bilan simplifié à une date : classes 1-5 + résultat de l'exercice en cours."""
    data = balance(None, date_to)
    actif, passif = [], []
    total_actif = total_passif = Decimal('0')
    for compte in data['comptes']:
        solde = Decimal(str(compte['debit'])) - Decimal(str(compte['credit']))
        if compte['classe'] in ('6', '7', '8'):
            continue   # → intégrés via le résultat net ci-dessous
        if solde > 0:
            actif.append({**compte, 'montant': float(solde)})
            total_actif += solde
        elif solde < 0:
            passif.append({**compte, 'montant': float(-solde)})
            total_passif += -solde
    # Résultat de l'exercice = produits − charges (solde créditeur ⇒ passif)
    resultat_net = Decimal('0')
    for compte in data['comptes']:
        solde = Decimal(str(compte['debit'])) - Decimal(str(compte['credit']))
        if compte['classe'] == '7':
            resultat_net += -solde
        elif compte['classe'] == '6':
            resultat_net -= solde
    if resultat_net >= 0:
        passif.append({'numero': '1191', 'intitule': "Résultat net de l'exercice",
                       'classe': '1', 'montant': float(resultat_net)})
        total_passif += resultat_net
    else:
        actif.append({'numero': '1199', 'intitule': "Résultat net (perte)",
                      'classe': '1', 'montant': float(-resultat_net)})
        total_actif += -resultat_net
    return {
        'date': date_to.isoformat(),
        'actif': sorted(actif, key=lambda c: c['numero']),
        'passif': sorted(passif, key=lambda c: c['numero']),
        'total_actif': float(total_actif),
        'total_passif': float(total_passif),
        'equilibre': total_actif == total_passif,
    }
