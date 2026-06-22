import csv
import io
from io import BytesIO
from datetime import date, timedelta
from xml.sax.saxutils import escape

from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rest_framework import permissions, status, viewsets
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import (
    DeliveryOrder,
    Driver,
    LogisticsNotification,
    LogisticsReport,
    LogisticsTask,
    LogisticsTaskAttachment,
    LogisticsTaskComment,
    LogisticsTaskHistory,
    Shipment,
    Vehicle,
    WarehouseTransfer,
)
from .serializers import (
    DeliveryOrderSerializer,
    DriverSerializer,
    LogisticsNotificationSerializer,
    LogisticsReportSerializer,
    LogisticsTaskAttachmentSerializer,
    LogisticsTaskCommentSerializer,
    LogisticsTaskHistorySerializer,
    LogisticsTaskSerializer,
    ShipmentSerializer,
    VehicleSerializer,
    WarehouseTransferSerializer,
)


def user_display_name(user):
    if not user:
        return 'SystÃ¨me'
    return user.get_full_name() or user.username


def task_queryset():
    return LogisticsTask.objects.select_related(
        'created_by',
        'delivery_order',
        'shipment',
        'transfer',
    ).prefetch_related(
        'assigned_employees',
        'assigned_employees__department',
        'assigned_employees__user',
        'comments',
        'comments__author',
        'attachments',
        'attachments__uploaded_by',
        'history',
        'history__actor',
    )


def task_filters(queryset, params):
    status_filter = params.get('status')
    employee = params.get('employee')
    priority = params.get('priority')
    date_filter = params.get('date')
    role = params.get('role')
    search = params.get('search')
    late_only = params.get('late')

    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if employee:
        queryset = queryset.filter(assigned_employees__id=employee)

    if priority:
        queryset = queryset.filter(priority=priority)

    if date_filter:
        queryset = queryset.filter(due_date=date_filter)

    if role:
        queryset = queryset.filter(assigned_role=role)

    if search:
        queryset = queryset.filter(
            Q(title__icontains=search)
            | Q(description__icontains=search)
            | Q(assigned_employees__first_name__icontains=search)
            | Q(assigned_employees__last_name__icontains=search)
        )

    if str(late_only).lower() in ('1', 'true', 'yes'):
        queryset = queryset.filter(due_date__lt=date.today()).exclude(
            status__in=[
                LogisticsTask.TaskStatus.DONE,
                LogisticsTask.TaskStatus.CANCELLED,
            ]
        )

    return queryset.distinct()


def create_history(task, actor, action, old_value='', new_value=''):
    return LogisticsTaskHistory.objects.create(
        task=task,
        actor=actor,
        action=action,
        old_value=str(old_value or ''),
        new_value=str(new_value or ''),
    )


def create_notification(
    task,
    notification_type,
    title,
    message,
    employees=None,
):
    employees = list(employees or task.assigned_employees.select_related('user'))

    for employee in employees:
        if not employee.user_id:
            continue

        LogisticsNotification.objects.create(
            recipient=employee.user,
            employee=employee,
            task=task,
            notification_type=notification_type,
            title=title,
            message=message,
        )


def notify_deadlines():
    limit = date.today() + timedelta(days=2)

    tasks = LogisticsTask.objects.filter(
        due_date__gte=date.today(),
        due_date__lte=limit,
    ).exclude(
        status__in=[
            LogisticsTask.TaskStatus.DONE,
            LogisticsTask.TaskStatus.CANCELLED,
        ]
    ).prefetch_related('assigned_employees__user')

    for task in tasks:
        for employee in task.assigned_employees.all():
            if not employee.user_id:
                continue

            already_exists = LogisticsNotification.objects.filter(
                recipient=employee.user,
                task=task,
                notification_type=(
                    LogisticsNotification.NotificationType.DEADLINE
                ),
                created_at__date=date.today(),
            ).exists()

            if not already_exists:
                LogisticsNotification.objects.create(
                    recipient=employee.user,
                    employee=employee,
                    task=task,
                    notification_type=(
                        LogisticsNotification.NotificationType.DEADLINE
                    ),
                    title='Ã‰chÃ©ance de tÃ¢che proche',
                    message=(
                        f'La tÃ¢che Â« {task.title} Â» arrive Ã  Ã©chÃ©ance '
                        f'le {task.due_date}.'
                    ),
                )


def csv_response(filename, headings, rows):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(headings)

    for row in rows:
        writer.writerow(row)

    return response


def excel_response(filename, sheet_name, headings, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.freeze_panes = 'A2'
    worksheet.append(headings)

    for cell in worksheet[1]:
        cell.fill = PatternFill('solid', fgColor='0F766E')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        worksheet.append([
            '' if value is None else value
            for value in row
        ])

    worksheet.auto_filter.ref = worksheet.dimensions

    for index, column in enumerate(worksheet.columns, start=1):
        width = max(len(str(cell.value or '')) for cell in column)
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(width + 2, 12),
            45,
        )

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    return response


ACTIVE_SHIPMENT_STATUSES = {
    Shipment.ShipmentStatus.PREPARATION,
    Shipment.ShipmentStatus.SHIPPED,
    Shipment.ShipmentStatus.IN_DELIVERY,
}


def sync_shipment_resources(*resources):
    vehicle_ids = {
        resource.id
        for resource in resources
        if isinstance(resource, Vehicle)
    }
    driver_ids = {
        resource.id
        for resource in resources
        if isinstance(resource, Driver)
    }
    employee_ids = {
        resource.id
        for resource in resources
        if resource.__class__.__name__ == 'Employee'
    }
    driver_ids.update(
        Driver.objects.filter(employee_id__in=employee_ids).values_list(
            'id',
            flat=True,
        )
    )

    for vehicle in Vehicle.objects.filter(id__in=vehicle_ids):
        desired = (
            Vehicle.VehicleStatus.IN_USE
            if (
                vehicle.shipments.filter(
                    status__in=ACTIVE_SHIPMENT_STATUSES,
                ).exists()
                or vehicle.warehouse_transfers.filter(
                    status__in=['draft', 'pending_approval', 'approved', 'in_transit'],
                ).exists()
                or vehicle.logistics_tasks.filter(
                    status__in=['todo', 'in_progress', 'waiting'],
                ).exists()
            )
            else Vehicle.VehicleStatus.AVAILABLE
        )
        if vehicle.status not in {
            Vehicle.VehicleStatus.MAINTENANCE,
            Vehicle.VehicleStatus.INACTIVE,
        } and vehicle.status != desired:
            vehicle.status = desired
            vehicle.save(update_fields=['status', 'updated_at'])

    for driver in Driver.objects.filter(id__in=driver_ids):
        desired = (
            Driver.DriverStatus.ASSIGNED
            if (
                driver.shipments.filter(
                    status__in=ACTIVE_SHIPMENT_STATUSES,
                ).exists()
                or driver.warehouse_transfers.filter(
                    status__in=['draft', 'pending_approval', 'approved', 'in_transit'],
                ).exists()
                or driver.employee.logistics_tasks.filter(
                    status__in=['todo', 'in_progress', 'waiting'],
                ).exists()
            )
            else Driver.DriverStatus.AVAILABLE
        )
        if driver.status not in {
            Driver.DriverStatus.ON_LEAVE,
            Driver.DriverStatus.INACTIVE,
        } and driver.status != desired:
            driver.status = desired
            driver.save(update_fields=['status', 'updated_at'])


def simple_pdf_response(filename, title, headings, rows):
    page_width = 842
    page_height = 595
    margin = 32
    row_height = 18
    column_count = max(len(headings), 1)
    column_width = (page_width - margin * 2) / column_count

    pages = []
    current_rows = []
    maximum_rows = 25

    for row in rows:
        current_rows.append(row)
        if len(current_rows) >= maximum_rows:
            pages.append(current_rows)
            current_rows = []

    if current_rows or not pages:
        pages.append(current_rows)

    objects = []
    page_ids = []
    font_id = 3

    for page_index, page_rows in enumerate(pages):
        content_parts = [
            'BT',
            '/F1 16 Tf',
            f'{margin} {page_height - 35} Td',
            f'({pdf_text(title)}) Tj',
            'ET',
        ]

        y = page_height - 65

        for column_index, heading in enumerate(headings):
            x = margin + column_index * column_width
            content_parts.extend([
                'BT',
                '/F1 8 Tf',
                f'{x:.2f} {y:.2f} Td',
                f'({pdf_text(heading)}) Tj',
                'ET',
            ])

        y -= row_height

        for row in page_rows:
            for column_index, value in enumerate(row):
                x = margin + column_index * column_width
                text = str(value if value is not None else '')
                text = text[:32]

                content_parts.extend([
                    'BT',
                    '/F1 7 Tf',
                    f'{x:.2f} {y:.2f} Td',
                    f'({pdf_text(text)}) Tj',
                    'ET',
                ])

            y -= row_height

        content = '\n'.join(content_parts).encode('latin-1', 'replace')
        content_id = 4 + page_index * 2
        page_id = content_id + 1
        page_ids.append(page_id)

        objects.append((
            content_id,
            b'<< /Length %d >>\nstream\n' % len(content)
            + content
            + b'\nendstream',
        ))
        objects.append((
            page_id,
            (
                f'<< /Type /Page /Parent 2 0 R '
                f'/MediaBox [0 0 {page_width} {page_height}] '
                f'/Resources << /Font << /F1 {font_id} 0 R >> >> '
                f'/Contents {content_id} 0 R >>'
            ).encode(),
        ))

    base_objects = [
        (1, b'<< /Type /Catalog /Pages 2 0 R >>'),
        (
            2,
            (
                f'<< /Type /Pages /Kids '
                f'[{" ".join(f"{page_id} 0 R" for page_id in page_ids)}] '
                f'/Count {len(page_ids)} >>'
            ).encode(),
        ),
        (
            font_id,
            b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
        ),
    ]

    all_objects = sorted(base_objects + objects, key=lambda item: item[0])
    output = bytearray(b'%PDF-1.4\n')
    offsets = {0: 0}

    for object_id, body in all_objects:
        offsets[object_id] = len(output)
        output.extend(f'{object_id} 0 obj\n'.encode())
        output.extend(body)
        output.extend(b'\nendobj\n')

    xref_position = len(output)
    maximum_id = max(offsets)
    output.extend(f'xref\n0 {maximum_id + 1}\n'.encode())
    output.extend(b'0000000000 65535 f \n')

    for object_id in range(1, maximum_id + 1):
        offset = offsets.get(object_id, 0)
        output.extend(f'{offset:010d} 00000 n \n'.encode())

    output.extend(
        (
            f'trailer\n<< /Size {maximum_id + 1} /Root 1 0 R >>\n'
            f'startxref\n{xref_position}\n%%EOF'
        ).encode()
    )

    response = HttpResponse(
        bytes(output),
        content_type='application/pdf',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    return response


def pdf_text(value):
    return (
        escape(str(value))
        .replace('\\', '\\\\')
        .replace('(', '\\(')
        .replace(')', '\\)')
    )


class VehicleViewSet(viewsets.ModelViewSet):
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Vehicle.objects.all()
        status_filter = self.request.query_params.get('status')
        search = self.request.query_params.get('search')

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if search:
            queryset = queryset.filter(
                Q(registration__icontains=search)
                | Q(vehicle_type__icontains=search)
            )

        return queryset

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        rows = [
            [
                vehicle.registration,
                vehicle.vehicle_type,
                f'{vehicle.capacity} kg' if vehicle.capacity else '',
                vehicle.get_status_display(),
                vehicle.service_date or '',
                vehicle.notes,
            ]
            for vehicle in self.get_queryset()
        ]
        return excel_response(
            'vehicules_logistique.xlsx',
            'Véhicules',
            ['Immatriculation', 'Type', 'Capacité', 'Statut', 'Mise en service', 'Notes'],
            rows,
        )


class DriverViewSet(viewsets.ModelViewSet):
    serializer_class = DriverSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Driver.objects.select_related(
            'employee',
            'employee__department',
        )
        status_filter = self.request.query_params.get('status')
        search = self.request.query_params.get('search')

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if search:
            queryset = queryset.filter(
                Q(employee__first_name__icontains=search)
                | Q(employee__last_name__icontains=search)
                | Q(employee__employee_number__icontains=search)
                | Q(license_number__icontains=search)
            )

        return queryset

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        rows = [
            [
                driver.employee.employee_number,
                driver.employee.full_name,
                driver.employee.department.name if driver.employee.department else '',
                driver.license_number,
                driver.license_expiry_date,
                driver.get_status_display(),
                driver.notes,
            ]
            for driver in self.get_queryset()
        ]
        return excel_response(
            'chauffeurs_logistique.xlsx',
            'Chauffeurs',
            ['Matricule', 'Employé', 'Département', 'Permis', 'Expiration', 'Statut', 'Notes'],
            rows,
        )


class DeliveryOrderViewSet(viewsets.ModelViewSet):
    serializer_class = DeliveryOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = DeliveryOrder.objects.prefetch_related(
            'lines',
            'lines__article',
        )
        status_filter = self.request.query_params.get('status')
        customer = self.request.query_params.get('customer')
        search = self.request.query_params.get('search')

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if customer:
            queryset = queryset.filter(customer__icontains=customer)

        if search:
            queryset = queryset.filter(
                Q(delivery_number__icontains=search)
                | Q(customer__icontains=search)
                | Q(delivery_address__icontains=search)
            )

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        rows = [
            [
                order.delivery_number,
                order.delivery_date,
                order.customer,
                order.delivery_address,
                order.get_status_display(),
                order.notes,
            ]
            for order in self.get_queryset()
        ]
        return excel_response(
            'ordres_livraison.xlsx',
            'Livraisons',
            ['Numéro', 'Date', 'Client', 'Adresse', 'Statut', 'Notes'],
            rows,
        )


class ShipmentViewSet(viewsets.ModelViewSet):
    serializer_class = ShipmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Shipment.objects.select_related(
            'delivery_order',
            'vehicle',
            'driver',
            'driver__employee',
        ).prefetch_related(
            'lines',
            'lines__article',
        )

        status_filter = self.request.query_params.get('status')
        driver = self.request.query_params.get('driver')
        vehicle = self.request.query_params.get('vehicle')
        search = self.request.query_params.get('search')

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if driver:
            queryset = queryset.filter(driver_id=driver)

        if vehicle:
            queryset = queryset.filter(vehicle_id=vehicle)

        if search:
            queryset = queryset.filter(
                Q(tracking_number__icontains=search)
                | Q(
                    delivery_order__delivery_number__icontains=search
                )
            )

        return queryset

    def perform_create(self, serializer):
        shipment = serializer.save(created_by=self.request.user)
        sync_shipment_resources(shipment.vehicle, shipment.driver)

    def perform_update(self, serializer):
        previous = self.get_object()
        old_vehicle = previous.vehicle
        old_driver = previous.driver
        shipment = serializer.save()
        sync_shipment_resources(
            old_vehicle,
            old_driver,
            shipment.vehicle,
            shipment.driver,
        )

    def perform_destroy(self, instance):
        vehicle = instance.vehicle
        driver = instance.driver
        instance.delete()
        sync_shipment_resources(vehicle, driver)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        rows = [
            [
                shipment.tracking_number,
                shipment.delivery_order.delivery_number if shipment.delivery_order else '',
                shipment.shipment_date,
                shipment.vehicle.registration if shipment.vehicle else '',
                shipment.driver.employee.full_name if shipment.driver else '',
                shipment.driver.employee.employee_number if shipment.driver else '',
                shipment.get_status_display(),
                shipment.notes,
            ]
            for shipment in self.get_queryset()
        ]
        return excel_response(
            'expeditions_logistique.xlsx',
            'Expéditions',
            ['Suivi', 'Ordre', 'Date', 'Véhicule', 'Chauffeur', 'Matricule', 'Statut', 'Notes'],
            rows,
        )


class WarehouseTransferViewSet(viewsets.ModelViewSet):
    serializer_class = WarehouseTransferSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = WarehouseTransfer.objects.select_related(
            'source_warehouse',
            'destination_warehouse',
            'requested_by',
            'approved_by',
            'vehicle',
            'driver',
        ).prefetch_related(
            'lines',
            'lines__article',
        )

        status_filter = self.request.query_params.get('status')
        source = self.request.query_params.get('source_warehouse')
        destination = self.request.query_params.get(
            'destination_warehouse'
        )
        search = self.request.query_params.get('search')

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if source:
            queryset = queryset.filter(source_warehouse_id=source)

        if destination:
            queryset = queryset.filter(
                destination_warehouse_id=destination
            )

        if search:
            queryset = queryset.filter(
                transfer_number__icontains=search
            )

        return queryset

    def perform_create(self, serializer):
        transfer = serializer.save(requested_by=self.request.user)
        sync_shipment_resources(transfer.vehicle, transfer.driver)

    def perform_update(self, serializer):
        previous = self.get_object()
        old_vehicle = previous.vehicle
        old_driver = previous.driver
        transfer = serializer.save()
        sync_shipment_resources(
            old_vehicle,
            transfer.vehicle,
            old_driver,
            transfer.driver,
        )

    def perform_destroy(self, instance):
        vehicle = instance.vehicle
        driver = instance.driver
        super().perform_destroy(instance)
        sync_shipment_resources(vehicle, driver)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        rows = [
            [
                transfer.transfer_number,
                transfer.requested_date,
                transfer.source_warehouse.nom,
                (
                    transfer.destination_warehouse.nom
                    if transfer.destination_warehouse
                    else transfer.external_destination
                ),
                transfer.get_destination_type_display(),
                transfer.get_status_display(),
                transfer.notes,
            ]
            for transfer in self.get_queryset()
        ]
        return excel_response(
            'transferts_logistique.xlsx',
            'Transferts',
            ['Numéro', 'Date', 'Source', 'Destination', 'Type destination', 'Statut', 'Notes'],
            rows,
        )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        transfer = self.get_object()

        if transfer.status not in [
            WarehouseTransfer.TransferStatus.PENDING_APPROVAL,
            WarehouseTransfer.TransferStatus.DRAFT,
        ]:
            return Response(
                {'detail': 'Ce transfert ne peut pas Ãªtre approuvÃ©.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        transfer.status = WarehouseTransfer.TransferStatus.APPROVED
        transfer.approved_by = request.user
        transfer.approved_at = timezone.now()
        transfer.save()

        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        transfer = self.get_object()

        if transfer.status not in [
            WarehouseTransfer.TransferStatus.PENDING_APPROVAL,
            WarehouseTransfer.TransferStatus.DRAFT,
        ]:
            return Response(
                {'detail': 'Ce transfert ne peut pas Ãªtre refusÃ©.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        transfer.status = WarehouseTransfer.TransferStatus.REJECTED
        transfer.save()

        return Response(self.get_serializer(transfer).data)


class LogisticsTaskViewSet(viewsets.ModelViewSet):
    serializer_class = LogisticsTaskSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return task_filters(
            task_queryset(),
            self.request.query_params,
        )

    def perform_create(self, serializer):
        task = serializer.save(created_by=self.request.user)
        sync_shipment_resources(task.vehicle, *task.assigned_employees.all())

        create_history(
            task=task,
            actor=self.request.user,
            action='CrÃ©ation',
            new_value='TÃ¢che crÃ©Ã©e',
        )

        create_notification(
            task=task,
            notification_type=(
                LogisticsNotification.NotificationType.ASSIGNMENT
            ),
            title='Nouvelle tÃ¢che assignÃ©e',
            message=f'La tÃ¢che Â« {task.title} Â» vous a Ã©tÃ© assignÃ©e.',
        )

    def perform_update(self, serializer):
        task_before = task_queryset().get(pk=self.get_object().pk)
        old_vehicle = task_before.vehicle
        old_status = task_before.status
        old_role = task_before.assigned_role
        old_title = task_before.title
        old_due_date = task_before.due_date
        old_employee_ids = set(
            task_before.assigned_employees.values_list('id', flat=True)
        )
        old_employees = list(task_before.assigned_employees.all())

        task = serializer.save()
        sync_shipment_resources(
            old_vehicle,
            task.vehicle,
            *old_employees,
            *task.assigned_employees.all(),
        )

        new_employee_ids = set(
            task.assigned_employees.values_list('id', flat=True)
        )

        changes = []

        if old_title != task.title:
            changes.append(f'Titre : {old_title} â†’ {task.title}')
            create_history(
                task,
                self.request.user,
                'Modification du titre',
                old_title,
                task.title,
            )

        if old_due_date != task.due_date:
            changes.append(
                f'Ã‰chÃ©ance : {old_due_date or "â€“"} â†’ '
                f'{task.due_date or "â€“"}'
            )
            create_history(
                task,
                self.request.user,
                "Modification de l'Ã©chÃ©ance",
                old_due_date or '',
                task.due_date or '',
            )

        if old_status != task.status:
            create_history(
                task,
                self.request.user,
                'Changement de statut',
                task_before.get_status_display(),
                task.get_status_display(),
            )

            notification_type = (
                LogisticsNotification.NotificationType.MODIFICATION
            )
            title = 'Statut de tÃ¢che modifiÃ©'

            if task.status == LogisticsTask.TaskStatus.DONE:
                notification_type = (
                    LogisticsNotification.NotificationType.COMPLETED
                )
                title = 'TÃ¢che terminÃ©e'
            elif task.status == LogisticsTask.TaskStatus.CANCELLED:
                notification_type = (
                    LogisticsNotification.NotificationType.CANCELLED
                )
                title = 'TÃ¢che annulÃ©e'

            create_notification(
                task,
                notification_type,
                title,
                (
                    f'Le statut de la tÃ¢che Â« {task.title} Â» est '
                    f'maintenant {task.get_status_display()}.'
                ),
            )

        if old_role != task.assigned_role:
            create_history(
                task,
                self.request.user,
                'Changement de rÃ´le',
                task_before.get_assigned_role_display(),
                task.get_assigned_role_display(),
            )
            changes.append(
                f'RÃ´le : {task_before.get_assigned_role_display()} â†’ '
                f'{task.get_assigned_role_display()}'
            )

        if old_employee_ids != new_employee_ids:
            old_names = ', '.join(
                task_before.assigned_employees.values_list(
                    'first_name',
                    flat=True,
                )
            )
            new_names = ', '.join(
                task.assigned_employees.values_list(
                    'first_name',
                    flat=True,
                )
            )

            create_history(
                task,
                self.request.user,
                "RÃ©affectation d'employÃ©s",
                old_names,
                new_names,
            )

            newly_assigned_ids = new_employee_ids - old_employee_ids
            newly_assigned = task.assigned_employees.filter(
                id__in=newly_assigned_ids
            )

            create_notification(
                task,
                LogisticsNotification.NotificationType.ASSIGNMENT,
                'Nouvelle tÃ¢che assignÃ©e',
                f'La tÃ¢che Â« {task.title} Â» vous a Ã©tÃ© assignÃ©e.',
                employees=newly_assigned,
            )

        if changes:
            create_notification(
                task,
                LogisticsNotification.NotificationType.MODIFICATION,
                'TÃ¢che modifiÃ©e',
                (
                    f'La tÃ¢che Â« {task.title} Â» a Ã©tÃ© modifiÃ©e : '
                    f'{"; ".join(changes)}.'
                ),
            )

    def perform_destroy(self, instance):
        vehicle = instance.vehicle
        employees = list(instance.assigned_employees.all())
        super().perform_destroy(instance)
        sync_shipment_resources(vehicle, *employees)

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        rows = [
            [
                task.title,
                task.get_priority_display(),
                task.get_status_display(),
                task.due_date or '',
                task.get_assigned_role_display(),
                ', '.join(
                    employee.full_name
                    for employee in task.assigned_employees.all()
                ),
                task.description,
            ]
            for task in queryset
        ]

        return csv_response(
            'taches-logistiques.csv',
            [
                'Titre',
                'PrioritÃ©',
                'Statut',
                'Date Ã©chÃ©ance',
                'RÃ´le',
                'EmployÃ©s',
                'Description',
            ],
            rows,
        )

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        rows = [
            [
                task.id,
                task.title,
                task.get_priority_display(),
                task.get_status_display(),
                task.due_date or '',
                task.get_assigned_role_display(),
                ', '.join(
                    f'{employee.full_name} ({employee.employee_number})'
                    for employee in task.assigned_employees.all()
                ),
                task.description,
                task.created_at.replace(tzinfo=None),
            ]
            for task in queryset
        ]
        return excel_response(
            'taches_logistiques.xlsx',
            'Tâches',
            [
                'ID',
                'Titre',
                'Priorité',
                'Statut',
                'Échéance',
                'Rôle',
                'Employés',
                'Description',
                'Créée le',
            ],
            rows,
        )

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        rows = [
            [
                task.title,
                task.get_priority_display(),
                task.get_status_display(),
                task.due_date or '',
                task.get_assigned_role_display(),
                ', '.join(
                    employee.full_name
                    for employee in task.assigned_employees.all()
                ),
            ]
            for task in queryset
        ]

        return simple_pdf_response(
            'taches-logistiques.pdf',
            'TÃ¢ches logistiques',
            [
                'Titre',
                'PrioritÃ©',
                'Statut',
                'Ã‰chÃ©ance',
                'RÃ´le',
                'EmployÃ©s',
            ],
            rows,
        )

    @action(detail=False, methods=['post'], url_path='import-csv')
    def import_csv(self, request):
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            return Response(
                {'file': 'Fichier obligatoire.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            decoded = uploaded_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
        except Exception:
            return Response(
                {'file': 'Fichier CSV invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        errors = []

        for row_number, row in enumerate(reader, start=2):
            title = (
                row.get('Titre')
                or row.get('title')
                or ''
            ).strip()

            if not title:
                errors.append(
                    f'Ligne {row_number} : titre obligatoire.'
                )
                continue

            priority = (
                row.get('PrioritÃ©')
                or row.get('priority')
                or LogisticsTask.Priority.MEDIUM
            )
            task_status = (
                row.get('Statut')
                or row.get('status')
                or LogisticsTask.TaskStatus.TODO
            )
            assigned_role = (
                row.get('RÃ´le')
                or row.get('role')
                or LogisticsTask.AssignedRole.WAREHOUSE_OPERATOR
            )

            valid_priorities = {
                choice[0]
                for choice in LogisticsTask.Priority.choices
            }
            valid_statuses = {
                choice[0]
                for choice in LogisticsTask.TaskStatus.choices
            }
            valid_roles = {
                choice[0]
                for choice in LogisticsTask.AssignedRole.choices
            }

            if priority not in valid_priorities:
                priority = LogisticsTask.Priority.MEDIUM

            if task_status not in valid_statuses:
                task_status = LogisticsTask.TaskStatus.TODO

            if assigned_role not in valid_roles:
                assigned_role = (
                    LogisticsTask.AssignedRole.WAREHOUSE_OPERATOR
                )

            task = LogisticsTask.objects.create(
                title=title,
                description=(
                    row.get('Description')
                    or row.get('description')
                    or ''
                ),
                priority=priority,
                status=task_status,
                assigned_role=assigned_role,
                other_role_description=(
                    row.get('Description rÃ´le')
                    or ''
                ),
                created_by=request.user,
            )

            create_history(
                task,
                request.user,
                'Import CSV',
                new_value='TÃ¢che importÃ©e',
            )
            created += 1

        return Response({
            'created': created,
            'errors': errors,
        })
    # -------------------------------------------------------------------
# Task Comments
# -------------------------------------------------------------------
class LogisticsTaskCommentViewSet(viewsets.ModelViewSet):
    serializer_class = LogisticsTaskCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = LogisticsTaskComment.objects.select_related(
            'task',
            'author',
        )

        task_id = self.request.query_params.get('task')
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset

    def perform_create(self, serializer):
        comment = serializer.save(author=self.request.user)

        create_history(
            comment.task,
            self.request.user,
            'Commentaire ajouté',
            '',
            comment.comment,
        )

        create_notification(
            comment.task,
            LogisticsNotification.NotificationType.COMMENT,
            'Nouveau commentaire',
            f'Un commentaire a Ã©tÃ© ajoutÃ© Ã  la tÃ¢che Â« {comment.task.title} Â».',
        )


# -------------------------------------------------------------------
# Task Attachments
# -------------------------------------------------------------------
class LogisticsTaskAttachmentViewSet(viewsets.ModelViewSet):
    serializer_class = LogisticsTaskAttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = LogisticsTaskAttachment.objects.select_related(
            'task',
            'uploaded_by',
        )

        task_id = self.request.query_params.get('task')
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset

    def perform_create(self, serializer):
        attachment = serializer.save(uploaded_by=self.request.user)

        create_history(
            attachment.task,
            self.request.user,
            'Pièce jointe ajoutée',
            '',
            attachment.file.name,
        )

        create_notification(
            attachment.task,
            LogisticsNotification.NotificationType.ATTACHMENT,
            'Nouvelle piÃ¨ce jointe',
            f'Une piÃ¨ce jointe a Ã©tÃ© ajoutÃ©e Ã  la tÃ¢che Â« {attachment.task.title} Â».',
        )


# -------------------------------------------------------------------
# Task History
# -------------------------------------------------------------------
class LogisticsTaskHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = LogisticsTaskHistorySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = LogisticsTaskHistory.objects.select_related(
            'task',
            'actor',
        )

        task_id = self.request.query_params.get('task')
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset


# -------------------------------------------------------------------
# Notifications
# -------------------------------------------------------------------
class LogisticsNotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = LogisticsNotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        notify_deadlines()

        queryset = LogisticsNotification.objects.select_related(
            'task',
            'employee',
            'recipient',
        ).filter(recipient=self.request.user)

        unread = self.request.query_params.get('unread')
        if unread in ('1', 'true', 'True'):
            queryset = queryset.filter(is_read=False)

        task_id = self.request.query_params.get('task')
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset

    @action(detail=True, methods=['post'], url_path='mark-read')
    def mark_read(self, request, pk=None):
        notification = self.get_object()

        if not notification.is_read:
            notification.is_read = True
            notification.read_at = timezone.now()
            notification.save(update_fields=['is_read', 'read_at'])

        return Response(self.get_serializer(notification).data)

    @action(detail=False, methods=['post'], url_path='mark-all-read')
    def mark_all_read(self, request):
        updated = self.get_queryset().filter(is_read=False).update(
            is_read=True,
            read_at=timezone.now(),
        )

        return Response({
            'detail': 'Toutes les notifications ont Ã©tÃ© marquÃ©es comme lues.',
            'updated': updated,
        })


class LogisticsReportJournalViewSet(viewsets.ModelViewSet):
    serializer_class = LogisticsReportSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = LogisticsReport.objects.select_related('created_by')
        report_type = self.request.query_params.get('report_type')
        search = self.request.query_params.get('search')
        if report_type:
            queryset = queryset.filter(report_type=report_type)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(content__icontains=search)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        rows = [
            [
                report.report_date,
                report.get_report_type_display(),
                report.title,
                report.content,
                user_display_name(report.created_by),
                report.created_at.replace(tzinfo=None),
            ]
            for report in self.get_queryset()
        ]
        return excel_response(
            'journal_rapports_logistique.xlsx',
            'Rapports',
            ['Date', 'Type', 'Titre', 'Contenu', 'Auteur', 'Créé le'],
            rows,
        )


# -------------------------------------------------------------------
# Reports
# -------------------------------------------------------------------
class LogisticsReportViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def _requested_format(self, request):
        return request.query_params.get('format', 'json').lower()

    def _employee_performance_data(self):
        from hr.models import Employee

        today = date.today()
        results = []

        employees = Employee.objects.select_related('department').order_by(
            'department__name',
            'last_name',
            'first_name',
        )

        for employee in employees:
            tasks = employee.logistics_tasks.all()
            total = tasks.count()

            if total == 0:
                continue

            completed = tasks.filter(
                status=LogisticsTask.TaskStatus.DONE
            ).count()

            cancelled = tasks.filter(
                status=LogisticsTask.TaskStatus.CANCELLED
            ).count()

            late = tasks.filter(
                due_date__lt=today,
            ).exclude(
                status__in=[
                    LogisticsTask.TaskStatus.DONE,
                    LogisticsTask.TaskStatus.CANCELLED,
                ]
            ).count()

            in_progress = tasks.filter(
                status=LogisticsTask.TaskStatus.IN_PROGRESS
            ).count()

            completion_rate = round((completed / total) * 100, 2) if total else 0

            results.append({
                'employee_id': employee.id,
                'employee_number': employee.employee_number,
                'employee_name': employee.full_name,
                'department': employee.department.name if employee.department else 'Sans dÃ©partement',
                'total_tasks': total,
                'completed_tasks': completed,
                'in_progress_tasks': in_progress,
                'late_tasks': late,
                'cancelled_tasks': cancelled,
                'completion_rate': completion_rate,
            })

        return results

    def _late_tasks_data(self):
        today = date.today()

        tasks = task_queryset().filter(
            due_date__lt=today,
        ).exclude(
            status__in=[
                LogisticsTask.TaskStatus.DONE,
                LogisticsTask.TaskStatus.CANCELLED,
            ]
        ).order_by('due_date', '-priority')

        results = []

        for task in tasks:
            employees = list(task.assigned_employees.all())

            results.append({
                'task_id': task.id,
                'title': task.title,
                'priority': task.get_priority_display(),
                'status': task.get_status_display(),
                'due_date': task.due_date,
                'days_late': (today - task.due_date).days,
                'assigned_employees': ', '.join(
                    employee.full_name for employee in employees
                ) or 'Non assignÃ©e',
                'departments': ', '.join(sorted({
                    employee.department.name
                    for employee in employees
                    if employee.department
                })) or 'Sans dÃ©partement',
            })

        return results

    def _workload_data(self):
        from hr.models import Employee

        results = []

        employees = Employee.objects.select_related('department').order_by(
            'department__name',
            'last_name',
            'first_name',
        )

        for employee in employees:
            tasks = employee.logistics_tasks.all()

            todo = tasks.filter(
                status=LogisticsTask.TaskStatus.TODO
            ).count()

            in_progress = tasks.filter(
                status=LogisticsTask.TaskStatus.IN_PROGRESS
            ).count()

            waiting = tasks.filter(
                status=LogisticsTask.TaskStatus.WAITING
            ).count()

            completed = tasks.filter(
                status=LogisticsTask.TaskStatus.DONE
            ).count()

            active = todo + in_progress + waiting
            total = tasks.count()

            if total == 0:
                continue

            results.append({
                'employee_id': employee.id,
                'employee_number': employee.employee_number,
                'employee_name': employee.full_name,
                'department': employee.department.name if employee.department else 'Sans dÃ©partement',
                'todo_tasks': todo,
                'in_progress_tasks': in_progress,
                'waiting_tasks': waiting,
                'active_tasks': active,
                'completed_tasks': completed,
                'total_tasks': total,
            })

        results.sort(
            key=lambda item: (
                -item['active_tasks'],
                item['department'],
                item['employee_name'],
            )
        )

        return results

    @action(detail=False, methods=['get'], url_path='employee-performance')
    def employee_performance(self, request):
        data = self._employee_performance_data()
        output_format = self._requested_format(request)

        headers = [
            'Matricule',
            'EmployÃ©',
            'DÃ©partement',
            'Total',
            'TerminÃ©es',
            'En cours',
            'En retard',
            'AnnulÃ©es',
            'Taux de rÃ©alisation',
        ]

        rows = [
            [
                item['employee_number'],
                item['employee_name'],
                item['department'],
                item['total_tasks'],
                item['completed_tasks'],
                item['in_progress_tasks'],
                item['late_tasks'],
                item['cancelled_tasks'],
                f"{item['completion_rate']} %",
            ]
            for item in data
        ]

        if output_format == 'csv':
            return csv_response(
                'performance_employes_logistique.csv',
                headers,
                rows,
            )

        if output_format == 'pdf':
            return simple_pdf_response(
                'performance_employes_logistique.pdf',
                'Rapport de performance des employÃ©s',
                headers,
                rows,
            )
        if output_format in ('xlsx', 'excel'):
            return excel_response(
                'performance_employes_logistique.xlsx',
                'Performance',
                headers,
                rows,
            )

        return Response(data)

    @action(detail=False, methods=['get'], url_path='late-tasks')
    def late_tasks(self, request):
        data = self._late_tasks_data()
        output_format = self._requested_format(request)

        headers = [
            'ID',
            'TÃ¢che',
            'PrioritÃ©',
            'Statut',
            'Ã‰chÃ©ance',
            'Jours de retard',
            'EmployÃ©s assignÃ©s',
            'DÃ©partements',
        ]

        rows = [
            [
                item['task_id'],
                item['title'],
                item['priority'],
                item['status'],
                item['due_date'],
                item['days_late'],
                item['assigned_employees'],
                item['departments'],
            ]
            for item in data
        ]

        if output_format == 'csv':
            return csv_response(
                'taches_logistiques_en_retard.csv',
                headers,
                rows,
            )

        if output_format == 'pdf':
            return simple_pdf_response(
                'taches_logistiques_en_retard.pdf',
                'Rapport des tÃ¢ches en retard',
                headers,
                rows,
            )
        if output_format in ('xlsx', 'excel'):
            return excel_response(
                'taches_logistiques_en_retard.xlsx',
                'Tâches en retard',
                headers,
                rows,
            )

        return Response(data)

    @action(detail=False, methods=['get'], url_path='workload')
    def workload(self, request):
        data = self._workload_data()
        output_format = self._requested_format(request)

        headers = [
            'Matricule',
            'EmployÃ©',
            'DÃ©partement',
            'Ã€ faire',
            'En cours',
            'En attente',
            'Charge active',
            'TerminÃ©es',
            'Total',
        ]

        rows = [
            [
                item['employee_number'],
                item['employee_name'],
                item['department'],
                item['todo_tasks'],
                item['in_progress_tasks'],
                item['waiting_tasks'],
                item['active_tasks'],
                item['completed_tasks'],
                item['total_tasks'],
            ]
            for item in data
        ]

        if output_format == 'csv':
            return csv_response(
                'charge_travail_logistique.csv',
                headers,
                rows,
            )

        if output_format == 'pdf':
            return simple_pdf_response(
                'charge_travail_logistique.pdf',
                'Rapport de charge de travail',
                headers,
                rows,
            )
        if output_format in ('xlsx', 'excel'):
            return excel_response(
                'charge_travail_logistique.xlsx',
                'Charge de travail',
                headers,
                rows,
            )

        return Response(data)


# -------------------------------------------------------------------
# Logistics Dashboard
# -------------------------------------------------------------------
class LogisticsDashboardView(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='')
    def dashboard(self, request):
        today = date.today()
        notify_deadlines()

        pending_shipments = Shipment.objects.filter(
            status=Shipment.ShipmentStatus.PENDING
        ).count()

        active_shipments = Shipment.objects.filter(
            status__in=[
                Shipment.ShipmentStatus.PREPARATION,
                Shipment.ShipmentStatus.SHIPPED,
                Shipment.ShipmentStatus.IN_DELIVERY,
            ]
        ).count()

        delivered_today = Shipment.objects.filter(
            status=Shipment.ShipmentStatus.DELIVERED,
            shipment_date=today,
        ).count()

        available_vehicles = Vehicle.objects.filter(
            status=Vehicle.VehicleStatus.AVAILABLE,
        ).count()

        unavailable_vehicles = Vehicle.objects.exclude(
            status=Vehicle.VehicleStatus.AVAILABLE,
        ).count()

        active_transfers = WarehouseTransfer.objects.filter(
            status__in=[
                WarehouseTransfer.TransferStatus.PENDING_APPROVAL,
                WarehouseTransfer.TransferStatus.APPROVED,
                WarehouseTransfer.TransferStatus.IN_TRANSIT,
            ]
        ).count()

        pending_tasks = LogisticsTask.objects.filter(
            status=LogisticsTask.TaskStatus.TODO
        ).count()

        active_tasks = LogisticsTask.objects.filter(
            status__in=[
                LogisticsTask.TaskStatus.IN_PROGRESS,
                LogisticsTask.TaskStatus.WAITING,
            ]
        ).count()

        completed_tasks = LogisticsTask.objects.filter(
            status=LogisticsTask.TaskStatus.DONE
        ).count()

        late_tasks = LogisticsTask.objects.filter(
            due_date__lt=today,
        ).exclude(
            status__in=[
                LogisticsTask.TaskStatus.DONE,
                LogisticsTask.TaskStatus.CANCELLED,
            ]
        ).count()

        unread_notifications = LogisticsNotification.objects.filter(
            recipient=request.user,
            is_read=False,
        ).count()

        return Response({
            'pending_shipments': pending_shipments,
            'active_shipments': active_shipments,
            'delivered_today': delivered_today,
            'available_vehicles': available_vehicles,
            'unavailable_vehicles': unavailable_vehicles,
            'active_transfers': active_transfers,
            'pending_tasks': pending_tasks,
            'active_tasks': active_tasks,
            'completed_tasks': completed_tasks,
            'late_tasks': late_tasks,
            'unread_notifications': unread_notifications,
        })
